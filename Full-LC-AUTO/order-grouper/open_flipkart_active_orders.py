from __future__ import annotations

import argparse
import html
import os
import re
import time
import tkinter as tk
from dataclasses import dataclass
from pathlib import Path
from time import sleep
from tkinter import messagebox, ttk

import pyautogui
from openpyxl import load_workbook
from pynput.mouse import Button as PynputButton
from pynput.mouse import Controller as MouseController
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

TARGET_URL = (
    "https://seller.flipkart.com/index.html#dashboard/active-orders?query=%7B%22activeShipmentTile%22%3A%22pendingToPack%22%7D"
)

FIREFOX_PROFILES = {
    "prabhu": Path(r"C:\Users\ESHAAN\Documents\Firefox-Profiles\0xe7h0bx.prabhu"),
}

DEFAULT_PROFILE_NAME = "prabhu"
VARIANT_SELECTION_WORKBOOK = Path(
    r"C:\work-mom\Code-Tools\Full-LC-AUTO\order-grouper\Variant Selection.xlsx"
)
DONE_SELECTION_IMAGE_DIRECTORY = Path(
    r"C:\work-mom\Code-Tools\Full-LC-AUTO\order-grouper\done-selection-images"
)


def log_event(stage: str, message: str) -> None:
    timestamp = time.strftime("%H:%M:%S")
    print(f"[{timestamp}] [{stage}] {message}")


@dataclass(slots=True)
class VariantSelectionRow:
    jeans_type: str
    size: str
    phrases: list[str]


@dataclass(slots=True)
class VariantSelectionChoice:
    jeans_type: str
    size: str


def load_variant_selection_options(workbook_path: Path) -> dict[str, list[str]]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Variant selection workbook was not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in worksheet[1]]
    options: dict[str, list[str]] = {}

    for row_index in range(2, worksheet.max_row + 1):
        row_values = {
            str(headers[column_index - 1]): worksheet.cell(row_index, column_index).value
            for column_index in range(1, worksheet.max_column + 1)
            if headers[column_index - 1]
        }
        row_type = str(row_values.get("JEANS-TYPE", "")).strip()
        row_size = str(row_values.get("SIZE", "")).strip()
        if not row_type or not row_size:
            continue
        options.setdefault(row_type, [])
        if row_size not in options[row_type]:
            options[row_type].append(row_size)

    if not options:
        raise ValueError("No JEANS-TYPE / SIZE options were found in the variant selection workbook.")

    return options


def resolve_profile_path(profile_name: str) -> Path:
    profile_name = profile_name.lower()
    if profile_name not in FIREFOX_PROFILES:
        available = ", ".join(sorted(FIREFOX_PROFILES))
        raise ValueError(
            f"Unknown Firefox profile '{profile_name}'. Choose one of: {available}."
        )
    profile_path = FIREFOX_PROFILES[profile_name]
    if not profile_path.exists():
        raise FileNotFoundError(
            f"Firefox profile directory was not found: {profile_path}"
        )
    return profile_path


def build_firefox_driver(profile_name: str = DEFAULT_PROFILE_NAME) -> webdriver.Firefox:
    options = FirefoxOptions()
    firefox_binary = os.getenv("FIREFOX_BINARY")
    geckodriver_path = os.getenv("GECKODRIVER_PATH")

    if firefox_binary:
        options.binary_location = firefox_binary

    profile_path = resolve_profile_path(profile_name)
    options.add_argument("-profile")
    options.add_argument(str(profile_path))

    if geckodriver_path:
        service = FirefoxService(executable_path=geckodriver_path)
        return webdriver.Firefox(service=service, options=options)

    return webdriver.Firefox(options=options)


def prompt_for_variant_selection() -> tuple[str, str]:
    options = load_variant_selection_options(VARIANT_SELECTION_WORKBOOK)
    selection = open_variant_selection_gui(options)
    return selection.jeans_type, selection.size


def open_variant_selection_gui(options: dict[str, list[str]]) -> VariantSelectionChoice:
    jeans_types = list(options.keys())
    if not jeans_types:
        raise ValueError("No JEANS-TYPE values are available for selection.")

    root = tk.Tk()
    root.title("Variant Row Selector")
    root.attributes("-topmost", True)
    root.resizable(False, False)

    selection_result: dict[str, str] = {}

    jeans_type_var = tk.StringVar(value=jeans_types[0])
    size_options = options[jeans_types[0]]
    size_var = tk.StringVar(value=size_options[0] if size_options else "")
    status_var = tk.StringVar(value="Choose JEANS-TYPE and SIZE, then click Start.")

    def refresh_sizes(*_args) -> None:
        selected_type = jeans_type_var.get()
        available_sizes = options[selected_type]
        size_combo["values"] = available_sizes
        if available_sizes:
            size_var.set(available_sizes[0])
        else:
            size_var.set("")

    def submit() -> None:
        selected_type = jeans_type_var.get().strip()
        selected_size = size_var.get().strip()
        if not selected_type or not selected_size:
            status_var.set("Please choose both JEANS-TYPE and SIZE.")
            return
        selection_result["jeans_type"] = selected_type
        selection_result["size"] = selected_size
        root.destroy()

    def cancel() -> None:
        root.destroy()

    container = ttk.Frame(root, padding=16)
    container.grid(row=0, column=0, sticky="nsew")

    ttk.Label(container, text="JEANS-TYPE").grid(row=0, column=0, sticky="w")
    type_combo = ttk.Combobox(
        container,
        textvariable=jeans_type_var,
        values=jeans_types,
        state="readonly",
        width=24,
    )
    type_combo.grid(row=1, column=0, sticky="ew", pady=(4, 12))

    ttk.Label(container, text="SIZE").grid(row=2, column=0, sticky="w")
    size_combo = ttk.Combobox(
        container,
        textvariable=size_var,
        values=size_options,
        state="readonly",
        width=24,
    )
    size_combo.grid(row=3, column=0, sticky="ew", pady=(4, 12))

    ttk.Label(container, textvariable=status_var, wraplength=260).grid(
        row=4, column=0, sticky="w", pady=(0, 12)
    )

    button_row = ttk.Frame(container)
    button_row.grid(row=5, column=0, sticky="ew")
    ttk.Button(button_row, text="Start", command=submit).grid(row=0, column=0, padx=(0, 8))
    ttk.Button(button_row, text="Cancel", command=cancel).grid(row=0, column=1)

    type_combo.bind("<<ComboboxSelected>>", refresh_sizes)
    root.protocol("WM_DELETE_WINDOW", cancel)
    root.mainloop()

    if "jeans_type" not in selection_result or "size" not in selection_result:
        raise ValueError("Variant row selection was cancelled.")

    return VariantSelectionChoice(
        jeans_type=selection_result["jeans_type"],
        size=selection_result["size"],
    )


def load_variant_selection_row(workbook_path: Path, jeans_type: str, size: str) -> VariantSelectionRow:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Variant selection workbook was not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in worksheet[1]]
    normalized_target_type = jeans_type.strip().lower()
    normalized_target_size = size.strip().lower()

    for row_index in range(2, worksheet.max_row + 1):
        row_values = {
            str(headers[column_index - 1]): worksheet.cell(row_index, column_index).value
            for column_index in range(1, worksheet.max_column + 1)
            if headers[column_index - 1]
        }
        row_type = str(row_values.get("JEANS-TYPE", "")).strip()
        row_size = str(row_values.get("SIZE", "")).strip()
        if row_type.lower() != normalized_target_type or row_size.lower() != normalized_target_size:
            continue

        phrases = []
        for key in ("Phrase-1", "Phrase-2"):
            raw_phrase = row_values.get(key)
            if raw_phrase is None:
                continue
            phrase = str(raw_phrase).strip()
            if phrase:
                phrases.append(phrase)

        return VariantSelectionRow(
            jeans_type=row_type,
            size=row_size,
            phrases=phrases,
        )

    raise ValueError(
        f"No Variant Selection row found for JEANS-TYPE='{jeans_type}' and SIZE='{size}'."
    )


def open_target_page(driver: webdriver.Firefox, url: str) -> None:
    driver.maximize_window()
    sleep(1.0)
    driver.get(url)
    sleep(2.0)
    current_url = ""
    try:
        current_url = driver.current_url
    except WebDriverException:
        current_url = ""

    if url not in current_url:
        log_event("NAV", f"Chrome did not land on the target URL yet. Current URL: {current_url or '<unknown>'}")
        log_event("NAV", "Retrying direct navigation to the target URL...")
        driver.get(url)
        sleep(2.0)
        try:
            current_url = driver.current_url
        except WebDriverException:
            current_url = ""

    if url not in current_url:
        log_event("NAV", "Direct navigation still did not stick. Forcing location change in the active tab...")
        driver.execute_script("window.location.href = arguments[0];", url)
        sleep(2.0)
        try:
            current_url = driver.current_url
        except WebDriverException:
            current_url = ""

    log_event("NAV", f"Opened target page. Current URL: {current_url or url}")


def wait_for_skip_button(driver: webdriver.Firefox, timeout: int = 30):
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable(
            (By.XPATH, "//button[.//span[text()='Skip for Later']]")
        )
    )


def wait_for_pagination_area(driver: webdriver.Firefox, timeout: int = 30):
    xpaths = [
        "//div[@data-testid='pagination-container']",
        "//div[contains(@class,'PaginationSelectContainer')]",
        "//div[contains(@class,'PageSelectorContainer')]",
    ]
    deadline = time.time() + timeout

    while time.time() < deadline:
        for xpath in xpaths:
            try:
                candidates = driver.find_elements(By.XPATH, xpath)
            except WebDriverException:
                continue

            for candidate in candidates:
                try:
                    if candidate.is_displayed():
                        log_event("PAGINATION", "Detected pagination area near the bottom of the page.")
                        return candidate
                except WebDriverException:
                    continue
        time.sleep(0.2)

    raise TimeoutException("Could not find the pagination area that contains the page-size selector.")


def scroll_to_pagination_area(driver: webdriver.Firefox, timeout: int = 30) -> None:
    pagination_area = wait_for_pagination_area(driver, timeout=timeout)
    log_event("SCROLL", "Scrolling to the pagination area at the bottom of the page...")
    driver.execute_script(
        "arguments[0].scrollIntoView({block: 'end', inline: 'nearest'});",
        pagination_area,
    )
    sleep(0.8)
    for step_index in range(1, 5):
        driver.execute_script("window.scrollBy(0, 220);")
        log_event("SCROLL", f"Bottom scroll step {step_index}/4 completed.")
        sleep(0.45)
    log_event("SCROLL", "Reached the page-size selector area.")


def wait_for_items_per_page_combobox(
    driver: webdriver.Firefox,
    timeout: int = 30,
):
    xpaths = [
        "//button[@role='combobox' and .//div[normalize-space(text())='Items / page']]",
        "//div[contains(@class,'PageSelectorContainer')]//button[@role='combobox' and .//div[normalize-space(text())='Items / page']]",
    ]
    deadline = time.time() + timeout

    while time.time() < deadline:
        for xpath in xpaths:
            try:
                candidates = driver.find_elements(By.XPATH, xpath)
            except WebDriverException:
                continue

            for candidate in candidates:
                try:
                    if candidate.is_displayed():
                        log_event("FORM", "Detected the 'Items / page' combobox.")
                        return candidate
                except WebDriverException:
                    continue
        time.sleep(0.15)

    raise TimeoutException("Could not find the 'Items / page' combobox.")


def click_element_with_gui(
    driver: webdriver.Firefox,
    element,
    label: str = "element",
    pre_click_sleep: float = 1.0,
    move_duration: float = 0.35,
    post_click_sleep: float = 0.5,
) -> None:
    driver.execute_script(
        "arguments[0].scrollIntoView({block: 'center', inline: 'center'});",
        element,
    )
    sleep(pre_click_sleep)

    rect = driver.execute_script(
        "const r = arguments[0].getBoundingClientRect();"
        "return {left: r.left, top: r.top, width: r.width, height: r.height, "
        "outerWidth: window.outerWidth, innerWidth: window.innerWidth, "
        "outerHeight: window.outerHeight, innerHeight: window.innerHeight, "
        "screenX: window.screenX, screenY: window.screenY};",
        element,
    )

    left_border = (rect["outerWidth"] - rect["innerWidth"]) / 2
    top_border = rect["outerHeight"] - rect["innerHeight"]

    screen_x = rect["screenX"] + rect["left"] + left_border + rect["width"] / 2
    screen_y = rect["screenY"] + rect["top"] + top_border + rect["height"] / 2

    screen_x = int(round(screen_x))
    screen_y = int(round(screen_y))
    log_event("MOUSE", f"Moving to {label} at screen coordinates {screen_x},{screen_y}")
    pyautogui.moveTo(screen_x, screen_y, duration=move_duration)
    mouse = MouseController()
    mouse.click(PynputButton.left, 1)
    log_event("MOUSE", f"Clicked {label}.")
    sleep(post_click_sleep)


def find_page_size_option_label(
    driver: webdriver.Firefox,
    option_text: str = "100",
    timeout: int = 30,
):
    normalized_target = option_text.strip()
    xpaths = [
        (
            "//div[contains(@class,'PageSelectorContainer')]"
            f"//div[@role='radiogroup']//label[.//span[normalize-space(text())='{normalized_target}']]"
        ),
        (
            "//div[contains(@class,'PageSelectorContainer')]"
            f"//div[@role='radiogroup']//span[normalize-space(text())='{normalized_target}']/ancestor::label[1]"
        ),
        f"//div[@role='radiogroup']//label[.//span[normalize-space(text())='{normalized_target}']]",
        f"//label[.//span[normalize-space(text())='{normalized_target}']]",
    ]
    deadline = time.time() + timeout

    while time.time() < deadline:
        for xpath in xpaths:
            try:
                candidates = driver.find_elements(By.XPATH, xpath)
            except WebDriverException:
                continue

            for candidate in candidates:
                try:
                    if candidate.is_displayed():
                        log_event("FORM", f"Detected page-size option row '{normalized_target}'.")
                        return candidate
                except WebDriverException:
                    continue
        time.sleep(0.15)

    all_texts = []
    for xpath in xpaths:
        try:
            candidates = driver.find_elements(By.XPATH, xpath)
        except WebDriverException:
            continue
        for candidate in candidates:
            try:
                text = candidate.text.strip()
            except WebDriverException:
                continue
            if text:
                all_texts.append(text)
    raise TimeoutException(
        f"Could not find page-size option '{option_text}'. "
        f"Searched DOM xpaths and saw options: {all_texts}"
    )


def select_page_size_option(driver: webdriver.Firefox, option_text: str = "100") -> None:
    log_event("FORM", f"Starting page-size selection flow for option '{option_text}'.")
    scroll_to_pagination_area(driver, timeout=30)
    items_per_page_combobox = wait_for_items_per_page_combobox(driver, timeout=30)
    click_element_with_gui(
        driver,
        items_per_page_combobox,
        label="Items / page combobox",
    )
    log_event("FORM", "Opened the Items / page dropdown.")
    sleep(0.75)
    option_label = find_page_size_option_label(driver, option_text=option_text, timeout=30)
    click_element_with_gui(driver, option_label, label=f"page size option {option_text}")
    log_event("FORM", f"Selected page-size option '{option_text}'.")


def wait_for_page_size_value(driver: webdriver.Firefox, option_text: str = "100", timeout: int = 30) -> None:
    normalized_target = option_text.strip()
    xpaths = [
        (
            "//div[contains(@class,'PageSelectorContainer')]"
            f"//button[@role='combobox'][.//div[normalize-space(text())='Items / page']]"
            f"//*[self::div or self::span][normalize-space(text())='{normalized_target}' or @title='{normalized_target}' or @aria-label='{normalized_target}']"
        ),
        (
            "//button[@role='combobox'][.//div[normalize-space(text())='Items / page']]"
            f"//*[self::div or self::span][normalize-space(text())='{normalized_target}' or @title='{normalized_target}' or @aria-label='{normalized_target}']"
        ),
    ]
    deadline = time.time() + timeout

    while time.time() < deadline:
        for xpath in xpaths:
            try:
                candidates = driver.find_elements(By.XPATH, xpath)
            except WebDriverException:
                continue

            for candidate in candidates:
                try:
                    if candidate.is_displayed():
                        log_event("VERIFY", f"Confirmed page-size selector now shows '{normalized_target}'.")
                        return
                except WebDriverException:
                    continue
        time.sleep(0.2)

    raise TimeoutException(f"Page-size selector did not update to '{normalized_target}'.")


def wait_for_order_rows(driver: webdriver.Firefox, timeout: int = 30):
    xpath = "//tr[starts-with(@data-testid, 'grid-component-row-')]"
    deadline = time.time() + timeout

    while time.time() < deadline:
        try:
            rows = driver.find_elements(By.XPATH, xpath)
        except WebDriverException:
            rows = []

        visible_rows = []
        for row in rows:
            try:
                if row.is_displayed():
                    visible_rows.append(row)
            except WebDriverException:
                continue

        if visible_rows:
            log_event("ROWS", f"Detected {len(visible_rows)} visible order row(s).")
            return visible_rows
        time.sleep(0.2)

    raise TimeoutException("Could not find any visible order rows on the page.")


def wait_for_order_rows_to_settle(
    driver: webdriver.Firefox,
    timeout: int = 45,
    settle_duration: float = 2.0,
) -> list:
    log_event("VERIFY", "Waiting for visible order rows after page-size change...")
    deadline = time.time() + timeout
    last_count = None
    last_change_time = time.time()
    last_rows = []

    while time.time() < deadline:
        rows = wait_for_order_rows(driver, timeout=10)
        visible_count = len(rows)
        if visible_count != last_count:
            log_event("VERIFY", f"Visible row count changed to {visible_count}.")
            last_count = visible_count
            last_change_time = time.time()
            last_rows = rows
        elif time.time() - last_change_time >= settle_duration:
            log_event(
                "VERIFY",
                f"Visible row count has stopped changing at {visible_count}. Starting fast row selection.",
            )
            return rows
        else:
            last_rows = rows
        time.sleep(0.25)

    if last_rows:
        log_event(
            "VERIFY",
            (
                f"Row count did not fully settle before timeout. "
                f"Continuing with the latest visible count: {len(last_rows)}."
            ),
        )
        return last_rows
    raise TimeoutException("Order rows did not load after changing page size.")


def extract_sku_and_fsn_from_row(row) -> tuple[str, str, str]:
    row_address = row.get_attribute("data-testid") or "<missing-data-testid>"
    row_text = " ".join(part.strip() for part in row.text.splitlines() if part.strip())
    match = re.search(r"SKU ID:\s*(.*?)\s*\|\s*FSN:\s*([A-Z0-9]+)", row_text, re.IGNORECASE)
    if not match:
        return row_address, "", ""
    sku_value = match.group(1).strip()
    fsn_value = match.group(2).strip()
    return row_address, sku_value, fsn_value


def get_row_checkbox_element(row):
    checkbox_xpaths = [
        ".//input[@type='checkbox']",
        ".//span[contains(@class,'SelectionWrapper')]//input[@type='checkbox']",
    ]
    for xpath in checkbox_xpaths:
        try:
            candidates = row.find_elements(By.XPATH, xpath)
        except WebDriverException:
            continue
        for candidate in candidates:
            try:
                if candidate.is_displayed():
                    return candidate
            except WebDriverException:
                continue
        if candidates:
            return candidates[0]
    raise TimeoutException("Could not find a checkbox input inside the matched row.")


def print_visible_row_skus(
    driver: webdriver.Firefox,
    selection_row: VariantSelectionRow,
) -> int:
    rows = wait_for_order_rows_to_settle(driver, timeout=45, settle_duration=2.0)
    normalized_phrases = [phrase.upper() for phrase in selection_row.phrases]
    log_event(
        "FILTER",
        (
            f"Filtering rows for JEANS-TYPE='{selection_row.jeans_type}', "
            f"SIZE='{selection_row.size}', phrases={selection_row.phrases}"
        ),
    )
    matched_count = 0
    log_event("ROWS", "Scanning visible order rows for phrase matches...")
    for row in rows:
        try:
            row_address, sku_value, fsn_value = extract_sku_and_fsn_from_row(row)
        except WebDriverException:
            continue

        if not sku_value:
            continue

        normalized_sku = sku_value.upper()
        matching_phrases = [
            phrase for phrase, normalized_phrase in zip(selection_row.phrases, normalized_phrases)
            if normalized_phrase and normalized_phrase.upper() in normalized_sku
        ]
        if matching_phrases:
            matched_count += 1
            checkbox_element = get_row_checkbox_element(row)
            click_element_with_gui(
                driver,
                checkbox_element,
                label=f"matched row checkbox #{matched_count}",
                pre_click_sleep=0.2,
                move_duration=0.15,
                post_click_sleep=0.15,
            )
    if matched_count == 0:
        log_event("ROWS", "No visible order rows matched the selected phrase filters.")
    else:
        log_event("ROWS", f"Selected {matched_count} matched row checkbox(es).")
    return matched_count


def show_completion_message(selection_row: VariantSelectionRow, matched_count: int) -> None:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    message = (
        "Done selecting matching rows.\n\n"
        f"JEANS-TYPE: {selection_row.jeans_type}\n"
        f"SIZE: {selection_row.size}\n"
        f"Matched rows selected: {matched_count}\n\n"
        "Returning to the selection window..."
    )
    messagebox.showinfo("Selection Complete", message, parent=root)
    root.destroy()


def create_done_selection_image(selection_row: VariantSelectionRow) -> Path:
    DONE_SELECTION_IMAGE_DIRECTORY.mkdir(parents=True, exist_ok=True)
    file_stem = (
        f"done-selecting-{selection_row.jeans_type}-{selection_row.size}"
        .replace(" ", "-")
        .replace("/", "-")
        .replace("\\", "-")
        .lower()
    )
    image_path = DONE_SELECTION_IMAGE_DIRECTORY / f"{file_stem}.svg"
    title_text = f"DONE SELECTING {selection_row.jeans_type.upper()} {selection_row.size}"
    escaped_title = html.escape(title_text)
    svg_markup = f"""<svg xmlns="http://www.w3.org/2000/svg" width="1400" height="800" viewBox="0 0 1400 800">
  <defs>
    <linearGradient id="bg" x1="0%" y1="0%" x2="100%" y2="100%">
      <stop offset="0%" stop-color="#e8f1ff"/>
      <stop offset="100%" stop-color="#cfe2ff"/>
    </linearGradient>
  </defs>
  <rect width="1400" height="800" fill="url(#bg)"/>
  <rect x="70" y="70" width="1260" height="660" rx="36" fill="#ffffff" stroke="#2a55e5" stroke-width="8"/>
  <text x="700" y="350" text-anchor="middle" font-family="Segoe UI, Arial, sans-serif" font-size="88" font-weight="700" fill="#1f3b8f">DONE SELECTING</text>
  <text x="700" y="470" text-anchor="middle" font-family="Segoe UI, Arial, sans-serif" font-size="76" font-weight="700" fill="#2a55e5">{escaped_title.replace('DONE SELECTING ', '')}</text>
  <text x="700" y="585" text-anchor="middle" font-family="Segoe UI, Arial, sans-serif" font-size="34" fill="#3b4b6b">Flipkart order row selection complete</text>
</svg>
"""
    image_path.write_text(svg_markup, encoding="utf-8")
    log_event("DONE", f"Created completion image: {image_path}")
    return image_path


def open_image_in_new_tab(driver: webdriver.Firefox, image_path: Path) -> None:
    image_uri = image_path.resolve().as_uri()
    driver.switch_to.new_window("tab")
    driver.get(image_uri)
    log_event("DONE", f"Opened completion image in a new tab: {image_uri}")


def prepare_orders_view(driver: webdriver.Firefox, url: str) -> None:
    open_target_page(driver, url)
    log_event("NAV", "Page opened successfully. Waiting for the Skip for Later button...")
    try:
        skip_button = wait_for_skip_button(driver, timeout=2)
        click_element_with_gui(driver, skip_button, label="Skip for Later button")
        log_event("NAV", "Skip for Later button clicked.")
    except TimeoutException:
        log_event("NAV", "Skip for Later button did not appear within the timeout window.")
    log_event("FORM", "Waiting for page size dropdown option 100 with no timeout cutoff...")
    select_page_size_option(driver, option_text="100")
    wait_for_page_size_value(driver, option_text="100", timeout=30)
    log_event("DONE", "Page size option 100 selected.")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Open Flipkart seller active orders in Firefox using the Prabhu profile."
    )
    parser.add_argument(
        "--profile",
        default=DEFAULT_PROFILE_NAME,
        help="Firefox profile name to use (default: prabhu).",
    )
    parser.add_argument(
        "--url",
        default=TARGET_URL,
        help="URL to open in Firefox.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    log_event("START", f"Opening Firefox profile '{args.profile}' and navigating to the target Flipkart URL...")
    driver = build_firefox_driver(args.profile)
    try:
        prepare_orders_view(driver, args.url)
        selection_cycle_index = 0
        while True:
            try:
                jeans_type, size = prompt_for_variant_selection()
            except ValueError:
                log_event("DONE", "Selection window was closed. Ending selection loop.")
                break
            if selection_cycle_index > 0:
                log_event(
                    "NAV",
                    "Starting a fresh selection cycle in a new tab for the next variation.",
                )
                driver.switch_to.new_window("tab")
                prepare_orders_view(driver, args.url)
            selection_row = load_variant_selection_row(
                VARIANT_SELECTION_WORKBOOK,
                jeans_type,
                size,
            )
            log_event(
                "DATA",
                (
                    f"Loaded Variant Selection row: JEANS-TYPE={selection_row.jeans_type}, "
                    f"SIZE={selection_row.size}, phrases={selection_row.phrases}"
                ),
            )
            matched_count = print_visible_row_skus(driver, selection_row)
            show_completion_message(selection_row, matched_count)
            # Completion image generation / tab opening retained but disabled for now.
            # completion_image_path = create_done_selection_image(selection_row)
            # open_image_in_new_tab(driver, completion_image_path)
            log_event("DONE", "Done selecting matching rows. Returning to selection window.")
            selection_cycle_index += 1
        log_event("DONE", "Browser remains open.")
    finally:
        pass


if __name__ == "__main__":
    main()
