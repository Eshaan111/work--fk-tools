from __future__ import annotations

import json
import shutil
import tkinter as tk
from copy import deepcopy
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
CONFIG = ROOT / "config.json"
DEMO = ROOT / "config-setup-demo.json"
ENV = ROOT / "setup-runtime.env"
REPORT = ROOT / "setup-diagnostics.txt"
DEFAULT_FLOW_STEPS = ["additional_description", "product_description", "price_stock_shipping", "images"]


def load_config() -> dict[str, object]:
    return json.loads(CONFIG.read_text(encoding="utf-8"))


def norm(value: str) -> str:
    value = value.strip().strip('"')
    return str(Path(value).expanduser()) if value else ""


def slug(value: str) -> str:
    return "_".join(part.lower() for part in value.replace("-", " ").split() if part)


def title_token(value: str) -> str:
    parts = [part for part in value.replace("_", " ").replace("-", " ").split() if part]
    return " ".join(part[:1].upper() + part[1:] for part in parts) or "Product"


def path_state(value: str) -> tuple[str, str]:
    path_value = norm(value)
    if not path_value:
        return "Empty", "#8a6d3b"
    return ("OK", "#2e7d32") if Path(path_value).exists() else ("Missing", "#c62828")


def ensure_routing(cfg: dict[str, object]) -> None:
    shared = cfg.setdefault("shared", {})
    routing = shared.setdefault("routing", {})
    routing.setdefault("surface_file_suffix_by_surface", {"flipkart": "", "shopsy": "-Shopsy", "default": ""})
    routing.setdefault("flow_directory_pattern", "{product_type}_{surface}")
    routing.setdefault("common_input_files", {
        "price_stock_shipping_excel": "common/Price-Stock-Shipping-inputs.xlsx",
        "price_stock_shipping_json": "Price-Stock-Shipping-inputs.json",
        "variants_excel": "common/Variants-excel.xlsx",
    })
    routing.setdefault("product_input_patterns", {
        "product_description_excel": "{product_type}/Product-Description-inputs{surface_suffix}.xlsx",
        "additional_description_excel": "{product_type}/Additional-Description-inputs{surface_suffix}.xlsx",
        "product_description_asset_json": "Product-Description-inputs-{product_title}.json",
        "additional_description_asset_json": "Additional-Description-inputs-{product_title}.json",
        "product_description_flow_json": "02_product_description.json",
        "additional_description_flow_json": "01_additional_description.json",
    })


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


class StatusLabel:
    def __init__(self, master: tk.Misc):
        self.var = tk.StringVar(value="Empty")
        self.label = ttk.Label(master, textvariable=self.var, width=9)

    def set_for_path(self, value: str) -> None:
        text, color = path_state(value)
        self.var.set(text)
        self.label.configure(foreground=color)


class BrandRow:
    def __init__(self, master: tk.Misc, index: int, on_change):
        self.label = ttk.Label(master, text=f"Brand {index}")
        self.name = tk.StringVar(value=f"Brand {index}")
        self.code = tk.StringVar(value=f"B{index}")
        self.name_entry = ttk.Entry(master, textvariable=self.name)
        self.code_entry = ttk.Entry(master, textvariable=self.code, width=12)
        self.name.trace_add("write", lambda *_: on_change())
        self.code.trace_add("write", lambda *_: on_change())

    def destroy(self) -> None:
        self.label.destroy(); self.name_entry.destroy(); self.code_entry.destroy()


class KindRow:
    def __init__(self, master: tk.Misc, vertical_name: tk.StringVar, index: int, on_change):
        self.label = ttk.Label(master, text=f"Kind {index}")
        self.name = tk.StringVar(value=f"Kind {index}")
        self.path = tk.StringVar()
        self.name_entry = ttk.Entry(master, textvariable=self.name)
        self.path_entry = ttk.Entry(master, textvariable=self.path)
        self.status = StatusLabel(master)
        self.name.trace_add("write", lambda *_: on_change())
        self.path.trace_add("write", lambda *_: on_change())
        self.browse_button = ttk.Button(master, text="Browse", command=lambda: self.browse(vertical_name))

    def browse(self, vertical_name: tk.StringVar) -> None:
        selected = filedialog.askdirectory(title=f"Select image folder for {vertical_name.get() or 'vertical'} / {self.name.get() or 'kind'}")
        if selected:
            self.path.set(selected)

    def destroy(self) -> None:
        self.label.destroy(); self.name_entry.destroy(); self.path_entry.destroy(); self.status.label.destroy(); self.browse_button.destroy()


class VerticalRow:
    def __init__(self, master: tk.Misc, index: int, on_change, scaffold_options: list[str]):
        self.on_change = on_change
        self.scaffold_options = scaffold_options
        self.frame = ttk.LabelFrame(master, text=f"Vertical {index}", padding=10)
        self.name = tk.StringVar(value=f"vertical{index}")
        self.default_size = tk.StringVar(value="28")
        self.default_kind = tk.StringVar(value="")
        self.scaffold_source = tk.StringVar(value=scaffold_options[0] if scaffold_options else "auto")
        self.kind_rows: list[KindRow] = []
        ttk.Label(self.frame, text="Vertical Name").grid(row=0, column=0, sticky="w")
        ttk.Entry(self.frame, textvariable=self.name).grid(row=1, column=0, sticky="ew", padx=(0, 8))
        ttk.Label(self.frame, text="Default Size").grid(row=0, column=1, sticky="w")
        ttk.Entry(self.frame, textvariable=self.default_size, width=12).grid(row=1, column=1, sticky="ew", padx=(0, 8))
        ttk.Label(self.frame, text="Default Kind").grid(row=0, column=2, sticky="w")
        ttk.Entry(self.frame, textvariable=self.default_kind).grid(row=1, column=2, sticky="ew", padx=(0, 8))
        ttk.Label(self.frame, text="Scaffold Source").grid(row=0, column=3, sticky="w")
        ttk.Combobox(self.frame, textvariable=self.scaffold_source, values=self.scaffold_options, state="readonly").grid(row=1, column=3, sticky="ew")
        self.frame.columnconfigure(0, weight=1)
        self.frame.columnconfigure(2, weight=1)
        self.frame.columnconfigure(3, weight=1)
        self.kinds_box = ttk.LabelFrame(self.frame, text="Kinds", padding=8)
        self.kinds_box.grid(row=2, column=0, columnspan=4, sticky="ew", pady=(12, 0))
        self.kinds_box.columnconfigure(2, weight=1)
        actions = ttk.Frame(self.kinds_box)
        actions.grid(row=0, column=0, columnspan=5, sticky="w", pady=(0, 8))
        ttk.Button(actions, text="Add Kind", command=self.add_kind_row).pack(side="left")
        ttk.Button(actions, text="Remove Last Kind", command=self.remove_kind_row).pack(side="left", padx=(8, 0))
        self.name.trace_add("write", lambda *_: on_change())
        self.default_size.trace_add("write", lambda *_: on_change())
        self.default_kind.trace_add("write", lambda *_: on_change())
        self.scaffold_source.trace_add("write", lambda *_: on_change())
        self.add_kind_row()

    def add_kind_row(self, kind_name: str = "", image_directory: str = "") -> None:
        row = KindRow(self.kinds_box, self.name, len(self.kind_rows) + 1, self.refresh_statuses)
        if kind_name:
            row.name.set(kind_name)
        if image_directory:
            row.path.set(image_directory)
        self.kind_rows.append(row)
        self.layout_kind_rows()

    def remove_kind_row(self) -> None:
        if len(self.kind_rows) <= 1:
            return
        self.kind_rows.pop().destroy()
        self.layout_kind_rows()

    def layout_kind_rows(self) -> None:
        for index, row in enumerate(self.kind_rows, start=1):
            row.label.configure(text=f"Kind {index}")
            row.label.grid(row=index, column=0, sticky="w", padx=(0, 8), pady=4)
            row.name_entry.grid(row=index, column=1, sticky="ew", pady=4)
            row.path_entry.grid(row=index, column=2, sticky="ew", pady=4)
            row.browse_button.grid(row=index, column=3, padx=(8, 0), pady=4)
            row.status.label.grid(row=index, column=4, padx=(8, 0), sticky="w", pady=4)
        self.refresh_statuses()

    def refresh_statuses(self) -> None:
        for row in self.kind_rows:
            row.status.set_for_path(row.path.get())
        self.on_change()


class AccountRow:
    def __init__(self, master: tk.Misc, index: int, on_change):
        self.on_change = on_change
        self.frame = ttk.LabelFrame(master, text=f"Account {index}", padding=10)
        self.profile = tk.StringVar(value=f"account{index}")
        self.alias = tk.StringVar(value=f"a{index}")
        self.firefox_profile = tk.StringVar()
        self.firefox_status = StatusLabel(self.frame)
        self.brand_rows: list[BrandRow] = []
        ttk.Label(self.frame, text="Profile").grid(row=0, column=0, sticky="w")
        ttk.Entry(self.frame, textvariable=self.profile, width=18).grid(row=1, column=0, sticky="ew", padx=(0, 8))
        ttk.Label(self.frame, text="Alias").grid(row=0, column=1, sticky="w")
        ttk.Entry(self.frame, textvariable=self.alias, width=12).grid(row=1, column=1, sticky="ew", padx=(0, 8))
        ttk.Label(self.frame, text="Firefox Profile Folder").grid(row=0, column=2, sticky="w")
        ttk.Entry(self.frame, textvariable=self.firefox_profile).grid(row=1, column=2, sticky="ew", padx=(0, 8))
        ttk.Button(self.frame, text="Browse", command=self.browse_firefox_profile).grid(row=1, column=3)
        self.firefox_status.label.grid(row=1, column=4, padx=(8, 0), sticky="w")
        self.frame.columnconfigure(2, weight=1)
        self.brands_box = ttk.LabelFrame(self.frame, text="Brands", padding=8)
        self.brands_box.grid(row=2, column=0, columnspan=5, sticky="ew", pady=(12, 0))
        self.brands_box.columnconfigure(1, weight=1)
        actions = ttk.Frame(self.brands_box)
        actions.grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 8))
        ttk.Button(actions, text="Add Brand", command=self.add_brand_row).pack(side="left")
        ttk.Button(actions, text="Remove Last Brand", command=self.remove_brand_row).pack(side="left", padx=(8, 0))
        self.profile.trace_add("write", lambda *_: on_change())
        self.alias.trace_add("write", lambda *_: on_change())
        self.firefox_profile.trace_add("write", lambda *_: self.refresh_statuses())
        self.add_brand_row()

    def browse_firefox_profile(self) -> None:
        selected = filedialog.askdirectory(title="Select Firefox profile folder")
        if selected:
            self.firefox_profile.set(selected)

    def add_brand_row(self, name: str = "", code: str = "") -> None:
        row = BrandRow(self.brands_box, len(self.brand_rows) + 1, self.on_change)
        if name:
            row.name.set(name)
        if code:
            row.code.set(code)
        self.brand_rows.append(row)
        self.layout_brand_rows()

    def remove_brand_row(self) -> None:
        if len(self.brand_rows) <= 1:
            return
        self.brand_rows.pop().destroy()
        self.layout_brand_rows()
        self.on_change()

    def layout_brand_rows(self) -> None:
        for index, row in enumerate(self.brand_rows, start=1):
            row.label.configure(text=f"Brand {index}")
            row.label.grid(row=index, column=0, sticky="w", padx=(0, 8), pady=4)
            row.name_entry.grid(row=index, column=1, sticky="ew", pady=4)
            row.code_entry.grid(row=index, column=2, sticky="ew", padx=(8, 0), pady=4)

    def refresh_statuses(self) -> None:
        self.firefox_status.set_for_path(self.firefox_profile.get())
        self.on_change()


class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Full LC Auto Setup Wizard")
        self.root.geometry("1320x920")
        self.root.minsize(1080, 740)
        self.cfg = load_config()
        ensure_routing(self.cfg)
        shared = dict(self.cfg.get("shared", {}))
        paths = dict(shared.get("project_paths", {}))
        self.laptops = dict(self.cfg.get("laptops", {}))
        self.laptop = tk.StringVar(value="NEW_LAPTOP")
        self.run_helpers = tk.StringVar(value=str(paths.get("run_helpers_directory", "run_helpers")))
        self.error_file = tk.StringVar(value=str(paths.get("error_latest_path", "run_helpers/error_latest.txt")))
        self.success_record = tk.StringVar(value=str(paths.get("success_run_record_path", "successful-run-record.xlsx")))
        self.flow_root = tk.StringVar(value=str(paths.get("flow_config_root", "json_LC_creation")))
        self.data_inputs_root = tk.StringVar(value=str(paths.get("data_inputs_root", "data inputs")))
        self.assets_root = tk.StringVar(value=str(paths.get("assets_root", "assets")))
        self.snapshot_dir = tk.StringVar(value="snapshots")
        self.apply_runtime = tk.BooleanVar(value=False)
        self.write_env = tk.BooleanVar(value=True)
        self.write_report = tk.BooleanVar(value=True)
        self.create_scaffold = tk.BooleanVar(value=True)
        self.preview_var = tk.StringVar(value="Waiting for setup details...")
        self.survey_note = tk.StringVar(value="Fill profiles, brands, verticals, kinds, and image folders. Excel and JSON scaffolds will be generated from your chosen source verticals.")
        self.path_statuses: list[tuple[tk.StringVar, StatusLabel]] = []
        self.account_rows: list[AccountRow] = []
        self.vertical_rows: list[VerticalRow] = []
        self.build_ui()
        self.add_account()
        self.add_vertical()
        self.bind_validators()

    def build_ui(self) -> None:
        style = ttk.Style(self.root)
        if "clam" in style.theme_names():
            style.theme_use("clam")
        style.configure("Shell.TFrame", background="#f3efe7")
        style.configure("Hero.TFrame", background="#17313e")
        style.configure("HeroTitle.TLabel", background="#17313e", foreground="#f6f1e8", font=("Segoe UI Semibold", 22))
        style.configure("HeroBody.TLabel", background="#17313e", foreground="#d8e0e4", font=("Segoe UI", 10))
        shell = ttk.Frame(self.root, style="Shell.TFrame", padding=16)
        shell.pack(fill="both", expand=True)
        hero = ttk.Frame(shell, style="Hero.TFrame", padding=(18, 16))
        hero.pack(fill="x")
        ttk.Label(hero, text="Full LC Auto Setup", style="HeroTitle.TLabel").pack(anchor="w")
        ttk.Label(hero, text="Define laptops, profiles, brands, verticals, and kinds. This wizard only captures setup information and generates scaffolds.", style="HeroBody.TLabel").pack(anchor="w", pady=(6, 0))
        ttk.Label(hero, textvariable=self.survey_note, style="HeroBody.TLabel").pack(anchor="w", pady=(4, 0))
        canvas = tk.Canvas(shell, background="#f3efe7", highlightthickness=0)
        scrollbar = ttk.Scrollbar(shell, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y", pady=(12, 0))
        canvas.pack(side="left", fill="both", expand=True, pady=(12, 0))
        body = ttk.Frame(canvas, style="Shell.TFrame")
        body_id = canvas.create_window((0, 0), window=body, anchor="nw")
        body.bind("<Configure>", lambda _e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure(body_id, width=e.width))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-e.delta / 120), "units"))

        top = ttk.Frame(body, style="Shell.TFrame")
        top.pack(fill="x")
        left = ttk.LabelFrame(top, text="Laptop", padding=12)
        right = ttk.LabelFrame(top, text="Shared Project Paths", padding=12)
        left.pack(side="left", fill="both", expand=True, padx=(0, 6))
        right.pack(side="left", fill="both", expand=True, padx=(6, 0))
        left.columnconfigure(0, weight=1); left.columnconfigure(2, weight=1); right.columnconfigure(0, weight=1)
        ttk.Label(left, text="Laptop Name").grid(row=0, column=0, sticky="w")
        ttk.Entry(left, textvariable=self.laptop).grid(row=1, column=0, sticky="ew", padx=(0, 8))
        ttk.Label(left, text="Snapshot Directory").grid(row=0, column=2, sticky="w")
        self.path_row(left, 1, self.snapshot_dir, 2, folder=True)
        self.path_row(right, 0, self.run_helpers, 0, folder=True, label="Run Helpers Directory")
        self.path_row(right, 2, self.error_file, 0, folder=False, label="Latest Error File")
        self.path_row(right, 4, self.success_record, 0, folder=False, label="Success Record Workbook")
        self.path_row(right, 6, self.flow_root, 0, folder=True, label="JSON Flow Root")
        self.path_row(right, 8, self.data_inputs_root, 0, folder=True, label="Data Inputs Root")
        self.path_row(right, 10, self.assets_root, 0, folder=True, label="Assets Root")

        accounts = ttk.LabelFrame(body, text="Accounts", padding=12)
        accounts.pack(fill="x", pady=(12, 0))
        bar = ttk.Frame(accounts); bar.pack(fill="x")
        ttk.Button(bar, text="Add Account", command=self.add_account).pack(side="left")
        ttk.Button(bar, text="Remove Last Account", command=self.remove_account).pack(side="left", padx=(8, 0))
        self.accounts_box = ttk.Frame(accounts)
        self.accounts_box.pack(fill="x", pady=(10, 0))

        verticals = ttk.LabelFrame(body, text="Verticals And Batch Intake", padding=12)
        verticals.pack(fill="x", pady=(12, 0))
        bar = ttk.Frame(verticals); bar.pack(fill="x")
        ttk.Button(bar, text="Add Vertical", command=self.add_vertical).pack(side="left")
        ttk.Button(bar, text="Remove Last Vertical", command=self.remove_vertical).pack(side="left", padx=(8, 0))
        self.verticals_box = ttk.Frame(verticals)
        self.verticals_box.pack(fill="x", pady=(10, 0))

        out = ttk.LabelFrame(body, text="Output", padding=12)
        out.pack(fill="x", pady=(12, 0))
        ttk.Checkbutton(out, text="Create starter scaffold files", variable=self.create_scaffold).pack(anchor="w")
        ttk.Checkbutton(out, text="Also replace config.json", variable=self.apply_runtime).pack(anchor="w", pady=(4, 0))
        ttk.Checkbutton(out, text="Write setup-runtime.env", variable=self.write_env).pack(anchor="w", pady=(4, 0))
        ttk.Checkbutton(out, text="Write setup-diagnostics.txt", variable=self.write_report).pack(anchor="w", pady=(4, 0))
        preview = ttk.LabelFrame(body, text="Preview", padding=12)
        preview.pack(fill="x", pady=(12, 0))
        ttk.Label(preview, textvariable=self.preview_var, justify="left").pack(anchor="w")
        actions = ttk.Frame(body, style="Shell.TFrame")
        actions.pack(fill="x", pady=(14, 20))
        ttk.Button(actions, text="Preview Output", command=self.preview).pack(side="left")
        ttk.Button(actions, text="Generate Files", command=self.generate).pack(side="left", padx=(10, 0))

    def path_row(self, master: tk.Misc, row: int, var: tk.StringVar, column: int, *, folder: bool, label: str | None = None) -> None:
        if label is not None:
            ttk.Label(master, text=label).grid(row=row, column=column, sticky="w", pady=(0 if row == 0 else 10, 0))
        ttk.Entry(master, textvariable=var).grid(row=row + 1, column=column, sticky="ew", padx=(0, 8))
        ttk.Button(master, text="Browse", command=lambda: self.pick_path(var, folder)).grid(row=row + 1, column=column + 1, sticky="w")
        status = StatusLabel(master)
        status.label.grid(row=row + 1, column=column + 2, sticky="w", padx=(8, 0))
        self.path_statuses.append((var, status))

    def bind_validators(self) -> None:
        for var in [self.laptop, self.run_helpers, self.error_file, self.success_record, self.flow_root, self.data_inputs_root, self.assets_root, self.snapshot_dir]:
            var.trace_add("write", lambda *_: self.refresh_statuses())
        self.refresh_statuses()

    def pick_path(self, var: tk.StringVar, folder: bool) -> None:
        selected = filedialog.askdirectory() if folder else filedialog.asksaveasfilename(initialfile=Path(var.get() or "path.txt").name)
        if selected:
            var.set(selected)

    def add_account(self, profile: str = "", alias: str = "", firefox_profile: str = "", brands: list[tuple[str, str]] | None = None) -> None:
        row = AccountRow(self.accounts_box, len(self.account_rows) + 1, self.refresh_statuses)
        if profile:
            row.profile.set(profile)
        if alias:
            row.alias.set(alias)
        if firefox_profile:
            row.firefox_profile.set(firefox_profile)
        if brands:
            while row.brand_rows:
                row.brand_rows.pop().destroy()
            for name, code in brands:
                row.add_brand_row(name, code)
        self.account_rows.append(row)
        for index, item in enumerate(self.account_rows, start=1):
            item.frame.configure(text=f"Account {index}")
            item.frame.pack(fill="x", pady=(0, 10))

    def remove_account(self) -> None:
        if len(self.account_rows) <= 1:
            return
        self.account_rows.pop().frame.destroy()
        self.refresh_statuses()

    def add_vertical(self, name: str = "", default_kind: str = "", default_size: str = "", kinds: list[tuple[str, str]] | None = None, scaffold_source: str = "auto") -> None:
        row = VerticalRow(self.verticals_box, len(self.vertical_rows) + 1, self.refresh_statuses, self.get_scaffold_template_options())
        if name:
            row.name.set(name)
        if default_kind:
            row.default_kind.set(default_kind)
        if default_size:
            row.default_size.set(default_size)
        if scaffold_source and scaffold_source in row.scaffold_options:
            row.scaffold_source.set(scaffold_source)
        if kinds:
            while row.kind_rows:
                row.kind_rows.pop().destroy()
            for kind_name, image_dir in kinds:
                row.add_kind_row(kind_name, image_dir)
        self.vertical_rows.append(row)
        for index, item in enumerate(self.vertical_rows, start=1):
            item.frame.configure(text=f"Vertical {index}")
            item.frame.pack(fill="x", pady=(0, 10))

    def remove_vertical(self) -> None:
        if len(self.vertical_rows) <= 1:
            return
        self.vertical_rows.pop().frame.destroy()
        self.refresh_statuses()

    def refresh_statuses(self) -> None:
        for var, status in self.path_statuses:
            status.set_for_path(var.get())
        for row in self.account_rows:
            row.firefox_status.set_for_path(row.firefox_profile.get())
        for row in self.vertical_rows:
            for kind_row in row.kind_rows:
                kind_row.status.set_for_path(kind_row.path.get())
        self.preview_var.set(self.build_preview_text())

    def build_preview_text(self) -> str:
        accounts = [row.profile.get().strip() for row in self.account_rows if row.profile.get().strip()]
        verticals = [row.name.get().strip() for row in self.vertical_rows if row.name.get().strip()]
        brand_count = sum(len(row.brand_rows) for row in self.account_rows)
        kind_count = sum(len(row.kind_rows) for row in self.vertical_rows)
        scaffold_lines = [f"{row.name.get().strip() or 'vertical'} -> {row.scaffold_source.get()}" for row in self.vertical_rows if row.name.get().strip()]
        return (
            f"Laptop: {self.laptop.get().strip().upper() or 'NEW_LAPTOP'}\n"
            f"Profiles: {len(accounts)} | Brands: {brand_count}\n"
            f"Verticals: {len(verticals)} | Kinds: {kind_count}\n"
            f"Scaffold: {'on' if self.create_scaffold.get() else 'off'}\n"
            f"Sources: {'; '.join(scaffold_lines) or 'none'}\n"
            f"Output: {DEMO.name}"
            + (" + config.json" if self.apply_runtime.get() else "")
            + (" + setup-runtime.env" if self.write_env.get() else "")
            + (" + setup-diagnostics.txt" if self.write_report.get() else "")
        )

    def resolve_user_path(self, value: str) -> Path:
        path = Path(value).expanduser()
        return path if path.is_absolute() else (ROOT / path).resolve()

    def get_scaffold_template_options(self) -> list[str]:
        options = ["auto"]
        seen: set[str] = {"auto"}
        for laptop in self.laptops.values():
            for name in laptop.get("verticals", {}).keys():
                normalized = slug(str(name))
                if normalized and normalized not in seen:
                    seen.add(normalized)
                    options.append(normalized)
        for name in self.cfg.get("shared", {}).get("products", {}).keys():
            normalized = slug(str(name))
            if normalized and normalized not in seen:
                seen.add(normalized)
                options.append(normalized)
        return options

    def get_surface_names(self) -> list[str]:
        mapping = self.cfg.get("shared", {}).get("surfaces", {}).get("folder_suffix_by_surface", {})
        names = [str(name).strip().lower() for name in mapping if str(name).strip().lower() != "default"]
        return names or ["flipkart"]

    def flow_dir_name(self, product_type: str, surface: str) -> str:
        pattern = str(self.cfg.get("shared", {}).get("routing", {}).get("flow_directory_pattern", "{product_type}_{surface}"))
        return pattern.format(product_type=slug(product_type), surface=surface.strip().lower())

    def choose_scaffold_source_vertical(self, target_vertical: str, requested_source: str, flow_root: Path, data_root: Path) -> str:
        candidates: list[str] = []
        requested = slug(requested_source)
        if requested and requested != "auto":
            candidates.append(requested)
        normalized = slug(target_vertical)
        if normalized in self.cfg.get("shared", {}).get("products", {}):
            candidates.append(normalized)
        for laptop in self.laptops.values():
            candidates.extend(slug(name) for name in laptop.get("verticals", {}).keys())
        candidates.extend(slug(name) for name in self.cfg.get("shared", {}).get("products", {}).keys())
        seen: set[str] = set()
        for candidate in candidates:
            if not candidate or candidate in seen:
                continue
            seen.add(candidate)
            if (data_root / candidate).exists() or any((flow_root / self.flow_dir_name(candidate, surface)).exists() for surface in self.get_surface_names()):
                return candidate
        raise ValueError("No existing vertical was found to use as a scaffold source.")

    def ensure_workbook_copy(self, source_path: Path, target_path: Path, sheet_renames: dict[str, str], extra_sheet_copies: dict[str, str] | None = None) -> None:
        target_path.parent.mkdir(parents=True, exist_ok=True)
        if not target_path.exists():
            shutil.copy2(source_path, target_path)
        workbook = load_workbook(target_path)
        for old_name, new_name in sheet_renames.items():
            if old_name in workbook.sheetnames and new_name not in workbook.sheetnames:
                workbook[old_name].title = new_name
        for source_sheet, new_sheet in (extra_sheet_copies or {}).items():
            if new_sheet in workbook.sheetnames:
                continue
            if source_sheet in workbook.sheetnames:
                clone = workbook.copy_worksheet(workbook[source_sheet])
                clone.title = new_sheet
            else:
                workbook.create_sheet(new_sheet)
        workbook.save(target_path)

    def update_flow_spec(self, path: Path, source_title: str, target_title: str, target_vertical: str, surface: str) -> None:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(payload, dict):
            if "product_type" in payload:
                payload["product_type"] = target_vertical
            if "surface" in payload:
                payload["surface"] = surface
            data_source = payload.get("data_source")
            if isinstance(data_source, dict):
                worksheet = str(data_source.get("worksheet", ""))
                replacements = {
                    f"{source_title} Product Inputs": f"{target_title} Product Inputs",
                    f"{source_title} Addl Desc Inputs": f"{target_title} Addl Desc Inputs",
                    f"{source_title} Variant Inputs": f"{target_title} Variant Inputs",
                }
                if worksheet in replacements:
                    data_source["worksheet"] = replacements[worksheet]
        write_json(path, payload)

    def scaffold_vertical_files(self, output: dict[str, object], verticals: list[dict[str, object]]) -> list[str]:
        source_paths = self.cfg.get("shared", {}).get("project_paths", {})
        source_flow_root = self.resolve_user_path(str(source_paths.get("flow_config_root", "json_LC_creation")))
        source_data_root = self.resolve_user_path(str(source_paths.get("data_inputs_root", "data inputs")))
        source_assets_root = self.resolve_user_path(str(source_paths.get("assets_root", "assets")))
        target_paths = output["shared"]["project_paths"]
        target_flow_root = self.resolve_user_path(str(target_paths["flow_config_root"]))
        target_data_root = self.resolve_user_path(str(target_paths["data_inputs_root"]))
        target_assets_root = self.resolve_user_path(str(target_paths["assets_root"]))
        routing = output["shared"]["routing"]
        product_patterns = routing["product_input_patterns"]
        surface_suffixes = routing["surface_file_suffix_by_surface"]
        created: list[str] = []

        common_source = source_data_root / "common"
        common_target = target_data_root / "common"
        common_target.mkdir(parents=True, exist_ok=True)
        pss_source = common_source / "Price-Stock-Shipping-inputs.xlsx"
        pss_target = common_target / "Price-Stock-Shipping-inputs.xlsx"
        if pss_source.exists() and not pss_target.exists():
            shutil.copy2(pss_source, pss_target)
            created.append(str(pss_target))
        pss_json_source = source_assets_root / "Price-Stock-Shipping-inputs.json"
        pss_json_target = target_assets_root / "Price-Stock-Shipping-inputs.json"
        if pss_json_source.exists() and not pss_json_target.exists():
            pss_json_target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(pss_json_source, pss_json_target)
            created.append(str(pss_json_target))

        for vertical in verticals:
            target_vertical = str(vertical["name"])
            target_title = title_token(target_vertical)
            source_vertical = self.choose_scaffold_source_vertical(target_vertical, str(vertical.get("scaffold_source", "auto")), source_flow_root, source_data_root)
            source_title = title_token(source_vertical)
            for surface in self.get_surface_names():
                suffix = str(surface_suffixes.get(surface, surface_suffixes.get("default", "")))
                pd_name = str(product_patterns["product_description_excel"]).format(product_type=target_vertical, product_title=target_title, surface=surface, surface_suffix=suffix)
                ad_name = str(product_patterns["additional_description_excel"]).format(product_type=target_vertical, product_title=target_title, surface=surface, surface_suffix=suffix)
                source_pd = source_data_root / str(product_patterns["product_description_excel"]).format(product_type=source_vertical, product_title=source_title, surface=surface, surface_suffix=suffix)
                source_ad = source_data_root / str(product_patterns["additional_description_excel"]).format(product_type=source_vertical, product_title=source_title, surface=surface, surface_suffix=suffix)
                target_pd = target_data_root / pd_name
                target_ad = target_data_root / ad_name
                if source_pd.exists() and target_pd.resolve() != source_pd.resolve():
                    self.ensure_workbook_copy(source_pd, target_pd, {f"{source_title} Product Inputs": f"{target_title} Product Inputs", f"{source_title} Field Guide": f"{target_title} Field Guide"})
                    created.append(str(target_pd))
                if source_ad.exists() and target_ad.resolve() != source_ad.resolve():
                    self.ensure_workbook_copy(source_ad, target_ad, {f"{source_title} Addl Desc Inputs": f"{target_title} Addl Desc Inputs", f"{source_title} Field Guide": f"{target_title} Field Guide"})
                    created.append(str(target_ad))
                source_flow_dir = source_flow_root / self.flow_dir_name(source_vertical, surface)
                target_flow_dir = target_flow_root / self.flow_dir_name(target_vertical, surface)
                if source_flow_dir.exists() and target_flow_dir.resolve() != source_flow_dir.resolve():
                    target_flow_dir.mkdir(parents=True, exist_ok=True)
                    for file in source_flow_dir.glob("*.json"):
                        shutil.copy2(file, target_flow_dir / file.name)
                        self.update_flow_spec(target_flow_dir / file.name, source_title, target_title, target_vertical, surface)
                    created.append(str(target_flow_dir))
            variants_source = common_source / "Variants-excel.xlsx"
            variants_target = common_target / "Variants-excel.xlsx"
            if variants_source.exists():
                self.ensure_workbook_copy(variants_source, variants_target, {}, {f"{source_title} Variant Inputs": f"{target_title} Variant Inputs", f"{source_title} Variant Guide": f"{target_title} Variant Guide"})
                created.append(str(variants_target))
        return sorted(set(created))

    def collect_accounts(self) -> list[dict[str, object]]:
        accounts = []
        for row in self.account_rows:
            profile = row.profile.get().strip()
            if not profile:
                continue
            brands = []
            for brand_row in row.brand_rows:
                name = brand_row.name.get().strip()
                code = brand_row.code.get().strip().upper()
                if name and code:
                    brands.append({"name": name, "code": code})
            accounts.append({"profile": profile, "alias": row.alias.get().strip() or profile[:1].lower(), "firefox_profile": norm(row.firefox_profile.get()), "brands": brands})
        if not accounts:
            raise ValueError("Add at least one account.")
        return accounts

    def collect_verticals(self) -> list[dict[str, object]]:
        verticals = []
        for row in self.vertical_rows:
            name = slug(row.name.get())
            if not name:
                continue
            kinds = []
            for kind_row in row.kind_rows:
                kind_name = kind_row.name.get().strip()
                image_directory = norm(kind_row.path.get())
                if kind_name and image_directory:
                    kinds.append({"kind": kind_name, "image_directory": image_directory})
            if not kinds:
                raise ValueError(f"Vertical '{row.name.get().strip() or name}' needs at least one kind with an image folder.")
            default_kind = row.default_kind.get().strip() or kinds[0]["kind"]
            if default_kind not in [item["kind"] for item in kinds]:
                raise ValueError(f"Default kind '{default_kind}' is not present under vertical '{name}'.")
            ordered = sorted(kinds, key=lambda item: 0 if item["kind"] == default_kind else 1)
            verticals.append({"name": name, "default_kind": default_kind, "default_size": row.default_size.get().strip() or "28", "kinds": ordered, "scaffold_source": row.scaffold_source.get().strip() or "auto"})
        if not verticals:
            raise ValueError("Add at least one vertical.")
        return verticals

    def build_output(self) -> tuple[dict[str, object], str, list[dict[str, object]]]:
        accounts = self.collect_accounts()
        verticals = self.collect_verticals()
        laptop_name = self.laptop.get().strip().upper()
        if not laptop_name:
            raise ValueError("Laptop name is required.")
        output = deepcopy(self.cfg)
        ensure_routing(output)
        shared = output.setdefault("shared", {})
        defaults = shared.setdefault("defaults", {})
        paths = shared.setdefault("project_paths", {})
        products = shared.setdefault("products", {})
        brands_cfg = shared.setdefault("brands", {})
        profiles_cfg = shared.setdefault("profiles", {})
        flows = shared.setdefault("legacy_product_page_flows", {})
        variants = shared.setdefault("common_inputs", {}).setdefault("variants", {}).setdefault("sheet_name_by_product_type", {})
        output.setdefault("laptops", {})

        paths["run_helpers_directory"] = self.run_helpers.get().strip() or "run_helpers"
        paths["error_latest_path"] = self.error_file.get().strip() or "run_helpers/error_latest.txt"
        paths["success_run_record_path"] = self.success_record.get().strip() or "successful-run-record.xlsx"
        paths["snapshot_directory"] = self.snapshot_dir.get().strip() or "snapshots"
        paths["flow_config_root"] = self.flow_root.get().strip() or "json_LC_creation"
        paths["data_inputs_root"] = self.data_inputs_root.get().strip() or "data inputs"
        paths["assets_root"] = self.assets_root.get().strip() or "assets"
        paths["image_folder_insight_output_path"] = str(Path("insights") / "image_folder_insight.xlsx")
        paths["legacy_root"] = "legacy"

        defaults["profile_name"] = str(accounts[0]["profile"])
        defaults["product_type"] = str(verticals[0]["name"])
        defaults["brand_name"] = next((str(account_brand["name"]) for account in accounts for account_brand in account["brands"]), "")
        defaults["listing_size"] = str(verticals[0]["default_size"])
        defaults.pop("jeans_kind", None)

        brand_code_map: dict[str, str] = {}
        profile_brand_codes: dict[str, list[str]] = {}
        aliases: dict[str, str] = {}
        firefox_profiles: dict[str, str] = {}
        for account in accounts:
            profile = str(account["profile"])
            aliases[str(account["alias"])] = profile
            aliases[profile] = profile
            firefox_profiles[profile] = str(account["firefox_profile"])
            profile_brand_codes[profile] = []
            for brand in account["brands"]:
                code = str(brand["code"]).upper()
                brand_code_map[code] = str(brand["name"])
                profile_brand_codes[profile].append(code)

        brands_cfg["brand_code_map"] = brand_code_map
        brands_cfg["profile_brand_codes"] = profile_brand_codes
        profiles_cfg["aliases"] = aliases
        shared.setdefault("success_run_record", {})["accounts"] = [str(account["profile"]) for account in accounts]

        laptop_verticals: dict[str, object] = {}
        for vertical in verticals:
            name = str(vertical["name"])
            title = title_token(name)
            laptop_verticals[name] = {"kinds": {str(index): {"kind": item["kind"], "image_directory": item["image_directory"]} for index, item in enumerate(vertical["kinds"], start=1)}}
            products[name] = {
                "default_kind_by_surface": {"default": str(vertical["default_kind"])},
                "default_size_by_surface": {"default": str(vertical["default_size"])},
                "sheet_names": {
                    "product_description_by_surface": {"default": f"{title} Product Inputs"},
                    "additional_description_by_surface": {"default": f"{title} Addl Desc Inputs"},
                },
            }
            flows[name] = list(DEFAULT_FLOW_STEPS)
            variants[name] = f"{title} Variant Inputs"

        output["laptops"][laptop_name] = {
            "firefox_profiles": firefox_profiles,
            "paths": {"snapshot_directory": self.snapshot_dir.get().strip() or "snapshots"},
            "verticals": laptop_verticals,
        }
        output["default_laptop_name"] = laptop_name
        env_text = f"FK_LAPTOP_NAME={laptop_name}\nFLIPKART_SNAPSHOT_DIR={self.snapshot_dir.get().strip() or 'snapshots'}\n"
        return output, env_text, verticals

    def preview(self) -> None:
        try:
            output, _, _ = self.build_output()
        except Exception as exc:
            messagebox.showerror("Preview failed", str(exc), parent=self.root)
            return
        snippet = json.dumps({
            "default_laptop_name": output["default_laptop_name"],
            "laptops": {output["default_laptop_name"]: output["laptops"][output["default_laptop_name"]]},
            "shared_defaults": {
                "profile_name": output["shared"]["defaults"].get("profile_name"),
                "product_type": output["shared"]["defaults"].get("product_type"),
                "brand_name": output["shared"]["defaults"].get("brand_name"),
            },
        }, indent=2)
        self.preview_var.set(snippet[:1600] + ("\n..." if len(snippet) > 1600 else ""))

    def generate(self) -> None:
        try:
            output, env_text, verticals = self.build_output()
            scaffold_paths = self.scaffold_vertical_files(output, verticals) if self.create_scaffold.get() else []
        except Exception as exc:
            messagebox.showerror("Generate failed", str(exc), parent=self.root)
            return
        DEMO.write_text(json.dumps(output, indent=2), encoding="utf-8")
        if self.apply_runtime.get():
            CONFIG.write_text(json.dumps(output, indent=2), encoding="utf-8")
        if self.write_env.get():
            ENV.write_text(env_text, encoding="utf-8")
        if self.write_report.get():
            lines = [
                f"Laptop: {output['default_laptop_name']}",
                f"Accounts: {', '.join(output['shared']['success_run_record']['accounts'])}",
                f"Verticals: {', '.join(output['laptops'][output['default_laptop_name']]['verticals'].keys())}",
                f"Scaffold created: {'yes' if self.create_scaffold.get() else 'no'}",
                f"Scaffold entries: {len(scaffold_paths)}",
                f"Config demo: {DEMO}",
                f"Runtime config updated: {'yes' if self.apply_runtime.get() else 'no'}",
                "",
                "Batch intake survey:",
            ]
            for vertical in verticals:
                lines.append(f"- {vertical['name']}: default kind={vertical['default_kind']}, kinds={len(vertical['kinds'])}, scaffold source={vertical.get('scaffold_source', 'auto')}")
            REPORT.write_text("\n".join(lines) + "\n", encoding="utf-8")
        messagebox.showinfo("Setup complete", f"Saved {DEMO.name}" + (f"\nScaffolded {len(scaffold_paths)} file targets" if self.create_scaffold.get() else ""), parent=self.root)


def main() -> None:
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
