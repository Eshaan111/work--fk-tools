from __future__ import annotations

import json
import re
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from app_paths import get_app_root

PROJECT_ROOT = get_app_root()
CONFIG_PATH = PROJECT_ROOT / "config.json"
OUTPUT_PATH = PROJECT_ROOT / "image_folder_insight.xlsx"
TOKEN_PATTERN = re.compile(r"([A-Z]+)('?[SF])?$")

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
KIND_FILL = PatternFill(fill_type="solid", fgColor="EAF3FF")
NONZERO_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
ZERO_FILL = PatternFill(fill_type="solid", fgColor="F3F4F6")
NOTES_LABEL_FILL = PatternFill(fill_type="solid", fgColor="FCE5CD")
WARNING_FILL = PatternFill(fill_type="solid", fgColor="FDE9D9")
HEADER_FONT = Font(bold=True, color="FFFFFF")
LABEL_FONT = Font(bold=True, color="1F1F1F")
ZERO_FONT = Font(color="7A7A7A")
NONZERO_FONT = Font(bold=True, color="215E21")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center")


def load_app_config() -> dict[str, object]:
    if not CONFIG_PATH.exists():
        raise FileNotFoundError(f"Config file was not found: {CONFIG_PATH}")
    with CONFIG_PATH.open("r", encoding="utf-8") as config_file:
        return json.load(config_file)


def resolve_config_path(path_value: str | None) -> Path | None:
    if path_value in (None, ""):
        return None
    resolved_path = Path(path_value).expanduser()
    if resolved_path.is_absolute():
        return resolved_path
    return (PROJECT_ROOT / resolved_path).resolve()


APP_CONFIG = load_app_config()
DEFAULT_LAPTOP_NAME = str(APP_CONFIG["default_laptop_name"]).upper()
SHARED_CONFIG: dict[str, object] = APP_CONFIG["shared"]
BRAND_CODE_MAP = OrderedDict(SHARED_CONFIG["brands"]["brand_code_map"])
PROFILE_BRAND_CODES = OrderedDict(
    (account_name, tuple(brand_codes))
    for account_name, brand_codes in SHARED_CONFIG["brands"]["profile_brand_codes"].items()
)
SURFACE_FOLDER_SUFFIX = OrderedDict(SHARED_CONFIG["surfaces"]["folder_suffix_by_surface"])
FOLDER_SUFFIX_SURFACE = {
    suffix.upper(): surface for surface, suffix in SURFACE_FOLDER_SUFFIX.items()
}
ACCOUNT_BRAND_COLUMNS: list[tuple[str, str, str]] = []
for account_name, brand_codes in PROFILE_BRAND_CODES.items():
    for brand_code in brand_codes:
        brand_name = BRAND_CODE_MAP[brand_code]
        for surface_name in SURFACE_FOLDER_SUFFIX:
            ACCOUNT_BRAND_COLUMNS.append((account_name, brand_name, surface_name))

BRAND_TO_ACCOUNT = {
    BRAND_CODE_MAP[brand_code]: account_name
    for account_name, brand_codes in PROFILE_BRAND_CODES.items()
    for brand_code in brand_codes
}


def get_kind_directories(laptop_name: str) -> OrderedDict[str, Path]:
    normalized_laptop_name = laptop_name.strip().upper()
    laptop_payload = APP_CONFIG["laptops"].get(normalized_laptop_name)
    if laptop_payload is None:
        raise ValueError(f"Unknown LAPTOP_NAME '{laptop_name}'. Choose ASUS or VAIO.")

    kind_directories: OrderedDict[str, Path] = OrderedDict()
    vertical_payloads = laptop_payload.get("verticals", {})
    for vertical_name in ("jeans", "trouser", "shorts"):
        vertical_payload = vertical_payloads.get(vertical_name, {})
        kinds_payload = vertical_payload.get("kinds", {})
        for option_key in sorted(kinds_payload, key=int):
            option_payload = kinds_payload[option_key]
            kind_directories[str(option_payload["kind"])] = resolve_config_path(option_payload["image_directory"])
    return kind_directories


def parse_folder_number(folder_name: str) -> int:
    first_token = folder_name.split("-", maxsplit=1)[0].strip()
    if not first_token.isdigit():
        raise ValueError(f"Folder name must start with a number: {folder_name}")
    return int(first_token)


def parse_brand_folder_token(token: str, folder_name: str) -> list[tuple[str, str]]:
    normalized_token = token.strip().upper()
    token_match = TOKEN_PATTERN.fullmatch(normalized_token)
    if token_match is None:
        raise ValueError(f"Unknown brand code '{token}' in folder '{folder_name}'")

    brand_code, suffix = token_match.groups()
    if brand_code not in BRAND_CODE_MAP:
        raise ValueError(f"Unknown brand code '{token}' in folder '{folder_name}'")

    brand_name = BRAND_CODE_MAP[brand_code]
    if not suffix:
        return [(brand_name, surface_name) for surface_name in SURFACE_FOLDER_SUFFIX]

    normalized_suffix = suffix if suffix.startswith("'") else f"'{suffix}"
    surface_name = FOLDER_SUFFIX_SURFACE.get(normalized_suffix.upper())
    if surface_name is None:
        raise ValueError(f"Unknown brand surface suffix '{token}' in folder '{folder_name}'")

    return [(brand_name, surface_name)]


def parse_exhausted_brand_surfaces(folder_name: str) -> list[tuple[str, str]]:
    parse_folder_number(folder_name)
    exhausted_pairs: list[tuple[str, str]] = []
    tokens = [part.strip() for part in folder_name.split("-")[1:] if part.strip()]
    for token in tokens:
        exhausted_pairs.extend(parse_brand_folder_token(token, folder_name))
    return exhausted_pairs


def scan_kind_directory(
    kind_name: str,
    folder_root: Path,
) -> tuple[dict[tuple[str, str, str], int], list[str], int]:
    counts = {
        (account_name, brand_name, surface_name): 0
        for account_name, brand_name, surface_name in ACCOUNT_BRAND_COLUMNS
    }
    warnings: list[str] = []
    total_numbered_folders = 0

    if not folder_root.exists():
        warnings.append(f"{kind_name}: missing directory -> {folder_root}")
        return counts, warnings, total_numbered_folders
    if not folder_root.is_dir():
        warnings.append(f"{kind_name}: not a directory -> {folder_root}")
        return counts, warnings, total_numbered_folders

    for child in sorted(folder_root.iterdir(), key=lambda path: path.name.lower()):
        if not child.is_dir():
            continue

        try:
            parse_folder_number(child.name)
        except ValueError as exc:
            warnings.append(f"{kind_name}: skipped folder '{child.name}' ({exc})")
            continue

        total_numbered_folders += 1

        try:
            exhausted_pairs = parse_exhausted_brand_surfaces(child.name)
        except ValueError as exc:
            warnings.append(f"{kind_name}: skipped folder tokens in '{child.name}' ({exc})")
            continue

        for brand_name, surface_name in exhausted_pairs:
            account_name = BRAND_TO_ACCOUNT.get(brand_name)
            if account_name is None:
                warnings.append(
                    f"{kind_name}: skipped brand '{brand_name}' from folder '{child.name}' because no account mapping exists"
                )
                continue
            counts[(account_name, brand_name, surface_name)] += 1

    return counts, warnings, total_numbered_folders


def autosize_columns(worksheet) -> None:
    widths: dict[int, int] = {}
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for column_index, width in widths.items():
        worksheet.column_dimensions[worksheet.cell(row=1, column=column_index).column_letter].width = min(width + 3, 32)


def style_header_cell(cell) -> None:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = THIN_BORDER
    cell.alignment = CENTER_ALIGNMENT


def style_kind_cell(cell) -> None:
    cell.font = LABEL_FONT
    cell.fill = KIND_FILL
    cell.border = THIN_BORDER
    cell.alignment = LEFT_ALIGNMENT


def style_count_cell(cell, value: int) -> None:
    cell.border = THIN_BORDER
    cell.alignment = CENTER_ALIGNMENT
    if value > 0:
        cell.fill = NONZERO_FILL
        cell.font = NONZERO_FONT
    else:
        cell.fill = ZERO_FILL
        cell.font = ZERO_FONT


def build_workbook(laptop_name: str | None = None) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Image Folder Insight"
    normalized_laptop_name = (laptop_name or DEFAULT_LAPTOP_NAME).strip().upper()
    kind_directories = get_kind_directories(normalized_laptop_name)

    header_cell = worksheet.cell(row=1, column=1, value="Kind")
    style_header_cell(header_cell)
    for column_index, (account_name, brand_name, surface_name) in enumerate(ACCOUNT_BRAND_COLUMNS, start=2):
        header = f"{account_name.title()} - {brand_name} - {surface_name.title()}"
        cell = worksheet.cell(row=1, column=column_index, value=header)
        style_header_cell(cell)

    available_sheet = workbook.create_sheet("Available Options")
    available_kind_header = available_sheet.cell(row=1, column=1, value="Kind")
    style_header_cell(available_kind_header)
    total_folders_header = available_sheet.cell(row=1, column=2, value="Total Numbered Folders")
    style_header_cell(total_folders_header)
    for column_index, (account_name, brand_name, surface_name) in enumerate(ACCOUNT_BRAND_COLUMNS, start=3):
        header = f"{account_name.title()} - {brand_name} - {surface_name.title()}"
        cell = available_sheet.cell(row=1, column=column_index, value=header)
        style_header_cell(cell)

    warning_messages: list[str] = []
    for row_index, (kind_name, folder_root) in enumerate(kind_directories.items(), start=2):
        kind_cell = worksheet.cell(row=row_index, column=1, value=kind_name)
        style_kind_cell(kind_cell)
        available_kind_cell = available_sheet.cell(row=row_index, column=1, value=kind_name)
        style_kind_cell(available_kind_cell)

        counts, warnings, total_numbered_folders = scan_kind_directory(kind_name, folder_root)
        warning_messages.extend(warnings)

        total_cell = available_sheet.cell(row=row_index, column=2, value=total_numbered_folders)
        style_count_cell(total_cell, total_numbered_folders)

        for column_index, key in enumerate(ACCOUNT_BRAND_COLUMNS, start=2):
            value = counts[key]
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            style_count_cell(cell, value)

        for column_index, key in enumerate(ACCOUNT_BRAND_COLUMNS, start=3):
            available_value = max(total_numbered_folders - counts[key], 0)
            cell = available_sheet.cell(row=row_index, column=column_index, value=available_value)
            style_count_cell(cell, available_value)

    worksheet.freeze_panes = "B2"
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 28
    autosize_columns(worksheet)

    available_sheet.freeze_panes = "C2"
    available_sheet.sheet_view.showGridLines = False
    available_sheet.row_dimensions[1].height = 28
    autosize_columns(available_sheet)

    notes = workbook.create_sheet("Notes")
    notes_rows = [
        ("Generated From", f"Folder names under the configured image roots for {normalized_laptop_name}"),
        ("Rule", "Each exhausted brand/surface token in a numbered folder counts as one successful listing"),
        ("Suffix 'f", "Flipkart"),
        ("Suffix 's", "Shopsy"),
        ("No suffix", "Counts for both Flipkart and Shopsy"),
        ("Accounts", "Brand ownership is inferred from config.json profile_brand_codes"),
        ("Available Options", "Each value is total numbered folders minus exhausted count for that account/brand/surface"),
    ]
    for row_index, (label, value) in enumerate(notes_rows, start=1):
        label_cell = notes.cell(row=row_index, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.fill = NOTES_LABEL_FILL
        label_cell.border = THIN_BORDER
        label_cell.alignment = LEFT_ALIGNMENT

        value_cell = notes.cell(row=row_index, column=2, value=value)
        value_cell.border = THIN_BORDER
        value_cell.alignment = LEFT_ALIGNMENT

    if warning_messages:
        warning_header = notes.cell(row=8, column=1, value="Warnings")
        warning_header.font = LABEL_FONT
        warning_header.fill = WARNING_FILL
        warning_header.border = THIN_BORDER
        warning_header.alignment = LEFT_ALIGNMENT
        for row_index, warning in enumerate(warning_messages, start=9):
            warning_cell = notes.cell(row=row_index, column=1, value=warning)
            warning_cell.border = THIN_BORDER
            warning_cell.alignment = LEFT_ALIGNMENT

    notes.sheet_view.showGridLines = False
    autosize_columns(notes)
    return workbook


def generate_workbook(laptop_name: str | None = None) -> Path:
    workbook = build_workbook(laptop_name=laptop_name)
    try:
        try:
            workbook.save(OUTPUT_PATH)
            return OUTPUT_PATH
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            fallback_path = PROJECT_ROOT / f"image_folder_insight_{timestamp}.xlsx"
            workbook.save(fallback_path)
            return fallback_path
    finally:
        workbook.close()

def main(laptop_name: str | None = None) -> Path:
    try:
        output_path = generate_workbook(laptop_name=laptop_name)
        print(f"Saved insight workbook: {output_path}")
        return output_path
    except KeyboardInterrupt:
        raise SystemExit("Interrupted by user.") from None

if __name__ == "__main__":
    main()

