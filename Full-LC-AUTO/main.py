from __future__ import annotations

import json
import os
import random
import re
import string
import subprocess
import sys
import threading
import tkinter as tk
from datetime import datetime
from dataclasses import dataclass, field
from pathlib import Path
from time import sleep

import msvcrt
from openpyxl import Workbook, load_workbook
import pyautogui
from pynput.mouse import Button as PynputButton
from pynput.mouse import Controller as MouseController

from selenium import webdriver
from selenium.common.exceptions import (
    StaleElementReferenceException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


PROJECT_ROOT = Path(__file__).resolve().parent
RUN_HELPERS_DIRECTORY = PROJECT_ROOT / "run_helpers"
ERROR_LATEST_PATH = RUN_HELPERS_DIRECTORY / "error_latest.txt"
SUCCESS_RUN_RECORD_PATH = PROJECT_ROOT / "successful-run-record.xlsx"
SUCCESS_RUN_ACCOUNTS = ("prabhu", "seema")
SUCCESS_RUN_DATA_START_ROW = 3
# LAPTOP_NAME = os.getenv("ASUS", "VAIO").upper()
LAPTOP_NAME = "ASUS"
# LAPTOP_NAME = "VAIO"

DEFAULT_IMAGE_DIRECTORY_ASUS = Path(r"C:\work-mom\LISTING IMAGES AUTOMATED\Shorts")
DEFAULT_IMAGE_DIRECTORY_VAIO: Path | None = Path(r"C:\LISTING IMAGES AUTOMATED\Shorts")
DEFAULT_TROUSER_IMAGE_DIRECTORY_ASUS = Path(r"C:\work-mom\LISTING IMAGES AUTOMATED\Trouser")
DEFAULT_TROUSER_IMAGE_DIRECTORY_VAIO = Path(r"G:\Other computers\My Laptop\work-mom\LISTING IMAGES AUTOMATED\Trouser")
JEANS_KIND_OPTIONS_ASUS: dict[str, tuple[str, Path]] = {
    "1": ("Beige", Path(r"C:\work-mom\JEANS\PRODUCT IMAGES\BEIGE\NEW IMAGES")),
    "2": ("Ice", Path(r"C:\work-mom\JEANS\PRODUCT IMAGES\ICE\NEW IMAGES")),
    "3": ("Black-baggy", Path(r"C:\work-mom\JEANS\PRODUCT IMAGES\BLACK-BAGGY\NEW IMAGES")),
    "4": ("Black-Plain", Path(r"C:\work-mom\JEANS\PRODUCT IMAGES\BLACK-PLAIN\NEW IMAGES")),
    "5": ("White-Plain", Path(r"C:\work-mom\JEANS\PRODUCT IMAGES\WHITE-PLAIN\NEW IMAGES")),
}
JEANS_KIND_OPTIONS_VAIO: dict[str, tuple[str, Path | None]] = {
    "1": ("Beige", Path(r"G:\Other computers\My Laptop\work-mom\LISTING IMAGES AUTOMATED\BEIGE")),
    "2": ("Ice", Path(r"G:\Other computers\My Laptop\work-mom\LISTING IMAGES AUTOMATED\ICE")),
    "3": ("Black-baggy", Path(r"G:\Other computers\My Laptop\work-mom\LISTING IMAGES AUTOMATED\Black-baggy")),
    "4": ("Black-Plain", Path(r"G:\Other computers\My Laptop\work-mom\LISTING IMAGES AUTOMATED\Black-Plain")),
    "5": ("White-Plain", Path(r"G:\Other computers\My Laptop\work-mom\LISTING IMAGES AUTOMATED\White")),
}
DEFAULT_SNAPSHOT_DIRECTORY_ASUS = PROJECT_ROOT / "snapshots"
DEFAULT_SNAPSHOT_DIRECTORY_VAIO = PROJECT_ROOT / "snapshots"
DATA_INPUTS_ROOT = PROJECT_ROOT / "data inputs"
COMMON_DATA_INPUTS_ROOT = DATA_INPUTS_ROOT / "common"
JEANS_DATA_INPUTS_ROOT = DATA_INPUTS_ROOT / "jeans"
SHORTS_DATA_INPUTS_ROOT = DATA_INPUTS_ROOT / "shorts"
TROUSER_DATA_INPUTS_ROOT = DATA_INPUTS_ROOT / "trouser"

DEFAULT_PRICE_STOCK_SHIPPING_EXCEL_ASUS = (
    COMMON_DATA_INPUTS_ROOT / "Price-Stock-Shipping-inputs.xlsx"
)
DEFAULT_PRICE_STOCK_SHIPPING_EXCEL_VAIO = (
    COMMON_DATA_INPUTS_ROOT / "Price-Stock-Shipping-inputs.xlsx"
)
DEFAULT_PRICE_STOCK_SHIPPING_JSON_ASUS = (
    PROJECT_ROOT / "assets" / "Price-Stock-Shipping-inputs.json"
)
DEFAULT_PRICE_STOCK_SHIPPING_JSON_VAIO = (
    PROJECT_ROOT / "assets" / "Price-Stock-Shipping-inputs.json"
)
DEFAULT_PRODUCT_DESCRIPTION_EXCEL_ASUS = (
    JEANS_DATA_INPUTS_ROOT / "Product-Description-inputs.xlsx"
)
DEFAULT_PRODUCT_DESCRIPTION_EXCEL_VAIO = (
    JEANS_DATA_INPUTS_ROOT / "Product-Description-inputs.xlsx"
)
DEFAULT_PRODUCT_DESCRIPTION_SHORTS_EXCEL_ASUS = (
    SHORTS_DATA_INPUTS_ROOT / "Product-Description-inputs.xlsx"
)
DEFAULT_PRODUCT_DESCRIPTION_SHORTS_EXCEL_VAIO = (
    SHORTS_DATA_INPUTS_ROOT / "Product-Description-inputs.xlsx"
)
DEFAULT_PRODUCT_DESCRIPTION_JSON_ASUS = (
    PROJECT_ROOT / "assets" / "Product-Description-inputs-Jeans.json"
)
DEFAULT_PRODUCT_DESCRIPTION_JSON_VAIO = (
    PROJECT_ROOT / "assets" / "Product-Description-inputs-Jeans.json"
)
DEFAULT_PRODUCT_DESCRIPTION_SHORTS_JSON_ASUS = (
    PROJECT_ROOT / "assets" / "Product-Description-inputs-Shorts.json"
)
DEFAULT_PRODUCT_DESCRIPTION_SHORTS_JSON_VAIO = (
    PROJECT_ROOT / "assets" / "Product-Description-inputs-Shorts.json"
)
DEFAULT_ADDITIONAL_DESCRIPTION_EXCEL_ASUS = (
    JEANS_DATA_INPUTS_ROOT / "Additional-Description-inputs.xlsx"
)
DEFAULT_ADDITIONAL_DESCRIPTION_EXCEL_VAIO = (
    JEANS_DATA_INPUTS_ROOT / "Additional-Description-inputs.xlsx"
)
DEFAULT_ADDITIONAL_DESCRIPTION_SHORTS_EXCEL_ASUS = (
    SHORTS_DATA_INPUTS_ROOT / "Additional-Description-inputs.xlsx"
)
DEFAULT_ADDITIONAL_DESCRIPTION_SHORTS_EXCEL_VAIO = (
    SHORTS_DATA_INPUTS_ROOT / "Additional-Description-inputs.xlsx"
)
DEFAULT_ADDITIONAL_DESCRIPTION_JSON_ASUS = (
    PROJECT_ROOT / "assets" / "Additional-Description-inputs-Jeans.json"
)
DEFAULT_ADDITIONAL_DESCRIPTION_JSON_VAIO = (
    PROJECT_ROOT / "assets" / "Additional-Description-inputs-Jeans.json"
)
DEFAULT_ADDITIONAL_DESCRIPTION_SHORTS_JSON_ASUS = (
    PROJECT_ROOT / "assets" / "Additional-Description-inputs-Shorts.json"
)
DEFAULT_ADDITIONAL_DESCRIPTION_SHORTS_JSON_VAIO = (
    PROJECT_ROOT / "assets" / "Additional-Description-inputs-Shorts.json"
)
DEFAULT_VARIANTS_EXCEL_ASUS = COMMON_DATA_INPUTS_ROOT / "Variants-excel.xlsx"
DEFAULT_VARIANTS_EXCEL_VAIO = COMMON_DATA_INPUTS_ROOT / "Variants-excel.xlsx"
FIREFOX_PROFILES_ASUS: dict[str, Path | None] = {
    "seema": Path(r"C:\Users\ESHAAN\AppData\Roaming\Mozilla\Firefox\Profiles\shW5ycIM.Profile 2"),
    # "prabhu": Path(r"C:\Users\ESHAAN\Documents\Firefox-Profiles\7kkhlz7p.prabhu-bt"),
    "prabhu": Path(r"C:\Users\ESHAAN\AppData\Roaming\Mozilla\Firefox\Profiles\kw3a25lf.default-release"),
}
FIREFOX_PROFILES_VAIO: dict[str, Path | None] = {
    "seema": Path(r"C:\Users\SONY\AppData\Roaming\Mozilla\Firefox\Profiles\S2AIZkXx.Profile 1"),
    "prabhu": Path(r"C:\Users\SONY\AppData\Roaming\Mozilla\Firefox\Profiles\gm1pmawk.default-release"),
}
LAPTOP_CONFIGS = {
    "ASUS": {
        "default_image_directory": DEFAULT_IMAGE_DIRECTORY_ASUS,
        "trouser_image_directory": DEFAULT_TROUSER_IMAGE_DIRECTORY_ASUS,
        "jeans_kind_options": JEANS_KIND_OPTIONS_ASUS,
        "snapshot_directory": DEFAULT_SNAPSHOT_DIRECTORY_ASUS,
        "price_stock_shipping_excel": DEFAULT_PRICE_STOCK_SHIPPING_EXCEL_ASUS,
        "price_stock_shipping_json": DEFAULT_PRICE_STOCK_SHIPPING_JSON_ASUS,
        "product_description_excel": DEFAULT_PRODUCT_DESCRIPTION_EXCEL_ASUS,
        "product_description_shorts_excel": DEFAULT_PRODUCT_DESCRIPTION_SHORTS_EXCEL_ASUS,
        "product_description_json": DEFAULT_PRODUCT_DESCRIPTION_JSON_ASUS,
        "product_description_shorts_json": DEFAULT_PRODUCT_DESCRIPTION_SHORTS_JSON_ASUS,
        "additional_description_excel": DEFAULT_ADDITIONAL_DESCRIPTION_EXCEL_ASUS,
        "additional_description_shorts_excel": DEFAULT_ADDITIONAL_DESCRIPTION_SHORTS_EXCEL_ASUS,
        "additional_description_json": DEFAULT_ADDITIONAL_DESCRIPTION_JSON_ASUS,
        "additional_description_shorts_json": DEFAULT_ADDITIONAL_DESCRIPTION_SHORTS_JSON_ASUS,
        "variants_excel": DEFAULT_VARIANTS_EXCEL_ASUS,
        "firefox_profiles": FIREFOX_PROFILES_ASUS,
    },
    "VAIO": {
        "default_image_directory": DEFAULT_IMAGE_DIRECTORY_VAIO,
        "trouser_image_directory": DEFAULT_TROUSER_IMAGE_DIRECTORY_VAIO,
        "jeans_kind_options": JEANS_KIND_OPTIONS_VAIO,
        "snapshot_directory": DEFAULT_SNAPSHOT_DIRECTORY_VAIO,
        "price_stock_shipping_excel": DEFAULT_PRICE_STOCK_SHIPPING_EXCEL_VAIO,
        "price_stock_shipping_json": DEFAULT_PRICE_STOCK_SHIPPING_JSON_VAIO,
        "product_description_excel": DEFAULT_PRODUCT_DESCRIPTION_EXCEL_VAIO,
        "product_description_shorts_excel": DEFAULT_PRODUCT_DESCRIPTION_SHORTS_EXCEL_VAIO,
        "product_description_json": DEFAULT_PRODUCT_DESCRIPTION_JSON_VAIO,
        "product_description_shorts_json": DEFAULT_PRODUCT_DESCRIPTION_SHORTS_JSON_VAIO,
        "additional_description_excel": DEFAULT_ADDITIONAL_DESCRIPTION_EXCEL_VAIO,
        "additional_description_shorts_excel": DEFAULT_ADDITIONAL_DESCRIPTION_SHORTS_EXCEL_VAIO,
        "additional_description_json": DEFAULT_ADDITIONAL_DESCRIPTION_JSON_VAIO,
        "additional_description_shorts_json": DEFAULT_ADDITIONAL_DESCRIPTION_SHORTS_JSON_VAIO,
        "variants_excel": DEFAULT_VARIANTS_EXCEL_VAIO,
        "firefox_profiles": FIREFOX_PROFILES_VAIO,
    },
}
if LAPTOP_NAME not in LAPTOP_CONFIGS:
    raise ValueError(f"Unknown FK_LAPTOP_NAME '{LAPTOP_NAME}'. Choose ASUS or VAIO.")
ACTIVE_LAPTOP_CONFIG = LAPTOP_CONFIGS[LAPTOP_NAME]
DEFAULT_LISTING_URL = (
    "https://seller.flipkart.com/index.html#dashboard/addListings/single"
    "?vertical=jean&vid=667"
)
DEFAULT_BRAND_NAME = "STARVIELLE"
DEFAULT_FLOW_SURFACE = "flipkart"
FLOW_CONFIG_ROOT = PROJECT_ROOT / "json_LC_creation"
USE_CHANGES_SAVED_TOAST_FOR_VERIFICATION = True
DEFAULT_JEANS_KIND = "Beige"
DEFAULT_LISTING_SIZE = "28"
PHASE_ONE_SNAPSHOT_NAME = "PHASE 1.html"
ENABLE_ENTER_COMMIT_FOR_TAG_INPUT_FIELDS = False
IMAGE_SLOT_IDS = [
    "thumbnail_0",
    "thumbnail_1",
    "thumbnail_2",
    "thumbnail_3",
    "thumbnail_4",
]
IMAGE_UPLOAD_VERIFY_TIMEOUT_SECONDS = 60
IMAGE_UPLOAD_RETRY_PASSES = 2
SUCCESS_CLOSE_DELAY_SECONDS = 5
# Legacy fixed-point VERIFYING CHANGES CYCLE values kept only for reference.
# PRODUCT_DESCRIPTION_TAB_CLICK_POINT_ASUS = (770, 380)
# SKU_PAGE_ALTERNATE_CLICK_POINT_ASUS = (518, 379)
BRAND_CODE_MAP = {
    "STAR": "STARVIELLE",
    "GENZ": "GENZ VANE",
    "IND": "INDIVANE",
    "FADE": "FADEVIELLE",
    "FLEE": "FLEECRANE",
}
PROFILE_BRAND_CODES = {
    "prabhu": ("STAR", "GENZ"),
    "seema": ("FADE", "FLEE", "IND"),
}
BRAND_NAME_TO_CODE = {
    normalize_name: code for code, normalize_name in (
        (code, " ".join(name.strip().upper().split())) for code, name in BRAND_CODE_MAP.items()
    )
}
SURFACE_FOLDER_SUFFIX = {
    "shopsy": "'s",
    "flipkart": "'f",
}
FOLDER_SUFFIX_SURFACE = {
    suffix.upper(): surface for surface, suffix in SURFACE_FOLDER_SUFFIX.items()
}
PROFILE_ALIASES = {
    "s": "seema",
    "seema": "seema",
    "p": "prabhu",
    "prabhu": "prabhu",
}
CURRENT_RUN_LABEL = "setup"


def set_current_run_label(run_label: str) -> None:
    global CURRENT_RUN_LABEL
    CURRENT_RUN_LABEL = run_label


def log_event(stage: str, message: str) -> None:
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp} {CURRENT_RUN_LABEL}] [{stage}] {message}")


def write_latest_error(error_message: str) -> None:
    RUN_HELPERS_DIRECTORY.mkdir(parents=True, exist_ok=True)
    ERROR_LATEST_PATH.write_text(f"{str(error_message).strip()}\n", encoding="utf-8")


def parse_brand_count_cell(cell_value: object) -> dict[str, int]:
    parsed_counts: dict[str, int] = {}
    if cell_value in (None, ""):
        return parsed_counts

    for token in str(cell_value).split():
        match = re.fullmatch(r"(\d+)-([A-Z0-9]+)", token.strip().upper())
        if match is None:
            continue
        parsed_counts[match.group(2)] = int(match.group(1))
    return parsed_counts


def format_brand_count_cell(brand_counts: dict[str, int]) -> str:
    return " ".join(
        f"{count}-{brand_code}"
        for brand_code, count in brand_counts.items()
        if count > 0
    )


def ensure_success_run_record_headers(worksheet) -> None:
    worksheet.title = "Successful Runs"
    worksheet.cell(row=1, column=1, value="Kind")
    worksheet.cell(row=1, column=2, value="Surface")
    worksheet.cell(row=2, column=1, value="")
    worksheet.cell(row=2, column=2, value="")


def ensure_date_account_columns(worksheet, run_date: str) -> dict[str, int]:
    account_columns: dict[str, int] = {}
    max_column = max(worksheet.max_column, 2)

    for column_index in range(3, max_column + 1):
        header_date = str(worksheet.cell(row=1, column=column_index).value or "").strip()
        header_account = str(worksheet.cell(row=2, column=column_index).value or "").strip().lower()
        if header_date == run_date and header_account in SUCCESS_RUN_ACCOUNTS:
            account_columns[header_account] = column_index

    if len(account_columns) == len(SUCCESS_RUN_ACCOUNTS):
        return account_columns

    next_column = max_column + 1
    for account_name in SUCCESS_RUN_ACCOUNTS:
        if account_name in account_columns:
            continue
        worksheet.cell(row=1, column=next_column, value=run_date)
        worksheet.cell(row=2, column=next_column, value=account_name.title())
        account_columns[account_name] = next_column
        next_column += 1

    return account_columns


def get_or_create_success_run_row(worksheet, listing_selection: ListingSelection) -> int:
    normalized_kind = listing_selection.kind.strip().lower()
    normalized_surface = listing_selection.surface.strip().lower()

    for row_index in range(SUCCESS_RUN_DATA_START_ROW, worksheet.max_row + 1):
        row_kind = str(worksheet.cell(row=row_index, column=1).value or "").strip().lower()
        row_surface = str(worksheet.cell(row=row_index, column=2).value or "").strip().lower()
        if row_kind == normalized_kind and row_surface == normalized_surface:
            return row_index

    target_row_index = max(worksheet.max_row + 1, SUCCESS_RUN_DATA_START_ROW)
    worksheet.cell(row=target_row_index, column=1, value=listing_selection.kind.strip())
    worksheet.cell(row=target_row_index, column=2, value=listing_selection.surface.strip())
    return target_row_index


def get_brand_code_for_record(brand_name: str) -> str:
    normalized_brand = normalize_brand_name(brand_name)
    if normalized_brand not in BRAND_NAME_TO_CODE:
        raise ValueError(f"Unknown brand for successful run record: {brand_name}")
    return BRAND_NAME_TO_CODE[normalized_brand]


def record_successful_run(listing_selection: ListingSelection, profile_name: str) -> Path:
    run_date = datetime.now().strftime("%Y-%m-%d")
    if SUCCESS_RUN_RECORD_PATH.exists():
        workbook = load_workbook(SUCCESS_RUN_RECORD_PATH)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
    ensure_success_run_record_headers(worksheet)

    normalized_profile = resolve_profile_name(profile_name)
    account_columns = ensure_date_account_columns(worksheet, run_date)
    target_row_index = get_or_create_success_run_row(worksheet, listing_selection)
    target_column_index = account_columns[normalized_profile]
    brand_code = get_brand_code_for_record(listing_selection.brand_name)

    existing_counts = parse_brand_count_cell(
        worksheet.cell(row=target_row_index, column=target_column_index).value
    )
    existing_counts[brand_code] = existing_counts.get(brand_code, 0) + 1
    worksheet.cell(
        row=target_row_index,
        column=target_column_index,
        value=format_brand_count_cell(existing_counts),
    )
    workbook.save(SUCCESS_RUN_RECORD_PATH)
    return SUCCESS_RUN_RECORD_PATH


def require_configured_path(path_value: Path | None, path_label: str) -> Path:
    if path_value is not None:
        return path_value

    manual_path = input(f"Paste the {LAPTOP_NAME} path for {path_label}: ").strip().strip('"')
    if not manual_path:
        raise ValueError(f"No path was provided for {path_label}.")
    return Path(manual_path).expanduser()


def active_path(config_key: str, path_label: str) -> Path:
    return require_configured_path(ACTIVE_LAPTOP_CONFIG[config_key], path_label)


@dataclass(slots=True)
class BotConfig:
    listing_url: str = DEFAULT_LISTING_URL
    image_directory: Path = field(
        default_factory=lambda: Path(
            os.getenv(
                "FLIPKART_IMAGE_DIR",
                str(active_path("default_image_directory", "default shorts image directory")),
            )
        ).expanduser()
    )
    snapshot_directory: Path = field(
        default_factory=lambda: Path(
            os.getenv("FLIPKART_SNAPSHOT_DIR", str(active_path("snapshot_directory", "snapshot directory")))
        ).expanduser()
    )
    price_stock_shipping_excel: Path = field(
        default_factory=lambda: Path(
            os.getenv(
                "PRICE_STOCK_SHIPPING_EXCEL",
                str(active_path("price_stock_shipping_excel", "Price/Stock/Shipping Excel")),
            )
        ).expanduser()
    )
    price_stock_shipping_json: Path = field(
        default_factory=lambda: Path(
            os.getenv(
                "PRICE_STOCK_SHIPPING_JSON",
                str(active_path("price_stock_shipping_json", "Price/Stock/Shipping JSON")),
            )
        ).expanduser()
    )
    product_description_excel: Path = field(
        default_factory=lambda: Path(
            os.getenv(
                "PRODUCT_DESCRIPTION_EXCEL",
                str(active_path("product_description_excel", "Jeans Product Description Excel")),
            )
        ).expanduser()
    )
    product_description_json: Path = field(
        default_factory=lambda: Path(
            os.getenv(
                "PRODUCT_DESCRIPTION_JSON",
                str(active_path("product_description_json", "Jeans Product Description JSON")),
            )
        ).expanduser()
    )
    additional_description_excel: Path = field(
        default_factory=lambda: Path(
            os.getenv(
                "ADDITIONAL_DESCRIPTION_EXCEL",
                str(active_path("additional_description_excel", "Jeans Additional Description Excel")),
            )
        ).expanduser()
    )
    additional_description_json: Path = field(
        default_factory=lambda: Path(
            os.getenv(
                "ADDITIONAL_DESCRIPTION_JSON",
                str(active_path("additional_description_json", "Jeans Additional Description JSON")),
            )
        ).expanduser()
    )
    variants_excel: Path = field(
        default_factory=lambda: Path(
            os.getenv("VARIANTS_EXCEL", str(active_path("variants_excel", "Variants Excel")))
        ).expanduser()
    )
    data_directory: Path = field(
        default_factory=lambda: Path(os.getenv("FLIPKART_DATA_DIR", "")).expanduser()
    )
    firefox_binary: str | None = os.getenv("FIREFOX_BINARY")
    geckodriver_path: str | None = os.getenv("GECKODRIVER_PATH")
    profile_name: str = "seema"
    headless: bool = os.getenv("HEADLESS", "0") == "1"

    @property
    def firefox_profile_path(self) -> Path:
        firefox_profiles = ACTIVE_LAPTOP_CONFIG["firefox_profiles"]
        if self.profile_name not in firefox_profiles:
            available_profiles = ", ".join(sorted(firefox_profiles))
            raise ValueError(
                f"Unknown Firefox profile '{self.profile_name}'. "
                f"Choose one of: {available_profiles}."
            )
        return resolve_profile_path(self.profile_name)


@dataclass(slots=True)
class ImageFolder:
    folder_path: Path
    folder_number: int
    exhausted_brand_surfaces: dict[str, set[str]]
    image_paths: list[Path]


@dataclass(slots=True)
class ProductInputRow:
    kind: str
    size: str
    values: dict[str, str]


@dataclass(slots=True)
class FieldDefinition:
    order: int
    label: str
    required: bool
    input_type: str
    locator_hint: str


@dataclass(slots=True)
class FillResult:
    generated_values: dict[str, str]
    skipped_fields: set[str]


@dataclass(slots=True)
class PendingImageFolderUse:
    image_folder: ImageFolder
    brand_name: str
    surface_name: str


@dataclass(slots=True)
class ListingSelection:
    product_type: str
    surface: str
    kind: str
    size: str
    brand_name: str
    image_directory: Path


@dataclass(slots=True)
class FlowTargetOption:
    product_type: str
    surface: str
    flow_directory: Path


def get_product_description_excel_path(product_type: str , surface_type: str) -> Path:
    if product_type == "jeans" and surface_type == "flipkart":
        return active_path("product_description_excel", "Jeans Product Description Excel")
    if product_type == "jeans" and surface_type == "shopsy":
        return JEANS_DATA_INPUTS_ROOT / "Product-Description-inputs-Shopsy.xlsx"
    if product_type == "trouser" and surface_type == "flipkart":
        return TROUSER_DATA_INPUTS_ROOT / "Product-Description-inputs.xlsx"
    if product_type == "trouser" and surface_type == "shopsy":
        return TROUSER_DATA_INPUTS_ROOT / "Product-Description-inputs-Shopsy.xlsx"
    return active_path("product_description_shorts_excel", "Shorts Product Description Excel")


def get_product_description_json_path(product_type: str, surface_type: str) -> Path:
    if product_type == "jeans" and surface_type == "flipkart":
        return active_path("product_description_json", "Jeans Product Description JSON")
    if product_type == "jeans" and surface_type == "shopsy":
        return PROJECT_ROOT / "json_LC_creation" / "jeans_shopsy" / "02_product_description.json"
    if product_type == "trouser" and surface_type == "flipkart":
        return PROJECT_ROOT / "json_LC_creation" / "trouser_flipkart" / "02_product_description.json"
    if product_type == "trouser" and surface_type == "shopsy":
        return PROJECT_ROOT / "json_LC_creation" / "trouser_shopsy" / "02_product_description.json"
    return active_path("product_description_shorts_json", "Shorts Product Description JSON")


def get_product_description_sheet_name(product_type: str, surface_type: str) -> str:
    if product_type == "jeans" and surface_type == "flipkart":
        return "Jeans Product Inputs"
    if product_type == "jeans" and surface_type == "shopsy":
        return "Jeans Product Inputs"
    if product_type == "trouser" and surface_type == "flipkart":
        return "Trouser Product Inputs"
    return "Shorts Product Inputs"


def get_additional_description_excel_path(product_type: str, surface_type: str) -> Path:
    if product_type == "jeans" and surface_type == "flipkart":
        return active_path("additional_description_excel", "Jeans Additional Description Excel")
    if product_type == "jeans" and surface_type == "shopsy":
        return JEANS_DATA_INPUTS_ROOT / "Additional-Description-inputs-Shopsy.xlsx"
    if product_type == "trouser" and surface_type == "flipkart":
        return TROUSER_DATA_INPUTS_ROOT / "Additional-Description-inputs.xlsx"
    if product_type == "trouser" and surface_type == "shopsy":
        return TROUSER_DATA_INPUTS_ROOT / "Additional-Description-inputs-Shopsy.xlsx"
    return active_path("additional_description_shorts_excel", "Shorts Additional Description Excel")


def get_additional_description_json_path(product_type: str, surface_type: str) -> Path:
    if product_type == "jeans" and surface_type == "flipkart":
        return active_path("additional_description_json", "Jeans Additional Description JSON")
    if product_type == "jeans" and surface_type == "shopsy":
        return PROJECT_ROOT / "json_LC_creation" / "jeans_shopsy" / "01_additional_description.json"
    if product_type == "trouser" and surface_type == "flipkart":
        return (
            PROJECT_ROOT
            / "json_LC_creation"
            / "trouser_flipkart"
            / "01_additional_description.json"
        )
    if product_type == "trouser" and surface_type == "shopsy":
        return (
            PROJECT_ROOT
            / "json_LC_creation"
            / "trouser_shopsy"
            / "01_additional_description.json"
        )
    return active_path("additional_description_shorts_json", "Shorts Additional Description JSON")


def get_additional_description_sheet_name(product_type: str, surface_type: str) -> str:
    if product_type == "jeans" and surface_type == "flipkart":
        return "Jeans Addl Desc Inputs"
    if product_type == "jeans" and surface_type == "shopsy":
        return "Jeans Addl Desc Inputs"
    if product_type == "trouser" and surface_type == "flipkart":
        return "Trouser Addl Desc Inputs"
    return "Shorts Addl Desc Inputs"


def get_variants_sheet_name(product_type: str) -> str:
    return "Jeans Variant Inputs"


class PauseController:
    def __init__(self) -> None:
        self.pause_requested = False
        self._stop_event = threading.Event()
        self._listener_thread: threading.Thread | None = None

    def start(self) -> None:
        if self._listener_thread is not None:
            return

        self._listener_thread = threading.Thread(target=self._listen_for_spacebar, daemon=True)
        self._listener_thread.start()

    def stop(self) -> None:
        self._stop_event.set()

    def pause_if_requested(
        self,
        stage_name: str,
        driver: webdriver.Firefox,
        config: BotConfig,
    ) -> None:
        if not self.pause_requested:
            return

        self.pause_requested = False
        snapshot_path = save_html_snapshot(driver, config.snapshot_directory, stage_name)
        log_event("PAUSE", f"Saved HTML snapshot: {snapshot_path}")
        input(f"[PAUSED] {stage_name}. Press Enter to continue...")

    def _listen_for_spacebar(self) -> None:
        while not self._stop_event.is_set():
            if msvcrt.kbhit():
                pressed_key = msvcrt.getwch()
                if pressed_key == " ":
                    self.pause_requested = True
                    log_event("PAUSE", "Pause requested. The bot will pause at the next safe step.")
            sleep(0.1)


def checkpoint_pause(
    pause_controller: PauseController,
    stage_name: str,
    driver: webdriver.Firefox,
    config: BotConfig,
) -> None:
    pause_controller.pause_if_requested(stage_name, driver, config)


def resolve_profile_name(selected_value: str) -> str:
    normalized_value = selected_value.strip().lower()
    if normalized_value in PROFILE_ALIASES:
        return PROFILE_ALIASES[normalized_value]

    available_profiles = ", ".join(sorted(ACTIVE_LAPTOP_CONFIG["firefox_profiles"]))
    raise ValueError(
        f"Unknown Firefox profile '{selected_value}'. Choose one of: {available_profiles}, s, p."
    )


def normalize_brand_name(brand_name: str) -> str:
    return " ".join(brand_name.strip().upper().split())


def normalize_field_label(label: str) -> str:
    cleaned = label.replace("*", " ")
    return " ".join(cleaned.strip().lower().split())


def parse_folder_number(folder_name: str) -> int:
    first_token = folder_name.split("-", maxsplit=1)[0].strip()
    if not first_token.isdigit():
        raise ValueError(f"Folder name must start with a number: {folder_name}")
    return int(first_token)


def parse_brand_folder_token(token: str, folder_name: str) -> tuple[str, set[str]]:
    normalized_token = token.strip().upper()
    token_match = re.fullmatch(r"([A-Z]+)('?[SF])?", normalized_token)
    if token_match is None:
        raise ValueError(f"Unknown brand code '{token}' in folder '{folder_name}'")

    brand_code, suffix = token_match.groups()
    if brand_code not in BRAND_CODE_MAP:
        raise ValueError(f"Unknown brand code '{token}' in folder '{folder_name}'")

    if not suffix:
        # Backward compatibility: old folder names without a suffix mean the brand
        # was already consumed for every supported surface.
        return BRAND_CODE_MAP[brand_code], set(SURFACE_FOLDER_SUFFIX)

    normalized_suffix = suffix if suffix.startswith("'") else f"'{suffix}"
    surface_name = FOLDER_SUFFIX_SURFACE.get(normalized_suffix.upper())
    if surface_name is None:
        raise ValueError(f"Unknown brand surface suffix '{token}' in folder '{folder_name}'")

    return BRAND_CODE_MAP[brand_code], {surface_name}


def parse_exhausted_brands(folder_name: str) -> dict[str, set[str]]:
    tokens = [part.strip().upper() for part in folder_name.split("-")[1:] if part.strip()]
    exhausted_brands: dict[str, set[str]] = {}

    for token in tokens:
        brand_name, surfaces = parse_brand_folder_token(token, folder_name)
        exhausted_brands.setdefault(brand_name, set()).update(surfaces)

    return exhausted_brands


def collect_ordered_images(folder_path: Path) -> list[Path]:
    image_candidates = [path for path in folder_path.iterdir() if path.is_file()]

    def image_sort_key(path: Path) -> tuple[int, str]:
        stem = path.stem.strip()
        if stem.isdigit():
            return (int(stem), path.name.lower())
        return (10**9, path.name.lower())

    return sorted(image_candidates, key=image_sort_key)


def load_image_folders(image_root: Path) -> list[ImageFolder]:
    if not image_root.exists():
        raise ValueError(f"Image directory does not exist: {image_root}")
    if not image_root.is_dir():
        raise ValueError(f"Image directory is not a folder: {image_root}")

    image_folders: list[ImageFolder] = []
    for folder_path in image_root.iterdir():
        if not folder_path.is_dir():
            continue

        image_folders.append(
            ImageFolder(
                folder_path=folder_path,
                folder_number=parse_folder_number(folder_path.name),
                exhausted_brand_surfaces=parse_exhausted_brands(folder_path.name),
                image_paths=collect_ordered_images(folder_path),
            )
        )

    return sorted(image_folders, key=lambda folder: folder.folder_number)


def choose_image_folder_for_brand(
    image_root: Path,
    brand_name: str,
    surface_name: str,
) -> ImageFolder | None:
    normalized_brand = normalize_brand_name(brand_name)

    for image_folder in load_image_folders(image_root):
        exhausted_surfaces = image_folder.exhausted_brand_surfaces.get(normalized_brand, set())
        if surface_name not in exhausted_surfaces:
            return image_folder

    return None


def get_brand_code(brand_name: str) -> str:
    normalized_brand = normalize_brand_name(brand_name)
    if normalized_brand not in BRAND_NAME_TO_CODE:
        raise ValueError(f"No short code configured for brand: {brand_name}")
    return BRAND_NAME_TO_CODE[normalized_brand]


def build_exhausted_folder_name(
    image_folder: ImageFolder,
    brand_name: str,
    surface_name: str,
) -> str:
    exhausted_brand_surfaces = {
        exhausted_brand: set(exhausted_surfaces)
        for exhausted_brand, exhausted_surfaces in image_folder.exhausted_brand_surfaces.items()
    }
    exhausted_brand_surfaces.setdefault(normalize_brand_name(brand_name), set()).add(surface_name)

    folder_tokens: list[str] = []
    for exhausted_brand in sorted(exhausted_brand_surfaces):
        brand_code = get_brand_code(exhausted_brand)
        for exhausted_surface in sorted(exhausted_brand_surfaces[exhausted_brand]):
            folder_tokens.append(f"{brand_code}{SURFACE_FOLDER_SUFFIX[exhausted_surface]}")

    return "-".join([str(image_folder.folder_number), *folder_tokens])


def mark_image_folder_exhausted(
    image_folder: ImageFolder,
    brand_name: str,
    surface_name: str,
) -> Path:
    new_folder_name = build_exhausted_folder_name(image_folder, brand_name, surface_name)
    new_folder_path = image_folder.folder_path.with_name(new_folder_name)

    if new_folder_path == image_folder.folder_path:
        log_event(
            "IMAGES",
            f"Image folder already marked for {brand_name}: {image_folder.folder_path.name}",
        )
        return image_folder.folder_path

    if new_folder_path.exists():
        raise ValueError(f"Cannot rename folder because target already exists: {new_folder_path}")

    image_folder.folder_path.rename(new_folder_path)
    image_folder.folder_path = new_folder_path
    image_folder.exhausted_brand_surfaces.setdefault(normalize_brand_name(brand_name), set()).add(
        surface_name
    )
    log_event("IMAGES", f"Renamed image folder to: {new_folder_path.name}")
    return new_folder_path


def queue_image_folder_exhaustion(
    flow_state: FlowState,
    image_folder: ImageFolder,
    brand_name: str,
    surface_name: str,
) -> None:
    flow_state.pending_image_folder_use = PendingImageFolderUse(
        image_folder=image_folder,
        brand_name=brand_name,
        surface_name=surface_name,
    )
    log_event(
        "IMAGES",
        "Image folder will be renamed only after Save & Go Back is clicked.",
    )


def commit_pending_image_folder_exhaustion(flow_state: FlowState) -> None:
    pending_use = flow_state.pending_image_folder_use
    if pending_use is None:
        return

    mark_image_folder_exhausted(
        pending_use.image_folder,
        pending_use.brand_name,
        pending_use.surface_name,
    )
    flow_state.pending_image_folder_use = None


def resolve_profile_path(profile_name: str) -> Path:
    firefox_profiles = ACTIVE_LAPTOP_CONFIG["firefox_profiles"]
    configured_path = firefox_profiles[profile_name]

    if configured_path is not None and configured_path.exists():
        return configured_path

    log_event(
        "PROFILE",
        f"Saved path for profile '{profile_name}' was not found: {configured_path}",
    )
    manual_path = input(
        f"Paste the actual Firefox profile directory for '{profile_name}': "
    ).strip().strip('"')

    if not manual_path:
        raise ValueError(f"No profile directory was provided for '{profile_name}'.")

    resolved_path = Path(manual_path).expanduser()
    if not resolved_path.exists():
        raise ValueError(f"Profile directory does not exist: {resolved_path}")

    firefox_profiles[profile_name] = resolved_path
    return resolved_path


def prompt_for_profile() -> str:
    env_profile = os.getenv("FIREFOX_PROFILE", "prabhu")
    prompt = f"Choose Firefox profile - seema (s) or prabhu (p) [{env_profile}]: "
    selected_value = input(prompt).strip()
    return resolve_profile_name(selected_value or env_profile)


# def prompt_for_additional_test_run() -> bool:
#     selected_value = input("Run Additional Description test flow only? (y/N): ").strip().lower()
#     return selected_value in {"y", "yes"}


def prompt_for_run_count() -> int:
    selected_value = input("How many runs should the bot execute? [5]: ").strip()
    run_count = int(selected_value or "5")
    if run_count < 1:
        raise ValueError("Run count must be at least 1.")
    return run_count


def get_brand_options_for_profile(profile_name: str) -> list[tuple[str, str]]:
    brand_codes = PROFILE_BRAND_CODES.get(profile_name, tuple(BRAND_CODE_MAP))
    unknown_codes = [brand_code for brand_code in brand_codes if brand_code not in BRAND_CODE_MAP]
    if unknown_codes:
        raise ValueError(
            f"Unknown brand code(s) configured for profile '{profile_name}': "
            f"{', '.join(unknown_codes)}"
        )
    return [(brand_code, BRAND_CODE_MAP[brand_code]) for brand_code in brand_codes]


def prompt_for_brand(profile_name: str) -> str:
    brand_options = get_brand_options_for_profile(profile_name)
    print(f"Choose brand for {profile_name} profile:")
    for index, (_, brand_name) in enumerate(brand_options, start=1):
        print(f"{index}. {brand_name}")

    default_index = next(
        (
            index
            for index, (_, brand_name) in enumerate(brand_options, start=1)
            if normalize_brand_name(brand_name) == normalize_brand_name(DEFAULT_BRAND_NAME)
        ),
        1,
    )
    selected_value = input(f"Enter option [{default_index}]: ").strip()
    selected_index = int(selected_value or str(default_index))
    if selected_index < 1 or selected_index > len(brand_options):
        raise ValueError(f"Please choose a valid brand option from 1 to {len(brand_options)}.")

    return brand_options[selected_index - 1][1]


def prompt_for_listing_selection(profile_name: str) -> ListingSelection:
    available_flow_targets = discover_flow_target_options()
    if not available_flow_targets:
        raise ValueError(f"No flow folders with flow.json were found in {FLOW_CONFIG_ROOT}")

    default_option_index = next(
        (
            index
            for index, option in enumerate(available_flow_targets, start=1)
            if option.product_type == "jeans" and option.surface == DEFAULT_FLOW_SURFACE
        ),
        1,
    )
    print("Choose listing flow:")
    for index, option in enumerate(available_flow_targets, start=1):
        print(f"{index}. {option.product_type} / {option.surface}")
    selected_flow_value = input(f"Enter option [{default_option_index}]: ").strip()
    selected_flow_index = int(selected_flow_value or str(default_option_index))
    if selected_flow_index < 1 or selected_flow_index > len(available_flow_targets):
        raise ValueError(
            f"Please choose a valid flow option from 1 to {len(available_flow_targets)}."
        )
    selected_flow = available_flow_targets[selected_flow_index - 1]
    product_type = selected_flow.product_type
    surface_type = selected_flow.surface

    if product_type == "jeans":
        DEFAULT_LISTING_SIZE = "28"
        print("Choose jeans kind:")
        print("1. Beige")
        print("2. Ice")
        print("3. Black-baggy")
        print("4. Black-Plain")
        print("5. White-Plain")
        kind_choice = input(f"Enter option [1]: ").strip() or "1"
        jeans_kind_options = ACTIVE_LAPTOP_CONFIG["jeans_kind_options"]
        if kind_choice not in jeans_kind_options:
            raise ValueError("Please choose a valid jeans option from 1 to 5.")
        selected_kind, configured_image_directory = jeans_kind_options[kind_choice]
        image_directory = require_configured_path(
            configured_image_directory,
            f"{selected_kind} jeans image directory",
        )
    elif product_type == "trouser" and surface_type == "flipkart":
        DEFAULT_LISTING_SIZE = "28"
        selected_kind = "Trouser"
        image_directory = require_configured_path(
            ACTIVE_LAPTOP_CONFIG.get("trouser_image_directory"),
            "default trouser image directory",
        )
    elif product_type == "trouser" and surface_type == "shopsy":
        DEFAULT_LISTING_SIZE = "M"
        selected_kind = "Trouser"
        image_directory = require_configured_path(
            ACTIVE_LAPTOP_CONFIG.get("trouser_image_directory"),
            "default trouser image directory",
        )

    size_value = input(f"Enter size [{DEFAULT_LISTING_SIZE}]: ").strip() or DEFAULT_LISTING_SIZE
    brand_name = prompt_for_brand(profile_name)
    return ListingSelection(
        product_type=product_type,
        surface=surface_type,
        kind=selected_kind,
        size=size_value,
        brand_name=brand_name,
        image_directory=image_directory,
    )


def build_firefox_driver(config: BotConfig) -> webdriver.Firefox:
    options = FirefoxOptions()

    if config.firefox_binary:
        options.binary_location = config.firefox_binary

    if config.headless:
        options.add_argument("-headless")

    # Use the on-disk Firefox profile directly to avoid Selenium/geckodriver
    # cloning it into a temporary profile on every startup.
    options.add_argument("-profile")
    options.add_argument(str(config.firefox_profile_path))

    if config.geckodriver_path:
        service = FirefoxService(executable_path=config.geckodriver_path)
        return webdriver.Firefox(service=service, options=options)

    # Selenium Manager can resolve the driver automatically in recent Selenium versions.
    return webdriver.Firefox(options=options)


def open_listing_page(driver: webdriver.Firefox, url: str) -> None:
    driver.maximize_window()
    driver.get(url)


def save_html_snapshot(driver: webdriver.Firefox, snapshot_directory: Path, stage_name: str) -> Path:
    snapshot_directory.mkdir(parents=True, exist_ok=True)
    safe_stage_name = "".join(
        character if character.isalnum() else "_" for character in stage_name.strip().lower()
    ).strip("_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    snapshot_path = snapshot_directory / f"{timestamp}_{safe_stage_name}.html"
    snapshot_path.write_text(driver.page_source, encoding="utf-8")
    return snapshot_path


def save_named_html_snapshot(
    driver: webdriver.Firefox,
    snapshot_directory: Path,
    file_name: str,
) -> Path:
    snapshot_directory.mkdir(parents=True, exist_ok=True)
    snapshot_path = snapshot_directory / file_name
    snapshot_path.write_text(driver.page_source, encoding="utf-8")
    return snapshot_path


def apply_surface_specific_field_aliases(
    row_values: dict[str, object],
    surface: str | None = None,
) -> dict[str, object]:
    resolved_values = dict(row_values)
    normalized_surface = (surface or "").strip().lower()
    if not normalized_surface:
        return resolved_values

    normalized_key_lookup = {
        key.strip().lower(): key
        for key in row_values
        if key is not None and str(key).strip()
    }

    surface_price_header = f"your selling price - {normalized_surface}"
    canonical_price_header = "Your selling price"
    surface_price_key = normalized_key_lookup.get(surface_price_header)
    if surface_price_key:
        resolved_values[canonical_price_header] = row_values.get(surface_price_key)

    return resolved_values


def load_product_input_row(
    workbook_path: Path,
    target_kind: str,
    target_size: str,
    worksheet_name: str | None = None,
    surface: str | None = None,
) -> ProductInputRow:
    if not workbook_path.exists():
        raise ValueError(f"Excel file was not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)
    if worksheet_name and worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]
    else:
        preferred_sheet = next(
            (sheet_name for sheet_name in workbook.sheetnames if "product inputs" in sheet_name.lower()),
            workbook.sheetnames[0],
        )
        worksheet = workbook[preferred_sheet]
    headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    normalized_headers = [
        str(header).strip() if header is not None else None
        for header in headers
    ]

    normalized_target_kind = target_kind.strip().lower()
    normalized_target_size = target_size.strip().lower()

    for row_index in range(2, worksheet.max_row + 1):
        row_values = {
            normalized_headers[column_index - 1]: worksheet.cell(row_index, column_index).value
            for column_index in range(1, worksheet.max_column + 1)
            if normalized_headers[column_index - 1]
        }
        row_values = apply_surface_specific_field_aliases(row_values, surface)
        row_kind = str(row_values.get("kind", "")).strip()
        row_size = str(row_values.get("size", "")).strip()

        if row_kind.lower() == normalized_target_kind and row_size.lower() == normalized_target_size:
            return ProductInputRow(
                kind=row_kind,
                size=row_size,
                values={key: "" if value is None else str(value).strip() for key, value in row_values.items()},
            )

    raise ValueError(
        f"No Excel row found for kind='{target_kind}' and size='{target_size}' in {workbook_path}"
    )


def load_field_definitions(json_path: Path) -> list[FieldDefinition]:
    if not json_path.exists():
        raise ValueError(f"Price/Stock/Shipping JSON file was not found: {json_path}")

    payload = json.loads(json_path.read_text(encoding="utf-8"))
    return [build_field_definition(field) for field in payload["fields"]]


def build_field_definition(field_payload: dict[str, object]) -> FieldDefinition:
    return FieldDefinition(
        order=int(field_payload["order"]),
        label=str(field_payload["label"]),
        required=bool(field_payload["required"]),
        input_type=str(field_payload["input_type"]),
        locator_hint=str(field_payload["locator_hint"]),
    )


def xpath_literal(value: str) -> str:
    if '"' not in value:
        return f'"{value}"'
    if "'" not in value:
        return f"'{value}'"
    parts = value.split('"')
    return 'concat(' + ', \'"\', '.join(f'"{part}"' for part in parts) + ')'


def generate_sku_suffix(length: int = 7) -> str:
    alphabet = string.ascii_uppercase + string.digits
    return "".join(random.choices(alphabet, k=length))


def build_sku_with_suffix(base_sku: str, shared_suffix: str) -> str:
    normalized_base_sku = str(base_sku).strip()
    if not normalized_base_sku:
        return ""
    return f"{normalized_base_sku}{shared_suffix}"


def extract_generated_sku_suffix(generated_sku: str, base_sku: str) -> str:
    normalized_generated_sku = str(generated_sku).strip()
    normalized_base_sku = str(base_sku).strip()
    if normalized_generated_sku and normalized_base_sku and normalized_generated_sku.startswith(normalized_base_sku):
        return normalized_generated_sku[len(normalized_base_sku):]
    return ""


def resolve_variant_shared_sku_suffix(
    config: BotConfig,
    listing_selection: ListingSelection,
    source_sku: str,
) -> str:
    try:
        source_price_row = load_product_input_row(
            config.price_stock_shipping_excel,
            listing_selection.kind,
            listing_selection.size,
            surface=listing_selection.surface,
        )
    except ValueError as exc:
        log_event(
            "VARIANT",
            "Could not load source Price/Stock/Shipping row to reuse its SKU suffix. "
            f"Generating a fresh suffix instead: {exc}",
        )
        return generate_sku_suffix()

    source_base_sku = source_price_row.values.get("Seller SKU ID", "").strip()
    suffix = extract_generated_sku_suffix(source_sku, source_base_sku)
    if suffix:
        log_event("VARIANT", f"Reusing generated SKU suffix from source listing: {suffix}")
        return suffix

    generated_suffix = generate_sku_suffix()
    log_event(
        "VARIANT",
        "Source SKU did not expose a reusable suffix from the Price workbook template. "
        f"Generated a fresh shared suffix instead: {generated_suffix}",
    )
    return generated_suffix


def normalize_field_value(value: str) -> str:
    normalized = str(value).strip()
    if not normalized:
        return ""

    try:
        numeric_value = float(normalized)
    except ValueError:
        return " ".join(normalized.split()).lower()

    if numeric_value.is_integer():
        return str(int(numeric_value))
    return ("%f" % numeric_value).rstrip("0").rstrip(".")


def find_matching_label_element(
    driver: webdriver.Firefox,
    field_label: str,
) -> WebElement | None:
    normalized_target = normalize_field_label(field_label)
    label_elements = driver.find_elements(
        By.XPATH,
        "//div[contains(@class,'AttributeItemLabelName')]",
    )
    exact_matches: list[WebElement] = []
    partial_matches: list[WebElement] = []
    for label_element in label_elements:
        try:
            normalized_text = normalize_field_label(label_element.text)
        except StaleElementReferenceException:
            continue
        if normalized_text == normalized_target:
            exact_matches.append(label_element)
        elif (
            normalized_target in normalized_text
            or normalized_text in normalized_target
        ):
            partial_matches.append(label_element)
    if exact_matches:
        return exact_matches[0]
    if partial_matches:
        return partial_matches[0]
    return None


def locate_field_label_element(
    driver: webdriver.Firefox,
    field_label: str,
    timeout_per_scroll: float = 3,
) -> WebElement:
    immediate_match = find_matching_label_element(driver, field_label)
    if immediate_match is not None:
        try:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", immediate_match)
            return immediate_match
        except StaleElementReferenceException:
            pass

    def wait_callback(_: webdriver.Firefox) -> WebElement | None:
        return find_matching_label_element(driver, field_label)

    for scroll_fraction in (0.35, 0.6, 0.85, 1.0):
        driver.execute_script(
            "window.scrollTo(0, document.body.scrollHeight * arguments[0]);",
            scroll_fraction,
        )
        try:
            label_element = WebDriverWait(driver, timeout_per_scroll).until(wait_callback)
            try:
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", label_element)
                return label_element
            except StaleElementReferenceException:
                continue
        except TimeoutException:
            continue

    raise TimeoutException(f"Could not find label element for field: {field_label}")


def get_field_wrapper(
    driver: webdriver.Firefox,
    field_label: str,
    timeout_per_scroll: float = 3,
) -> WebElement:
    label_element = locate_field_label_element(
        driver,
        field_label,
        timeout_per_scroll=timeout_per_scroll,
    )
    wrapper = driver.execute_script(
        """
        const label = arguments[0];
        const wrapperSelectors = [
            ".styles__FocusWrapper-sc-7uiywl-3",
            ".styles__EditAttributeItemWrapper-sc-gni56x-0",
            ".styles__AttributeItemFieldWrapper-sc-ske8mu-0",
        ];
        for (const selector of wrapperSelectors) {
            const candidate = label.closest(selector);
            if (candidate) {
                return candidate;
            }
        }
        let current = label.parentElement;
        while (current) {
            if (
                current.querySelector("input, [role='combobox'], #trigger-single-select")
                && current.textContent.includes(label.textContent)
            ) {
                return current;
            }
            current = current.parentElement;
        }
        return label.parentElement;
        """,
        label_element,
    )
    if wrapper is None:
        raise TimeoutException(f"Could not find field wrapper for label: {field_label}")
    return wrapper


def get_editable_field_element(
    driver: webdriver.Firefox,
    field_label: str,
    prefer_combobox: bool,
    timeout_per_scroll: float = 3,
) -> WebElement:
    last_stale_error: StaleElementReferenceException | None = None
    for _ in range(3):
        try:
            label_element = locate_field_label_element(
                driver,
                field_label,
                timeout_per_scroll=timeout_per_scroll,
            )
            field_element = driver.execute_script(
                """
                const label = arguments[0];
                const preferCombobox = arguments[1];
                const wrapperSelectors = [
                    ".styles__AttributeWrapper-sc-ske8mu-5",
                    ".styles__FocusWrapper-sc-7uiywl-3",
                    ".styles__EditAttributeItemWrapper-sc-gni56x-0",
                    ".styles__AttributeItemFieldWrapper-sc-ske8mu-0",
                ];

                function isEditableInput(element) {
                    if (!element || element.tagName !== "INPUT") {
                        return false;
                    }
                    if (element.type === "radio" || element.readOnly || element.disabled) {
                        return false;
                    }
                    return true;
                }

                function isEditableTextarea(element) {
                    if (!element || element.tagName !== "TEXTAREA") {
                        return false;
                    }
                    return !element.readOnly && !element.disabled;
                }

                function findControl(root) {
                    if (!root) {
                        return null;
                    }
                    if (preferCombobox) {
                        return root.querySelector("button[role='combobox'], [role='combobox']");
                    }
                    const inputs = Array.from(root.querySelectorAll("input"));
                    const textareas = Array.from(root.querySelectorAll("textarea"));
                    return inputs.find(isEditableInput) || textareas.find(isEditableTextarea) || null;
                }

                for (const selector of wrapperSelectors) {
                    const wrapper = label.closest(selector);
                    const control = findControl(wrapper);
                    if (control) {
                        return control;
                    }
                }

                let current = label.parentElement;
                while (current) {
                    const control = findControl(current);
                    if (control) {
                        return control;
                    }
                    current = current.parentElement;
                }
                return null;
                """,
                label_element,
                prefer_combobox,
            )
            if field_element is None:
                raise TimeoutException(f"Could not find editable element for label: {field_label}")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", field_element)
            return field_element
        except StaleElementReferenceException as error:
            last_stale_error = error
            sleep(0.2)

    raise StaleElementReferenceException(
        f"Editable element for field '{field_label}' kept going stale."
    ) from last_stale_error


def set_input_value(
    driver: webdriver.Firefox,
    input_element: WebElement,
    field_value: str,
    field_label: str | None = None,
) -> None:
    def reacquire_input() -> WebElement:
        if not field_label:
            return input_element
        return get_editable_field_element(
            driver,
            field_label,
            prefer_combobox=False,
            timeout_per_scroll=1,
        )

    def handle_length_field_bug() -> bool:
        if field_label != "Length" or not field_value:
            return False

        length_input = reacquire_input()
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", length_input)
        driver.execute_script("arguments[0].focus();", length_input)
        sleep(0.05)
        length_input.send_keys(Keys.CONTROL, "a")
        sleep(0.05)
        length_input.send_keys(Keys.DELETE)
        sleep(0.05)
        length_input.send_keys(field_value[0])
        log_event(
            "FORM",
            "Applied Length-field bug workaround: typed the first character to trigger the blur behavior.",
        )
        sleep(0.15)
        ActionChains(driver).send_keys(Keys.TAB).perform()
        sleep(0.15)

        length_input = reacquire_input()
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", length_input)
        driver.execute_script("arguments[0].focus();", length_input)
        sleep(0.05)
        length_input.send_keys(Keys.CONTROL, "a")
        sleep(0.05)
        length_input.send_keys(Keys.DELETE)
        sleep(0.05)
        length_input.send_keys(field_value)
        sleep(0.1)
        current_value = (length_input.get_attribute("value") or "").strip()
        if current_value == field_value:
            log_event(
                "FORM",
                f"Length workaround completed successfully with final value: {field_value}",
            )
            return True
        return False

    def paste_value_via_clipboard(active_input: WebElement) -> bool:
        if field_label != "Description" or len(field_value) < 120:
            return False

        try:
            driver.execute_script(
                """
                const element = arguments[0];
                const value = arguments[1];
                const prototype = element instanceof HTMLTextAreaElement
                    ? HTMLTextAreaElement.prototype
                    : HTMLInputElement.prototype;
                const valueSetter = Object.getOwnPropertyDescriptor(prototype, "value")?.set;

                element.scrollIntoView({block: "center", inline: "nearest"});
                element.focus();
                if (valueSetter) {
                    valueSetter.call(element, value);
                } else {
                    element.value = value;
                }
                element.dispatchEvent(new InputEvent("input", {bubbles: true, inputType: "insertText", data: value}));
                element.dispatchEvent(new Event("change", {bubbles: true}));
                element.blur();
                """,
                active_input,
                field_value,
            )
            log_event("FORM", "Set long Description text via browser input/change events.")
            return True
        except WebDriverException as exc:
            log_event(
                "FORM",
                f"Browser event-based Description fill failed, trying clipboard paste fallback: {exc.__class__.__name__}",
            )

        original_clipboard_text: str | None = None
        clipboard_root = tk.Tk()
        clipboard_root.withdraw()
        try:
            try:
                original_clipboard_text = clipboard_root.clipboard_get()
            except tk.TclError:
                original_clipboard_text = None

            clipboard_root.clipboard_clear()
            clipboard_root.clipboard_append(field_value)
            clipboard_root.update()

            click_element_via_autogui(driver, active_input, "Description input")
            sleep(0.1)
            pyautogui.hotkey("ctrl", "a")
            sleep(0.05)
            pyautogui.press("delete")
            sleep(0.05)
            pyautogui.hotkey("ctrl", "v")
            sleep(1.0)
            log_event(
                "FORM",
                "Pasted long Description text via pyautogui. "
                "Skipping immediate Selenium read-back to avoid WebDriver stalls.",
            )
            return True
        finally:
            clipboard_root.clipboard_clear()
            if original_clipboard_text:
                clipboard_root.clipboard_append(original_clipboard_text)
            clipboard_root.update()
            clipboard_root.destroy()

    if handle_length_field_bug():
        return

    active_input = input_element
    for attempt in range(3):
        try:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", active_input)
            driver.execute_script("arguments[0].focus();", active_input)
            sleep(0.05)
            active_input.send_keys(Keys.CONTROL, "a")
            sleep(0.05)
            active_input.send_keys(Keys.DELETE)
            sleep(0.05)

            if paste_value_via_clipboard(active_input):
                return

            for character in field_value:
                try:
                    active_input.send_keys(character)
                except WebDriverException:
                    active_input = reacquire_input()
                    driver.execute_script("arguments[0].focus();", active_input)
                    sleep(0.02)
                    current_partial = (active_input.get_attribute("value") or "").strip()
                    remaining_value = field_value[len(current_partial):]
                    if remaining_value:
                        active_input.send_keys(remaining_value)
                    break
                sleep(0.02)

            current_value = (active_input.get_attribute("value") or "").strip()
            if current_value != field_value:
                active_input = reacquire_input()
                active_input.send_keys(Keys.CONTROL, "a")
                sleep(0.05)
                active_input.send_keys(Keys.DELETE)
                sleep(0.05)
                active_input.send_keys(field_value)
                current_value = (active_input.get_attribute("value") or "").strip()

            if current_value == field_value:
                return
        except WebDriverException:
            if attempt == 2:
                raise
            sleep(0.15)
            active_input = reacquire_input()

    current_value = (active_input.get_attribute("value") or "").strip()
    raise TimeoutException(
        f"Typed value did not stick. Expected '{field_value}', got '{current_value}'."
    )


def get_field_current_value(
    driver: webdriver.Firefox,
    field_label: str,
    timeout_per_scroll: float = 0.35,
) -> str:
    input_element = get_editable_field_element(
        driver,
        field_label,
        prefer_combobox=False,
        timeout_per_scroll=timeout_per_scroll,
    )
    return (input_element.get_attribute("value") or "").strip()


def normalize_field_value(value: str) -> str:
    return " ".join((value or "").strip().split()).lower()


def field_values_match(current_value: str, expected_value: str) -> bool:
    return normalize_field_value(current_value) == normalize_field_value(expected_value)


def get_combobox_current_value(
    driver: webdriver.Firefox,
    field_label: str,
    timeout_per_scroll: float = 0.35,
) -> str:
    combobox = get_editable_field_element(
        driver,
        field_label,
        prefer_combobox=True,
        timeout_per_scroll=timeout_per_scroll,
    )
    button_text = combobox.text.strip()
    if button_text:
        return button_text
    return (combobox.get_attribute("value") or "").strip()


def fill_text_or_number_field(
    driver: webdriver.Firefox,
    field_label: str,
    field_value: str,
    timeout_per_scroll: float = 3,
) -> None:
    current_value = get_field_current_value(
        driver,
        field_label,
        timeout_per_scroll=min(timeout_per_scroll, 1),
    )
    if field_values_match(current_value, field_value):
        log_event("FORM", f"Skipping {field_label}: already filled with expected value '{current_value}'.")
        return

    input_element = get_editable_field_element(
        driver,
        field_label,
        prefer_combobox=False,
        timeout_per_scroll=timeout_per_scroll,
    )
    set_input_value(driver, input_element, field_value, field_label=field_label)
    log_event("FORM", f"Filled {field_label}: {field_value}")


def click_field_by_mouse(
    driver: webdriver.Firefox,
    field_label: str,
    timeout_per_scroll: float = 1,
) -> None:
    input_element = get_editable_field_element(
        driver,
        field_label,
        prefer_combobox=False,
        timeout_per_scroll=timeout_per_scroll,
    )
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", input_element)
    click_element_resilient(driver, input_element)
    log_event("FORM", f"Clicked {field_label} field by mouse.")


def click_screen_point_multiple_times(
    x: int,
    y: int,
    click_count: int = 4,
    pause_between_clicks: float = 0.2,
) -> None:
    mouse = MouseController()
    original_position = mouse.position
    log_event(
        "MOUSE",
        f"Preparing screen-pixel click sequence at ({x}, {y}) for {click_count} click(s).",
    )
    pyautogui.moveTo(x, y, duration=0.15)
    mouse.position = (x, y)
    for click_index in range(1, click_count + 1):
        mouse.click(PynputButton.left, 1)
        log_event(
            "MOUSE",
            f"Clicked screen pixel ({x}, {y}) [{click_index}/{click_count}].",
        )
        sleep(pause_between_clicks)
    mouse.position = original_position
    log_event(
        "MOUSE",
        f"Restored mouse position to ({int(original_position[0])}, {int(original_position[1])}).",
    )


def get_element_screen_center(driver: webdriver.Firefox, element: WebElement) -> tuple[int, int]:
    geometry = driver.execute_script(
        """
        const element = arguments[0];
        const rect = element.getBoundingClientRect();
        return {
            left: rect.left,
            top: rect.top,
            width: rect.width,
            height: rect.height,
            screenX: window.screenX,
            screenY: window.screenY,
            outerWidth: window.outerWidth,
            outerHeight: window.outerHeight,
            innerWidth: window.innerWidth,
            innerHeight: window.innerHeight,
            devicePixelRatio: window.devicePixelRatio || 1
        };
        """,
        element,
    )
    device_pixel_ratio = float(geometry["devicePixelRatio"] or 1)
    horizontal_chrome = max(
        0.0,
        (float(geometry["outerWidth"]) - float(geometry["innerWidth"])) / 2.0,
    )
    vertical_chrome = max(
        0.0,
        float(geometry["outerHeight"]) - float(geometry["innerHeight"]) - horizontal_chrome,
    )
    center_x_css = float(geometry["screenX"]) + horizontal_chrome + float(geometry["left"]) + (
        float(geometry["width"]) / 2.0
    )
    center_y_css = float(geometry["screenY"]) + vertical_chrome + float(geometry["top"]) + (
        float(geometry["height"]) / 2.0
    )
    return (
        int(round(center_x_css * device_pixel_ratio)),
        int(round(center_y_css * device_pixel_ratio)),
    )


def click_element_via_autogui(driver: webdriver.Firefox, element: WebElement, label: str) -> None:
    driver.execute_script(
        "arguments[0].scrollIntoView({block: 'center', inline: 'nearest'});",
        element,
    )
    sleep(0.1)
    screen_x, screen_y = get_element_screen_center(driver, element)
    log_event(
        "MOUSE",
        f"Clicking {label} via screen coordinates at ({screen_x}, {screen_y}).",
    )
    click_screen_point_multiple_times(
        screen_x,
        screen_y,
        click_count=1,
        pause_between_clicks=0.05,
    )


def click_checkbox_option_via_autogui(
    driver: webdriver.Firefox,
    checkbox_input: WebElement,
    label: str,
) -> None:
    clickable_target = driver.execute_script(
        """
        const checkbox = arguments[0];
        return checkbox.closest('.styles__InputCheckboxWrapper-sc-qliyra-0') ||
            checkbox.parentElement ||
            checkbox;
        """,
        checkbox_input,
    )
    if clickable_target is None:
        clickable_target = checkbox_input
    click_element_via_autogui(driver, clickable_target, label)


def fill_combobox_field(
    driver: webdriver.Firefox,
    field_label: str,
    field_value: str,
    timeout_per_scroll: float = 3,
) -> None:
    current_value = get_combobox_current_value(
        driver,
        field_label,
        timeout_per_scroll=min(timeout_per_scroll, 1),
    )
    if field_values_match(current_value, field_value):
        log_event("FORM", f"Skipping {field_label}: already selected as '{current_value}'.")
        return

    combobox = get_editable_field_element(
        driver,
        field_label,
        prefer_combobox=True,
        timeout_per_scroll=timeout_per_scroll,
    )
    select_combobox_option(driver, combobox, field_value, field_label)
    log_event("FORM", f"Selected {field_label}: {field_value}")


def click_element_resilient(driver: webdriver.Firefox, element: WebElement) -> None:
    driver.execute_script(
        "arguments[0].scrollIntoView({block: 'center', inline: 'nearest'});",
        element,
    )
    try:
        element.click()
        return
    except WebDriverException:
        pass

    driver.execute_script("arguments[0].click();", element)


def click_element_without_js(driver: webdriver.Firefox, element: WebElement) -> None:
    driver.execute_script(
        "arguments[0].scrollIntoView({block: 'center', inline: 'nearest'});",
        element,
    )
    try:
        element.click()
        return
    except WebDriverException:
        pass

    ActionChains(driver).move_to_element(element).pause(0.05).click().perform()


def get_visible_dropdown_container(
    driver: webdriver.Firefox,
    combobox: WebElement | None = None,
    timeout_seconds: float = 2,
) -> WebElement:
    deadline = datetime.now().timestamp() + timeout_seconds

    while datetime.now().timestamp() < deadline:
        if combobox is not None:
            try:
                local_container = driver.execute_script(
                    """
                    const combobox = arguments[0];
                    const controlledId = (combobox.getAttribute('aria-controls') || '').trim();

                    function isVisible(element) {
                        if (!element) {
                            return false;
                        }
                        const style = window.getComputedStyle(element);
                        const rect = element.getBoundingClientRect();
                        return (
                            style.display !== 'none' &&
                            style.visibility !== 'hidden' &&
                            rect.width > 0 &&
                            rect.height > 0
                        );
                    }

                    const localRoot =
                        combobox.closest("[class*='SingleSelectContainer']") ||
                        combobox.closest("[class*='MultiSelect']") ||
                        combobox.parentElement;
                    if (localRoot) {
                        const localContainers = Array.from(
                            localRoot.querySelectorAll(
                                "[data-testid='content-single-select'], [data-testid='content-multi-select'], " +
                                ".styles__DropdownContent-sc-zkytp-1, .styles__DropdownContent-sc-lf8o9y-2"
                            )
                        ).filter(isVisible);
                        if (localContainers.length) {
                            return localContainers[0];
                        }
                    }

                    const candidates = Array.from(
                        document.querySelectorAll(
                            "[data-testid='content-single-select'], [data-testid='content-multi-select'], " +
                            ".styles__DropdownContent-sc-zkytp-1, .styles__DropdownContent-sc-lf8o9y-2"
                        )
                    ).filter(isVisible);
                    if (!candidates.length) {
                        return null;
                    }

                    const comboRect = combobox.getBoundingClientRect();
                    const comboCenterX = comboRect.left + comboRect.width / 2;
                    const comboBottom = comboRect.bottom;
                    candidates.sort((left, right) => {
                        function score(element) {
                            const rect = element.getBoundingClientRect();
                            const centerX = rect.left + rect.width / 2;
                            const verticalGap = Math.abs(rect.top - comboBottom);
                            const horizontalGap = Math.abs(centerX - comboCenterX);
                            const idPenalty = controlledId && element.id === controlledId ? 0 : 500;
                            return verticalGap + horizontalGap + idPenalty;
                        }
                        return score(left) - score(right);
                    });
                    return candidates[0];
                    """,
                    combobox,
                )
                if local_container is not None:
                    return local_container
            except WebDriverException:
                pass

        dropdown_candidates = driver.find_elements(
            By.CSS_SELECTOR,
            "[data-testid='content-single-select'], [data-testid='content-multi-select'], "
            ".styles__DropdownContent-sc-zkytp-1, .styles__DropdownContent-sc-lf8o9y-2",
        )
        for candidate in dropdown_candidates:
            try:
                if candidate.is_displayed():
                    return candidate
            except WebDriverException:
                continue
        sleep(0.1)

    raise TimeoutException("Could not find visible dropdown container after opening combobox.")


def close_dropdown_with_escape(driver: webdriver.Firefox, field_label: str) -> None:
    log_event("FORM", f"Closing dropdown for {field_label} with Escape.")
    ActionChains(driver).send_keys(Keys.ESCAPE).perform()
    sleep(0.15)


def select_combobox_option(
    driver: webdriver.Firefox,
    combobox: WebElement,
    field_value: str,
    field_label: str,
) -> None:
    click_element_without_js(driver, combobox)
    sleep(0.15)
    try:
        dropdown_container = get_visible_dropdown_container(driver, combobox=combobox)
    except TimeoutException:
        click_element_without_js(driver, combobox)
        sleep(0.2)
        dropdown_container = get_visible_dropdown_container(driver, combobox=combobox)

    search_inputs = dropdown_container.find_elements(
        By.CSS_SELECTOR,
        "input[aria-label='Search'], input[placeholder='Select'], input[type='search'], input[type='text']",
    )
    visible_search_input = next(
        (element for element in search_inputs if element.is_displayed() and element.is_enabled()),
        None,
    )
    if visible_search_input is not None:
        log_event("FORM", f"Typing into dropdown search for {field_label}: {field_value}")
        driver.execute_script("arguments[0].focus();", visible_search_input)
        sleep(0.05)
        visible_search_input.send_keys(Keys.CONTROL, "a")
        sleep(0.05)
        visible_search_input.send_keys(Keys.DELETE)
        sleep(0.05)
        for character in field_value:
            visible_search_input.send_keys(character)
            sleep(0.02)
        sleep(0.25)

    normalized_target = field_value.strip().lower()
    matching_label = None

    for _ in range(10):
        option_labels = dropdown_container.find_elements(
            By.CSS_SELECTOR,
            "label[for], .style__LabelWrapper-sc-n7qfg8-0, [role='radio'] label",
        )
        for option_label in option_labels:
            try:
                if not option_label.is_displayed():
                    continue
                option_text = " ".join(option_label.text.strip().split()).lower()
            except WebDriverException:
                continue
            if option_text == normalized_target:
                matching_label = option_label
                break
        if matching_label is not None:
            break
        sleep(0.1)

    if matching_label is None:
        option_labels = dropdown_container.find_elements(
            By.CSS_SELECTOR,
            "label[for], .style__LabelWrapper-sc-n7qfg8-0, [role='radio'] label",
        )
        available_options = []
        for option_label in option_labels:
            try:
                if option_label.is_displayed() and option_label.text.strip():
                    available_options.append(" ".join(option_label.text.strip().split()))
            except WebDriverException:
                continue
        raise TimeoutException(
            f"Could not find dropdown option '{field_value}' for {field_label}. "
            f"Available options: {available_options}"
        )

    # Click the radio input directly via JS so React synthetic events fire.
    # Clicking only the label via pyautogui does not trigger React onChange
    # on readonly radio inputs, leaving the Create button permanently disabled.
    clicked_radio = driver.execute_script(
        """
        const label = arguments[0];
        const forAttr = label.getAttribute('for');
        const radio = forAttr
            ? document.getElementById(forAttr)
            : label.closest('[class*="CheckMarkOptionWrapper"]')?.querySelector('input[type="radio"]');
        if (radio) { radio.click(); return true; }
        label.click();
        return false;
        """,
        matching_label,
    )
    log_event("FORM", f"Clicked {field_label} option '{field_value}' via JS radio click (radio={clicked_radio}).")
    sleep(0.3)
    close_dropdown_with_escape(driver, field_label)


def get_tag_input_wrapper(driver: webdriver.Firefox, field_label: str) -> WebElement:
    wrapper = get_field_wrapper(driver, field_label, timeout_per_scroll=1)
    tag_wrapper = driver.execute_script(
        """
        const wrapper = arguments[0];
        return (
            wrapper.querySelector('.rti--container') ||
            wrapper.querySelector('.multi-select-field-wrapper') ||
            wrapper.querySelector("[data-testid='trigger-multi-select']")?.closest('.multi-select-field-wrapper') ||
            wrapper.querySelector("[data-testid='trigger-multi-select']") ||
            wrapper
        );
        """,
        wrapper,
    )
    if tag_wrapper is None:
        raise TimeoutException(f"Could not find tag input wrapper for field: {field_label}")
    return tag_wrapper


def get_tag_input_current_value(driver: webdriver.Firefox, field_label: str) -> str:
    tag_wrapper = get_tag_input_wrapper(driver, field_label)
    values = driver.execute_script(
        """
        const wrapper = arguments[0];
        const helperText = wrapper.parentElement?.querySelector('.styles__MultiSelectHelperText-sc-ske8mu-16');
        if (helperText && helperText.textContent.trim()) {
            return helperText.textContent
                .split(',')
                .map((value) => value.trim())
                .filter(Boolean);
        }

        const checkedCheckboxValues = Array.from(
            wrapper.querySelectorAll("input[type='checkbox']:checked")
        )
            .map((element) => element.value || element.getAttribute('data-label') || '')
            .filter(Boolean);
        if (checkedCheckboxValues.length) {
            return checkedCheckboxValues;
        }

        return Array.from(wrapper.querySelectorAll('[role="tab"][label], [role="tab"][value]'))
            .map((element) => element.getAttribute('label') || element.getAttribute('value') || element.textContent.trim())
            .filter(Boolean);
        """,
        tag_wrapper,
    )
    return ", ".join(values)


def fill_tag_input_field(
    driver: webdriver.Firefox,
    field_label: str,
    field_value: str,
    commit_each_value_with_enter: bool = True,
) -> None:
    desired_values = [value.strip() for value in field_value.split(",") if value.strip()]
    current_value = get_tag_input_current_value(driver, field_label)
    current_values = [value.strip() for value in current_value.split(",") if value.strip()]
    normalized_current_values = {normalize_field_value(value) for value in current_values}
    missing_values = [
        value for value in desired_values
        if normalize_field_value(value) not in normalized_current_values
    ]
    if not missing_values:
        log_event(
            "FORM",
            f"Skipping {field_label}: all desired value(s) are already present in '{current_value}'.",
        )
        return

    tag_wrapper = get_tag_input_wrapper(driver, field_label)
    existing_tag_elements = tag_wrapper.find_elements(
        By.CSS_SELECTOR,
        "[role='tab'][label], [role='tab'][value]",
    )
    multi_select_combobox = driver.execute_script(
        """
        const wrapper = arguments[0];
        return wrapper.matches("[data-testid='trigger-multi-select'], button[role='combobox']")
            ? wrapper
            : wrapper.querySelector("[data-testid='trigger-multi-select'], button[role='combobox']");
        """,
        tag_wrapper,
    )
    if multi_select_combobox is not None:
        def open_multi_select_dropdown() -> WebElement:
            click_element_without_js(driver, multi_select_combobox)
            sleep(0.2)
            try:
                return get_visible_dropdown_container(
                    driver,
                    combobox=multi_select_combobox,
                )
            except TimeoutException:
                click_element_without_js(driver, multi_select_combobox)
                sleep(0.2)
                return get_visible_dropdown_container(
                    driver,
                    combobox=multi_select_combobox,
                )

        def get_search_input(dropdown_container: WebElement) -> WebElement:
            search_input = next(
                (
                    element
                    for element in dropdown_container.find_elements(
                        By.CSS_SELECTOR,
                        "input[id*='checkbox-tree'][type='text'], input[placeholder='Search Paramter'], input[aria-label='Search']",
                    )
                    if element.is_displayed() and element.is_enabled()
                ),
                None,
            )
            if search_input is None:
                raise TimeoutException(f"Could not find Brand Color search input for field: {field_label}")
            return search_input

        for desired_value in missing_values:
            selection_stuck = False
            normalized_target = normalize_field_value(desired_value)

            for attempt in range(2):
                dropdown_container = open_multi_select_dropdown()
                search_input = get_search_input(dropdown_container)

                log_event("FORM", f"Typing into multi-select search for {field_label}: {desired_value}")
                driver.execute_script("arguments[0].focus();", search_input)
                sleep(0.05)
                search_input.send_keys(Keys.CONTROL, "a")
                sleep(0.05)
                search_input.send_keys(Keys.DELETE)
                sleep(0.05)
                for character in desired_value:
                    search_input.send_keys(character)
                    sleep(0.02)
                sleep(0.25)

                matching_label = None
                for _ in range(10):
                    option_labels = dropdown_container.find_elements(
                        By.CSS_SELECTOR,
                        "label[for], .style__LabelWrapper-sc-n7qfg8-0",
                    )
                    for option_label in option_labels:
                        try:
                            if not option_label.is_displayed():
                                continue
                            option_text = normalize_field_value(option_label.text)
                        except WebDriverException:
                            continue
                        if option_text == normalized_target:
                            matching_label = option_label
                            break
                    if matching_label is not None:
                        break
                    sleep(0.1)

                if matching_label is None:
                    raise TimeoutException(
                        f"Could not find multi-select option '{desired_value}' for {field_label}."
                    )

                checkbox_id = matching_label.get_attribute("for")
                if checkbox_id:
                    checkbox = dropdown_container.find_element(By.ID, checkbox_id)
                    if checkbox.is_selected():
                        selection_stuck = True
                    else:
                        try:
                            click_checkbox_option_via_autogui(
                                driver,
                                checkbox,
                                f"{field_label} checkbox '{desired_value}'",
                            )
                        except Exception:
                            try:
                                click_element_without_js(driver, checkbox)
                            except Exception:
                                click_element_without_js(driver, matching_label)
                        sleep(0.15)
                        try:
                            checkbox = dropdown_container.find_element(By.ID, checkbox_id)
                            if checkbox.is_selected():
                                selection_stuck = True
                        except WebDriverException:
                            pass
                else:
                    try:
                        click_element_via_autogui(driver, matching_label, f"{field_label} option '{desired_value}'")
                    except Exception:
                        click_element_without_js(driver, matching_label)
                    sleep(0.15)

                close_dropdown_with_escape(driver, field_label)
                sleep(0.2)
                current_multi_value = get_tag_input_current_value(driver, field_label)
                current_multi_values = {
                    normalize_field_value(value)
                    for value in current_multi_value.split(",")
                    if value.strip()
                }
                if normalized_target in current_multi_values:
                    selection_stuck = True
                    break
                log_event(
                    "FORM",
                    f"{field_label} option '{desired_value}' did not read back after attempt {attempt + 1}. Retrying...",
                )

            if not selection_stuck:
                raise TimeoutException(
                    f"Selected multi-select option '{desired_value}' for {field_label}, "
                    "but it did not appear in the field state afterward."
                )

        log_event("FORM", f"Filled {field_label}: added missing value(s) {', '.join(missing_values)}")
        return

    input_element = driver.execute_script(
        """
        const wrapper = arguments[0];
        return wrapper.querySelector('.rti--input, input:not([type="radio"]):not([type="checkbox"])');
        """,
        tag_wrapper,
    )
    if input_element is None:
        if existing_tag_elements:
            log_event(
                "FORM",
                f"Skipping {field_label}: tag input is not currently editable. Existing values are "
                f"'{current_value}', missing desired values are '{', '.join(missing_values)}'.",
            )
            return
        raise TimeoutException(f"Could not find tag input element for field: {field_label}")

    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", input_element)
    driver.execute_script("arguments[0].focus();", input_element)
    try:
        input_element.click()
    except WebDriverException:
        pass

    if commit_each_value_with_enter:
        for value in missing_values:
            input_element.send_keys(value)
            input_element.send_keys(Keys.ENTER)
            sleep(0.05)
    else:
        # Keep the old Enter-per-value path above for later use.
        # For now, these fields accept comma-separated input from Excel directly,
        # and the page treats commas the same as Enter while typing.
        input_element.send_keys(", ".join(missing_values))

        # Old commit path intentionally retained:
        # for value in missing_values:
        #     input_element.send_keys(value)
        #     input_element.send_keys(Keys.ENTER)
        #     sleep(0.05)

    input_element.send_keys(Keys.ESCAPE)
    sleep(0.05)

    log_event("FORM", f"Filled {field_label}: added missing value(s) {', '.join(missing_values)}")


def fill_price_stock_shipping_fields(
    driver: webdriver.Firefox,
    field_definitions: list[FieldDefinition],
    product_input_row: ProductInputRow,
) -> FillResult:
    generated_values: dict[str, str] = {}
    skipped_fields: set[str] = set()
    handling_fee_fields_unavailable = False

    for field in field_definitions:
        if handling_fee_fields_unavailable and field.label in {
            "Zonal handling fee",
            "National handling fee",
        }:
            skipped_fields.add(field.label)
            log_event(
                "PRICE",
                f"Skipping {field.label}: handling fee fields are not present in current listing state.",
            )
            continue

        raw_value = product_input_row.values.get(field.label, "").strip()
        if not raw_value:
            log_event("PRICE", f"Skipping {field.label}: no Excel value provided.")
            continue

        if field.label == "Seller SKU ID":
            raw_value = f"{raw_value}{generate_sku_suffix()}"
            generated_values[field.label] = raw_value
            log_event("PRICE", f"Generated Seller SKU ID value: {raw_value}")

        timeout_per_scroll = 3 if field.required else 0.25
        try:
            if field.input_type == "combobox":
                fill_combobox_field(
                    driver,
                    field.label,
                    raw_value,
                    timeout_per_scroll=timeout_per_scroll,
                )
            else:
                fill_text_or_number_field(
                    driver,
                    field.label,
                    raw_value,
                    timeout_per_scroll=timeout_per_scroll,
                )
        except TimeoutException:
            if field.required:
                raise
            if field.label == "Local handling fee":
                handling_fee_fields_unavailable = True
                skipped_fields.update(
                    {"Local handling fee", "Zonal handling fee", "National handling fee"}
                )
                log_event(
                    "PRICE",
                    "Skipping Local handling fee: field not present in current listing state. "
                    "Assuming Zonal and National handling fee fields are also absent.",
                )
                continue
            skipped_fields.add(field.label)
            log_event("PRICE", f"Skipping {field.label}: field not present in current listing state.")

    generated_sku_value = generated_values.get("Seller SKU ID", "")
    if generated_sku_value:
        try:
            current_sku_value = get_field_current_value(driver, "Seller SKU ID")
            if current_sku_value != generated_sku_value:
                log_event(
                    "PRICE",
                    "Seller SKU ID changed after later field updates. "
                    f"Re-applying generated value: {generated_sku_value}",
                )
                fill_text_or_number_field(driver, "Seller SKU ID", generated_sku_value)
            else:
                log_event("PRICE", f"Seller SKU ID persisted: {generated_sku_value}")
        except TimeoutException:
            log_event("PRICE", "Could not re-check Seller SKU ID at the end of fill.")

    return FillResult(generated_values=generated_values, skipped_fields=skipped_fields)


def build_expected_field_values(
    field_definitions: list[FieldDefinition],
    product_input_row: ProductInputRow,
    fill_result: FillResult,
) -> dict[str, str]:
    expected_values: dict[str, str] = {}
    for field in field_definitions:
        if field.label in fill_result.skipped_fields:
            continue

        if field.label in fill_result.generated_values:
            expected_values[field.label] = fill_result.generated_values[field.label]
            continue

        raw_value = product_input_row.values.get(field.label, "").strip()
        if raw_value:
            expected_values[field.label] = raw_value
    return expected_values


def verify_and_refill_price_stock_shipping_fields(
    driver: webdriver.Firefox,
    field_definitions: list[FieldDefinition],
    expected_values: dict[str, str],
    skipped_fields: set[str],
    max_passes: int = 2,
) -> None:
    for verification_pass in range(1, max_passes + 1):
        mismatches: list[tuple[FieldDefinition, str, str]] = []

        for field in field_definitions:
            if field.label in skipped_fields:
                continue
            expected_value = expected_values.get(field.label, "").strip()
            if not expected_value:
                continue

            try:
                if field.input_type == "combobox":
                    current_value = get_combobox_current_value(
                        driver,
                        field.label,
                        timeout_per_scroll=0.2 if not field.required else 0.35,
                    )
                else:
                    current_value = get_field_current_value(
                        driver,
                        field.label,
                        timeout_per_scroll=0.2 if not field.required else 0.35,
                    )
            except TimeoutException:
                if field.required:
                    raise
                continue

            if normalize_field_value(current_value) != normalize_field_value(expected_value):
                mismatches.append((field, expected_value, current_value))

        if not mismatches:
            log_event(
                "VERIFY",
                f"Verification pass {verification_pass}: all filled values matched expected data.",
            )
            return

        log_event(
            "VERIFY",
            f"Verification pass {verification_pass}: found {len(mismatches)} mismatch(es). "
            "Refilling changed fields.",
        )
        for field, expected_value, current_value in mismatches:
            log_event(
                "VERIFY",
                f"Re-filling {field.label}: expected '{expected_value}', current '{current_value}'.",
            )
            if field.input_type == "combobox":
                fill_combobox_field(driver, field.label, expected_value, timeout_per_scroll=1)
            else:
                fill_text_or_number_field(driver, field.label, expected_value, timeout_per_scroll=1)

    log_event("VERIFY", "Verification finished after maximum refill passes.")


def dismiss_optional_ad_popup(driver: webdriver.Firefox, timeout_seconds: int = 5) -> None:
    close_button_locator = (
        By.XPATH,
        "//button[@data-testid='button' and normalize-space()='Close']",
    )

    try:
        close_button = WebDriverWait(driver, timeout_seconds).until(
            EC.element_to_be_clickable(close_button_locator)
        )
    except TimeoutException:
        log_event("PAGE", "Optional ad popup did not appear.")
        return

    close_button.click()
    log_event("PAGE", "Optional ad popup closed.")


def wait_for_clickable(
    driver: webdriver.Firefox,
    locator: tuple[str, str],
    timeout_seconds: int = 15,
) -> WebElement:
    return WebDriverWait(driver, timeout_seconds).until(EC.element_to_be_clickable(locator))


def fill_brand_name(driver: webdriver.Firefox, brand_name: str) -> None:
    brand_input_locator = (By.CSS_SELECTOR, "input[placeholder='Enter Brand Name']")
    check_brand_button_locator = (
        By.XPATH,
        "//button[@data-testid='button' and normalize-space()='Check Brand']",
    )

    brand_input = wait_for_clickable(driver, brand_input_locator)
    brand_input.clear()
    brand_input.send_keys(brand_name)
    log_event("LISTING", f"Entered brand name: {brand_name}")

    check_brand_button = wait_for_clickable(driver, check_brand_button_locator)
    check_brand_button.click()
    log_event("LISTING", "Clicked Check Brand.")


def click_create_new_listing(driver: webdriver.Firefox) -> None:
    create_listing_button_locator = (
        By.XPATH,
        "//button[@data-testid='button' and normalize-space()='Create new listing']",
    )
    create_listing_button = wait_for_clickable(driver, create_listing_button_locator)
    create_listing_button.click()
    log_event("LISTING", "Clicked Create new listing.")


def click_optional_continue(driver: webdriver.Firefox, timeout_seconds: int = 5) -> None:
    continue_button_locator = (
        By.XPATH,
        "//button[@type='button' and .//span[normalize-space()='Continue']]",
    )

    try:
        continue_button = wait_for_clickable(driver, continue_button_locator, timeout_seconds)
    except TimeoutException:
        log_event("LISTING", "Optional Continue button did not appear.")
        return

    continue_button.click()
    log_event("LISTING", "Clicked optional Continue button.")


def preview_selected_image_folder(
    config: BotConfig,
    brand_name: str,
    surface_name: str,
) -> ImageFolder | None:
    if not str(config.image_directory):
        log_event("IMAGES", "Image directory not configured yet, skipping image folder selection.")
        return None

    selected_folder = choose_image_folder_for_brand(
        config.image_directory,
        brand_name,
        surface_name,
    )
    if selected_folder is None:
        log_event("IMAGES", f"No image folder is available for brand: {brand_name}")
        return None

    log_event(
        "IMAGES",
        f"Selected image folder: {selected_folder.folder_path.name} "
        f"with {len(selected_folder.image_paths)} image(s)",
    )
    for image_path in selected_folder.image_paths:
        log_event("IMAGES", f"Queued image file: {image_path.name}")
    log_event(
        "IMAGES",
        "Folder name after successful upload would become: "
        f"{build_exhausted_folder_name(selected_folder, brand_name, surface_name)}",
    )
    return selected_folder


def click_image_slot(driver: webdriver.Firefox, slot_id: str) -> None:
    image_slot = wait_for_clickable(driver, (By.ID, slot_id))
    image_slot.click()
    log_event("IMAGES", f"Selected image slot: {slot_id}")


def upload_image_to_selected_slot(driver: webdriver.Firefox, image_path: Path) -> None:
    upload_input = WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.ID, "upload-image"))
    )
    upload_input.send_keys(str(image_path))
    WebDriverWait(driver, 10).until(lambda _: upload_input.get_attribute("value"))
    sleep(0.5)
    log_event("IMAGES", f"Uploaded image file into active slot: {image_path.name}")


def is_image_slot_uploaded(driver: webdriver.Firefox, slot_id: str) -> bool:
    slot_element = driver.find_element(By.ID, slot_id)
    has_thumbnail = bool(slot_element.find_elements(By.CSS_SELECTOR, "img.styles__Img-sc-1o2k4cf-0"))
    has_check_icon = bool(slot_element.find_elements(By.CSS_SELECTOR, "i.fa-check"))
    has_plus_icon = bool(slot_element.find_elements(By.CSS_SELECTOR, "i.fa-plus"))
    return has_thumbnail and has_check_icon and not has_plus_icon


def get_incomplete_uploaded_image_slots(
    driver: webdriver.Firefox,
    slot_ids: list[str],
) -> list[str]:
    return [
        slot_id for slot_id in slot_ids
        if not is_image_slot_uploaded(driver, slot_id)
    ]


def wait_for_uploaded_image_slots(
    driver: webdriver.Firefox,
    slot_ids: list[str],
    timeout_seconds: int = IMAGE_UPLOAD_VERIFY_TIMEOUT_SECONDS,
) -> None:
    def all_slots_uploaded(_: webdriver.Firefox) -> bool:
        return not get_incomplete_uploaded_image_slots(driver, slot_ids)

    try:
        WebDriverWait(driver, timeout_seconds).until(all_slots_uploaded)
    except TimeoutException as error:
        incomplete_slots = get_incomplete_uploaded_image_slots(driver, slot_ids)
        raise TimeoutException(
            "Image upload verification timed out. Incomplete slot(s): "
            + ", ".join(incomplete_slots)
        ) from error

    log_event("IMAGES", f"Verified uploaded image slot(s): {', '.join(slot_ids)}")


def upload_image_folder(
    driver: webdriver.Firefox,
    image_folder: ImageFolder,
    brand_name: str,
    surface_name: str,
    pause_controller: PauseController,
    config: BotConfig,
    slot_ids: list[str] | None = None,
) -> None:
    if not image_folder.image_paths:
        raise ValueError(f"No images found in folder: {image_folder.folder_path}")

    target_slot_ids = slot_ids or IMAGE_SLOT_IDS
    if not target_slot_ids:
        raise ValueError("No image upload slot ids are configured.")

    upload_count = min(len(image_folder.image_paths), len(target_slot_ids))
    slot_image_paths = list(zip(target_slot_ids[:upload_count], image_folder.image_paths))
    for index, (slot_id, image_path) in enumerate(slot_image_paths, start=1):
        checkpoint_pause(pause_controller, f"Before upload slot {index}", driver, config)
        click_image_slot(driver, slot_id)
        upload_image_to_selected_slot(driver, image_path)
        checkpoint_pause(pause_controller, f"After upload slot {index}", driver, config)

    uploaded_slot_ids = target_slot_ids[:upload_count]
    slot_image_path_by_id = dict(slot_image_paths)
    for retry_pass in range(IMAGE_UPLOAD_RETRY_PASSES + 1):
        try:
            wait_for_uploaded_image_slots(driver, uploaded_slot_ids)
            break
        except TimeoutException:
            incomplete_slot_ids = get_incomplete_uploaded_image_slots(driver, uploaded_slot_ids)
            if retry_pass >= IMAGE_UPLOAD_RETRY_PASSES:
                raise
            log_event(
                "IMAGES",
                "Retrying incomplete image upload slot(s): "
                + ", ".join(incomplete_slot_ids),
            )
            for slot_id in incomplete_slot_ids:
                image_path = slot_image_path_by_id[slot_id]
                click_image_slot(driver, slot_id)
                upload_image_to_selected_slot(driver, image_path)
    checkpoint_pause(pause_controller, "After verifying uploaded image slots", driver, config)

    if len(image_folder.image_paths) > len(target_slot_ids):
        log_event(
            "IMAGES",
            f"Only uploaded the first {len(target_slot_ids)} image(s); "
            f"{len(image_folder.image_paths) - len(target_slot_ids)} extra image(s) were skipped.",
        )


def open_selling_info_tab(driver: webdriver.Firefox) -> None:
    selling_info_tab_locator = (
        By.XPATH,
        "//button[@role='tab' and .//span[contains(normalize-space(), 'Price, Stock and Shipping Information')]]",
    )
    selling_info_tab = wait_for_clickable(driver, selling_info_tab_locator)
    selling_info_tab.click()
    log_event("NAV", "Opened Price, Stock and Shipping Information tab.")


def open_product_description_tab(driver: webdriver.Firefox) -> None:
    product_description_tab_locator = (
        By.XPATH,
        "//button[@role='tab' and .//span[contains(normalize-space(), 'Product Description')]]",
    )
    product_description_tab = wait_for_clickable(driver, product_description_tab_locator)
    product_description_tab.click()
    log_event("NAV", "Opened Product Description tab.")


def open_additional_description_tab(driver: webdriver.Firefox) -> None:
    additional_description_tab_locator = (
        By.XPATH,
        "//button[@role='tab' and .//span[contains(normalize-space(), 'Additional Description')]]",
    )
    additional_description_tab = wait_for_clickable(driver, additional_description_tab_locator)
    additional_description_tab.click()
    log_event("NAV", "Opened Additional Description tab.")


def open_tab_via_autogui(
    driver: webdriver.Firefox,
    tab_label: str,
    xpath_locator: str,
    settle_seconds: float = 2,
) -> None:
    last_error: Exception | None = None
    for attempt in range(3):
        try:
            tab_button = wait_for_clickable(driver, (By.XPATH, xpath_locator))
            click_element_via_autogui(driver, tab_button, f"{tab_label} tab")
            break
        except WebDriverException as exc:
            last_error = exc
            log_event(
                "NAV",
                f"Retrying {tab_label} tab click after transient WebDriver issue on attempt {attempt + 1}: {exc.__class__.__name__}",
            )
            sleep(0.2)
    else:
        raise last_error if last_error is not None else TimeoutException(
            f"Could not click {tab_label} tab."
        )
    log_event("NAV", f"Clicked {tab_label} tab via pyautogui.")
    if settle_seconds > 0:
        log_event("NAV", f"Waiting {settle_seconds} seconds for {tab_label} tab to settle...")
        sleep(settle_seconds)


def click_save_and_go_back_button(driver: webdriver.Firefox) -> None:
    save_and_go_back_button = WebDriverWait(driver, 15).until(
        EC.presence_of_element_located(
            (By.XPATH, "//button[@data-testid='button' and normalize-space()='Save & Go Back']")
        )
    )
    click_element_via_autogui(driver, save_and_go_back_button, "Save & Go Back button")
    log_event("DONE", "Clicked Save & Go Back.")


def fill_size_qualifier_field(driver: webdriver.Firefox, field_value: str) -> None:
    size_wrapper = get_field_wrapper(driver, "Size", timeout_per_scroll=1)
    qualifier_combobox = driver.execute_script(
        """
        const wrapper = arguments[0];
        const comboboxes = Array.from(
            wrapper.querySelectorAll("button[role='combobox'], [role='combobox']")
        ).filter((element) => {
            if (!element) {
                return false;
            }
            const style = window.getComputedStyle(element);
            return style.display !== 'none' && style.visibility !== 'hidden';
        });
        if (comboboxes.length < 2) {
            return comboboxes.length === 1 ? comboboxes[0] : null;
        }
        return comboboxes[1];
        """,
        size_wrapper,
    )
    if qualifier_combobox is None:
        raise TimeoutException("Could not find Size Qualifier combobox inside the Size field wrapper.")
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", qualifier_combobox)
    log_event(
        "DESC",
        f"Preparing second Size dropdown selection (current button like 'Number') -> {field_value}",
    )
    select_combobox_option(driver, qualifier_combobox, field_value, "Size Qualifier")
    log_event("DESC", f"Selected second Size dropdown (Size Qualifier): {field_value}")


def fill_product_description_fields(
    driver: webdriver.Firefox,
    field_definitions: list[FieldDefinition],
    product_input_row: ProductInputRow,
) -> FillResult:
    generated_values: dict[str, str] = {}
    skipped_fields: set[str] = set()

    size_qualifier_value = product_input_row.values.get("Size Qualifier", "").strip()
    main_size_value = product_input_row.size.strip()

    for field in field_definitions:
        raw_value = product_input_row.values.get(field.label, "").strip()

        if field.label == "Size":
            if size_qualifier_value:
                fill_size_qualifier_field(driver, size_qualifier_value)
            else:
                log_event("DESC", "Skipping Size Qualifier: no Excel value provided.")

            if not main_size_value:
                log_event("DESC", "Skipping Size: no row match size value was available.")
                continue

            log_event("DESC", f"Preparing main Size dropdown selection after qualifier -> {main_size_value}")
            fill_combobox_field(driver, field.label, main_size_value, timeout_per_scroll=1)
            continue

        if not raw_value:
            log_event("DESC", f"Skipping {field.label}: no Excel value provided.")
            continue

        if field.label == "Style Code":
            raw_value = f"{raw_value}{generate_sku_suffix()}"
            generated_values[field.label] = raw_value
            log_event("DESC", f"Generated Style Code value: {raw_value}")

        if field.input_type == "combobox":
            fill_combobox_field(driver, field.label, raw_value, timeout_per_scroll=1)
        elif field.input_type == "tag_input":
            fill_tag_input_field(driver, field.label, raw_value)
        elif field.input_type == "tag_input_commit":
            fill_tag_input_field(
                driver,
                field.label,
                raw_value,
                commit_each_value_with_enter=ENABLE_ENTER_COMMIT_FOR_TAG_INPUT_FIELDS,
            )
        else:
            fill_text_or_number_field(driver, field.label, raw_value, timeout_per_scroll=1)

    return FillResult(generated_values=generated_values, skipped_fields=skipped_fields)


def fill_additional_description_fields(
    driver: webdriver.Firefox,
    field_definitions: list[FieldDefinition],
    product_input_row: ProductInputRow,
) -> FillResult:
    generated_values: dict[str, str] = {}
    skipped_fields: set[str] = set()

    for field in field_definitions:
        raw_value = product_input_row.values.get(field.label, "").strip()

        if field.input_type == "skip":
            skipped_fields.add(field.label)
            log_event("ADDL", f"Skipping {field.label}: marked as skip in JSON.")
            continue

        if not raw_value:
            skipped_fields.add(field.label)
            log_event("ADDL", f"Skipping {field.label}: no Excel value provided.")
            continue

        if field.input_type == "combobox":
            fill_combobox_field(driver, field.label, raw_value, timeout_per_scroll=1)
        elif field.input_type == "tag_input":
            fill_tag_input_field(driver, field.label, raw_value)
        elif field.input_type == "tag_input_commit":
            fill_tag_input_field(
                driver,
                field.label,
                raw_value,
                commit_each_value_with_enter=ENABLE_ENTER_COMMIT_FOR_TAG_INPUT_FIELDS,
            )
        else:
            fill_text_or_number_field(driver, field.label, raw_value, timeout_per_scroll=1)

    return FillResult(generated_values=generated_values, skipped_fields=skipped_fields)


def get_variant_creation_controls(
    driver: webdriver.Firefox,
) -> tuple[WebElement, WebElement | None, WebElement]:
    """Return (qualifier_combobox_or_None, size_combobox, create_button).
    Scoped strictly to the wrapper containing the Create button so
    page-level dropdowns are never mistaken for variant controls.
    Trouser: 1 combobox (size only). Jeans: 2 (qualifier + size).
    """
    deadline = datetime.now().timestamp() + 6
    last_error = "Variant controls not found."
    while datetime.now().timestamp() < deadline:
        try:
            controls = driver.execute_script("""
                function isVisible(el) {
                    if (!el) return false;
                    const s = window.getComputedStyle(el);
                    const r = el.getBoundingClientRect();
                    return s.display!=='none' && s.visibility!=='hidden'
                           && r.width>0 && r.height>0;
                }
                function getCreateBtn(root) {
                    return Array.from(root.querySelectorAll('button')).find(btn => {
                        const span = btn.querySelector('span');
                        return isVisible(btn) && span && span.textContent.trim() === 'Create';
                    }) || null;
                }
                const wrappers = Array.from(document.querySelectorAll(
                    "[class*='AttributesAdditionWrapper'], [class*='OptionWrapper']"
                )).filter(isVisible);
                for (const w of wrappers) {
                    const createBtn = getCreateBtn(w);
                    if (!createBtn) continue;
                    const combos = Array.from(
                        w.querySelectorAll("button[role='combobox']")
                    ).filter(isVisible);
                    if (!combos.length) continue;
                    if (combos.length === 1) return [null, combos[0], createBtn];
                    // Structural signal: the qualifier ("Select Unit") combobox sits
                    // inside [class*='AttributeItemQualifierWrapper']; the size
                    // ("Select Size") combobox does not.
                    const qual = combos.find(b => b.closest("[class*='AttributeItemQualifierWrapper']"));
                    const size = combos.find(b => b !== qual && !b.closest("[class*='AttributeItemQualifierWrapper']"))
                                  || combos.find(b => b !== qual);
                    if (size) return [qual||null, size, createBtn];
                }
                return null;
            """)
            if controls and len(controls) >= 3:
                return controls[0], controls[1], controls[2]
        except Exception as exc:
            last_error = str(exc)
        sleep(0.2)
    raise TimeoutException(f"Could not locate variant controls. {last_error}")

def get_variant_sizes_to_create(product_input_row: ProductInputRow) -> list[str]:
    variant_sizes: list[str] = []
    for column_name in ("size_variant_1", "size_variant_2", "size_variant_3", "size_variant_4"):
        raw_value = product_input_row.values.get(column_name, "").strip()
        if raw_value and raw_value not in variant_sizes:
            variant_sizes.append(raw_value)
    return variant_sizes


def click_variant_create_button(driver: webdriver.Firefox, variant_size: str) -> None:
    """Wait for the Size section Create button to be enabled, then JS-click it.

    The button textContent includes the SVG <title> "AddCircle" so we CANNOT
    match on full textContent === "Create".  Instead we find the <span> child
    whose text is exactly "Create" and whose button has aria-disabled="false"
    and no disabled attribute.  We also prefer the button whose SVG icon color
    is "blue" (the enabled state) over the greyed-out Brand Color one.
    """
    def _find_enabled(d: webdriver.Firefox) -> "WebElement | bool":
        try:
            btn = d.execute_script("""
                function isVisible(el) {
                    if (!el) return false;
                    const s = window.getComputedStyle(el);
                    const r = el.getBoundingClientRect();
                    return s.display!=='none' && s.visibility!=='hidden'
                           && r.width>0 && r.height>0;
                }
                // Find all buttons that have a <span> child with text "Create"
                // and are NOT disabled.
                const candidates = Array.from(document.querySelectorAll('button')).filter(btn => {
                    if (!isVisible(btn)) return false;
                    if (btn.hasAttribute('disabled')) return false;
                    if (btn.getAttribute('aria-disabled') === 'true') return false;
                    // Check the span child text, not full textContent (which includes SVG title)
                    const span = btn.querySelector('span');
                    return span && span.textContent.trim() === 'Create';
                });
                if (!candidates.length) return null;
                // Prefer the one whose SVG color is "blue" (the enabled Size button)
                const blue = candidates.find(btn => {
                    const svg = btn.querySelector('svg');
                    return svg && (svg.getAttribute('color') === 'blue'
                                   || window.getComputedStyle(svg).color === 'blue');
                });
                return blue || candidates[0];
            """)
        except Exception:
            return False
        return btn if btn else False

    log_event("VARIANT", f"Waiting for enabled Create button for size {variant_size}...")
    create_button = WebDriverWait(driver, 15).until(_find_enabled)
    driver.execute_script("arguments[0].scrollIntoView({block:'center',inline:'nearest'});", create_button)
    sleep(0.4)
    log_event("VARIANT", f"JS-clicking Create button for variant size {variant_size}.")
    driver.execute_script("arguments[0].click();", create_button)
    sleep(0.25)
    try:
        still_enabled = driver.execute_script(
            "return !arguments[0].hasAttribute('disabled') "
            "&& arguments[0].getAttribute('aria-disabled') !== 'true';",
            create_button,
        )
    except Exception:
        still_enabled = False
    if still_enabled:
        log_event("VARIANT", "JS click may have been swallowed; retrying with ActionChains.")
        try:
            ActionChains(driver).move_to_element(create_button).pause(0.1).click().perform()
            sleep(0.25)
        except Exception:
            pass


def wait_for_variant_create_button_enabled(
    driver: webdriver.Firefox,
    timeout_seconds: float = 15,
) -> WebElement:
    def _locate_enabled_button(current_driver: webdriver.Firefox) -> WebElement | bool:
        create_button = current_driver.execute_script(
            """
            function isVisible(element) {
                if (!element) {
                    return false;
                }
                const style = window.getComputedStyle(element);
                const rect = element.getBoundingClientRect();
                return (
                    style.display !== 'none' &&
                    style.visibility !== 'hidden' &&
                    rect.width > 0 &&
                    rect.height > 0
                );
            }

            function isEnabled(button) {
                return (
                    button.getAttribute('aria-disabled') !== 'true' &&
                    !button.hasAttribute('disabled') &&
                    !button.classList.contains('disabled')
                );
            }

            const createButtons = Array.from(document.querySelectorAll('button')).filter((button) => {
                const text = (button.textContent || '').replace(/\\s+/g, ' ').trim();
                return isVisible(button) && isEnabled(button) && text === 'Create';
            });
            return (
                createButtons.find((button) => {
                    const iconTitle = button.querySelector('svg title');
                    return iconTitle && iconTitle.textContent.trim() === 'AddCircle';
                }) ||
                createButtons[0] ||
                null
            );
            """
        )
        if create_button is not None:
            return create_button
        return False

    return WebDriverWait(driver, timeout_seconds).until(_locate_enabled_button)


def wait_for_variant_row_creation(
    driver: webdriver.Firefox,
    variant_qualifier: str,
    variant_size: str,
    timeout_seconds: float = 12,
) -> None:
    expected_size_text = build_variant_size_display_text(variant_size, variant_qualifier)
    size_xpath = (
        "//div[contains(@id,'-size') and contains(@class,'variant-table-cell')]"
        f"//span[@title={xpath_literal(expected_size_text)}"
        f" or contains(@title, {xpath_literal(variant_size)})"
        f" or normalize-space()={xpath_literal(expected_size_text)}]"
    )
    WebDriverWait(driver, timeout_seconds).until(
        EC.presence_of_element_located((By.XPATH, size_xpath))
    )
    log_event("VARIANT", f"Detected created variant row for {expected_size_text}.")


def get_variant_rows(driver: webdriver.Firefox) -> list[WebElement]:
    return driver.find_elements(By.XPATH, "//tr[starts-with(@data-testid,'grid-component-row-')]")


def get_variant_row_sku(row: WebElement) -> str:
    sku_xpath = ".//div[contains(@id,'-sku_id')]//span[contains(@class,'variant-cell-readonly-value')]"
    sku_elements = row.find_elements(By.XPATH, sku_xpath)
    if not sku_elements:
        return ""
    sku_element = sku_elements[0]
    return (sku_element.get_attribute("title") or sku_element.text or "").strip()


def get_variant_row_size_text(row: WebElement) -> str:
    size_xpath = ".//div[contains(@id,'-size')]//span[contains(@class,'variant-cell-readonly-value')]"
    size_elements = row.find_elements(By.XPATH, size_xpath)
    if not size_elements:
        return ""
    size_element = size_elements[0]
    return (size_element.get_attribute("title") or size_element.text or "").strip()


def build_variant_size_display_text(size_value: str, variant_qualifier: str = "") -> str:
    normalized_size = str(size_value).strip()
    normalized_qualifier = str(variant_qualifier).strip()
    if not normalized_qualifier:
        normalized_qualifier = "Number" if normalized_size.isdigit() else "Regular"
    return f"{normalized_size} {normalized_qualifier}".strip()


def get_variant_row_by_size_text(
    driver: webdriver.Firefox,
    size_text: str,
    timeout_seconds: float = 10,
) -> WebElement:
    def _locate_row(current_driver: webdriver.Firefox) -> WebElement | bool:
        for row in get_variant_rows(current_driver):
            try:
                if get_variant_row_size_text(row) == size_text:
                    return row
            except StaleElementReferenceException:
                continue
        return False

    return WebDriverWait(driver, timeout_seconds).until(_locate_row)


def get_variant_row_checkbox_input(
    driver: webdriver.Firefox,
    row_size_text: str,
) -> WebElement:
    row = get_variant_row_by_size_text(driver, row_size_text)
    return row.find_element(By.XPATH, ".//input[starts-with(@id,'checkbox-')]")


def click_variant_row_copy_button(
    driver: webdriver.Firefox,
    source_size_text: str,
) -> None:
    row = get_variant_row_by_size_text(driver, source_size_text)
    copy_button = row.find_element(
        By.XPATH,
        ".//button[.//*[local-name()='title' and normalize-space()='ContentCopy']]",
    )
    click_element_via_autogui(driver, copy_button, f"Copy variant row for size {source_size_text}")
    log_event("VARIANT", f"Clicked copy for source row size: {source_size_text}")


def select_variant_target_rows_for_paste(
    driver: webdriver.Firefox,
    source_size_text: str,
) -> list[str]:
    candidate_target_sizes: list[str] = []
    for row in get_variant_rows(driver):
        try:
            row_size_text = get_variant_row_size_text(row)
            if not row_size_text or row_size_text == source_size_text:
                continue

            checkbox_inputs = row.find_elements(By.XPATH, ".//input[starts-with(@id,'checkbox-')]")
            if not checkbox_inputs:
                continue

            checkbox_input = checkbox_inputs[0]
            if checkbox_input.get_attribute("disabled") is not None:
                continue
            candidate_target_sizes.append(row_size_text)
        except StaleElementReferenceException:
            continue

    selected_target_sizes: list[str] = []
    for row_size_text in candidate_target_sizes:
        try:
            checkbox_input = get_variant_row_checkbox_input(driver, row_size_text)
            if checkbox_input.get_attribute("disabled") is not None:
                continue

            if not checkbox_input.is_selected():
                row = get_variant_row_by_size_text(driver, row_size_text)
                checkbox_wrapper = row.find_element(
                    By.XPATH,
                    ".//div[contains(@class,'InputCheckboxWrapper')]",
                )
                click_element_via_autogui(
                    driver,
                    checkbox_wrapper,
                    f"Select variant checkbox for size {row_size_text}",
                )
                WebDriverWait(driver, 5).until(
                    lambda current_driver, expected=row_size_text: get_variant_row_checkbox_input(
                        current_driver,
                        expected,
                    ).is_selected()
                )

            selected_target_sizes.append(row_size_text)
        except StaleElementReferenceException:
            continue

    return selected_target_sizes


def get_selected_variant_row_sizes(driver: webdriver.Firefox) -> list[str]:
    selected_sizes: list[str] = []
    for row in get_variant_rows(driver):
        try:
            checkbox_inputs = row.find_elements(By.XPATH, ".//input[starts-with(@id,'checkbox-')]")
            if not checkbox_inputs or not checkbox_inputs[0].is_selected():
                continue
            row_size_text = get_variant_row_size_text(row)
            if row_size_text:
                selected_sizes.append(row_size_text)
        except StaleElementReferenceException:
            continue
    return selected_sizes


def get_variant_row_sku_input(
    driver: webdriver.Firefox,
    row_size_text: str,
) -> WebElement:
    row = get_variant_row_by_size_text(driver, row_size_text)
    return row.find_element(By.XPATH, ".//input[contains(@id,'-sku_id') and @type='text']")


def get_variant_row_listing_status_combobox(
    driver: webdriver.Firefox,
    row_size_text: str,
) -> WebElement:
    row = get_variant_row_by_size_text(driver, row_size_text)
    status_xpath = (
        ".//*[contains(@id,'-listing_status')]"
        "//*[self::button or self::input or self::div]"
        "["
        "@role='combobox'"
        " or @aria-haspopup='listbox'"
        " or @aria-haspopup='dialog'"
        " or contains(@class,'SelectInput')"
        "]"
    )
    return row.find_element(By.XPATH, status_xpath)


def get_variant_row_listing_status_value(
    driver: webdriver.Firefox,
    row_size_text: str,
) -> str:
    row = get_variant_row_by_size_text(driver, row_size_text)
    readonly_status_xpath = (
        ".//div[contains(@id,'-listing_status')]"
        "//span[contains(@class,'variant-cell-readonly-value')]"
    )
    readonly_status_elements = row.find_elements(By.XPATH, readonly_status_xpath)
    if readonly_status_elements:
        readonly_status_element = readonly_status_elements[0]
        return (
            readonly_status_element.get_attribute("title")
            or readonly_status_element.text
            or ""
        ).strip()

    status_combobox = get_variant_row_listing_status_combobox(driver, row_size_text)
    status_value = driver.execute_script(
        """
        const element = arguments[0];

        function cleanText(node) {
            if (!node) {
                return "";
            }
            const clone = node.cloneNode(true);
            clone.querySelectorAll("svg, title").forEach((child) => child.remove());
            return (clone.textContent || "").trim();
        }

        const candidates = [
            element.querySelector("input"),
            element.querySelector("[class*='SingleValue']"),
            element.querySelector("[class*='ButtonText']"),
            element.querySelector("[class*='LabelText']"),
            element.querySelector("span"),
            element,
        ].filter(Boolean);

        for (const candidate of candidates) {
            const parts = [
                candidate.getAttribute("value"),
                candidate.getAttribute("title"),
                candidate.getAttribute("aria-label"),
                cleanText(candidate),
            ];
            for (const part of parts) {
                if (part && String(part).trim()) {
                    return String(part).trim();
                }
            }
        }
        return "";
        """,
        status_combobox,
    )
    return str(status_value or "").strip()


def build_variant_row_sku(
    source_sku: str,
    source_size_value: str,
    target_row_size_text: str,
) -> str:
    base_size = str(source_size_value).strip()
    target_size = target_row_size_text.strip().split()[0]
    size_token = f"-{base_size}-"
    replacement_token = f"-{target_size}-"
    if size_token in source_sku:
        return source_sku.replace(size_token, replacement_token, 1)
    return source_sku


def update_variant_row_sku(
    driver: webdriver.Firefox,
    row_size_text: str,
    new_sku: str,
) -> None:
    deadline = datetime.now().timestamp() + 8
    last_error: str = "SKU input was not updated."

    while datetime.now().timestamp() < deadline:
        try:
            sku_input = get_variant_row_sku_input(driver, row_size_text)
            current_value = (sku_input.get_attribute("value") or "").strip()
            if current_value == new_sku:
                log_event("VARIANT", f"Variant row {row_size_text} already has SKU: {new_sku}")
                return

            set_input_value(driver, sku_input, new_sku)
            refreshed_input = get_variant_row_sku_input(driver, row_size_text)
            refreshed_value = (refreshed_input.get_attribute("value") or "").strip()
            if refreshed_value == new_sku:
                log_event("VARIANT", f"Updated variant row {row_size_text} SKU to: {new_sku}")
                return
            last_error = (
                f"Variant row {row_size_text} SKU read back as '{refreshed_value}' "
                f"instead of '{new_sku}'."
            )
        except StaleElementReferenceException:
            last_error = f"Variant row {row_size_text} SKU input went stale while updating."

        sleep(0.2)

    raise TimeoutException(last_error)


def update_variant_row_listing_status(
    driver: webdriver.Firefox,
    row_size_text: str,
    listing_status: str,
) -> None:
    normalized_status = str(listing_status).strip()
    if not normalized_status:
        log_event("VARIANT", f"Skipping Listing Status for row {row_size_text}: no value provided.")
        return

    current_value = get_variant_row_listing_status_value(driver, row_size_text)
    if field_values_match(current_value, normalized_status):
        log_event(
            "VARIANT",
            f"Variant row {row_size_text} already has Listing Status: {current_value}",
        )
        return

    status_combobox = get_variant_row_listing_status_combobox(driver, row_size_text)
    select_combobox_option(
        driver,
        status_combobox,
        normalized_status,
        f"Variant row {row_size_text} Listing Status",
    )
    refreshed_value = get_variant_row_listing_status_value(driver, row_size_text)
    if not field_values_match(refreshed_value, normalized_status):
        raise TimeoutException(
            f"Variant row {row_size_text} Listing Status read back as '{refreshed_value}' "
            f"instead of '{normalized_status}'."
        )
    log_event("VARIANT", f"Updated variant row {row_size_text} Listing Status to: {normalized_status}")


def update_pasted_variant_row_skus(
    config: BotConfig,
    listing_selection: ListingSelection,
    driver: webdriver.Firefox,
    source_sku: str,
    target_row_sizes: list[str],
) -> None:
    if not target_row_sizes:
        log_event("VARIANT", "Skipping variant SKU rewrite: no target rows were selected.")
        return

    for row_size_text in target_row_sizes:
        target_size = row_size_text.strip().split()[0]
        target_listing_status = ""
        try:
            price_row = load_product_input_row(
                config.price_stock_shipping_excel,
                listing_selection.kind,
                target_size,
                surface=listing_selection.surface,
            )
        except ValueError as exc:
            log_event(
                "VARIANT",
                "Could not find Price/Stock/Shipping row for variant size "
                f"{target_size}. Falling back to copied-row size replacement: {exc}",
            )
            updated_sku = build_variant_row_sku(source_sku, listing_selection.size, row_size_text)
        else:
            target_base_sku = price_row.values.get("Seller SKU ID", "").strip()
            if not target_base_sku:
                log_event(
                    "VARIANT",
                    f"Price/Stock/Shipping row for size {target_size} has no Seller SKU ID. "
                    "Falling back to copied-row size replacement.",
                )
                updated_sku = build_variant_row_sku(source_sku, listing_selection.size, row_size_text)
            else:
                updated_sku = build_sku_with_suffix(target_base_sku, generate_sku_suffix())
            target_listing_status = price_row.values.get("Listing Status", "").strip()
        log_event(
            "VARIANT",
            f"Rewriting SKU for row {row_size_text}: source '{source_sku}' -> '{updated_sku}'",
        )
        update_variant_row_sku(driver, row_size_text, updated_sku)
        if target_listing_status:
            update_variant_row_listing_status(driver, row_size_text, target_listing_status)


def wait_for_variant_paste_button_enabled(
    driver: webdriver.Firefox,
    timeout_seconds: float = 10,
) -> WebElement:
    paste_xpath = (
        "//button[.//span[normalize-space()='Paste in selected']]"
    )

    def _locate_enabled_button(current_driver: webdriver.Firefox) -> WebElement | bool:
        buttons = current_driver.find_elements(By.XPATH, paste_xpath)
        for button in buttons:
            try:
                aria_disabled = (button.get_attribute("aria-disabled") or "").strip().lower()
                disabled_attribute = button.get_attribute("disabled")
                if button.is_enabled() and aria_disabled != "true" and disabled_attribute is None:
                    return button
            except StaleElementReferenceException:
                continue
        return False

    return WebDriverWait(driver, timeout_seconds).until(_locate_enabled_button)


def copy_source_variant_into_selected_rows(
    driver: webdriver.Firefox,
    source_size_text: str,
) -> list[str]:
    click_variant_row_copy_button(driver, source_size_text)
    selected_target_sizes = select_variant_target_rows_for_paste(driver, source_size_text)
    if not selected_target_sizes:
        log_event("VARIANT", "No eligible target variant rows were found for paste.")
        return []

    log_event(
        "VARIANT",
        "Selected target variant size rows for paste: " + ", ".join(selected_target_sizes),
    )
    confirmed_selected_sizes = get_selected_variant_row_sizes(driver)
    log_event(
        "VARIANT",
        f"Confirmed selected target rows before paste [{len(confirmed_selected_sizes)}]: "
        + ", ".join(confirmed_selected_sizes),
    )
    log_event("VARIANT", f"Source size row for paste: {source_size_text}")
    paste_button = wait_for_variant_paste_button_enabled(driver)
    click_element_via_autogui(driver, paste_button, "Paste in selected variants")
    log_event("VARIANT", "Clicked Paste in selected for copied source variant data.")
    sleep(1)
    return selected_target_sizes


def fill_variant_page(
    config: BotConfig,
    listing_selection: ListingSelection,
    driver: webdriver.Firefox,
    pause_controller: PauseController,
    product_input_row: ProductInputRow,
    source_sku: str,
) -> None:
    variant_qualifier = product_input_row.values.get("variant_qualifier", "").strip()
    variant_sizes = get_variant_sizes_to_create(product_input_row)

    log_event(
        "VARIANT",
        f"Loaded variant_qualifier='{variant_qualifier}' (raw row keys: {list(product_input_row.values.keys())}).",
    )

    if not variant_sizes:
        log_event("VARIANT", "Skipping Variant page: no size_variant values were provided in Excel.")
        return

    should_select_qualifier = bool(variant_qualifier)
    if not should_select_qualifier:
        log_event(
            "VARIANT",
            "Skipping Variant Qualifier: no Excel value provided for variant_qualifier.",
        )

    for variant_size in variant_sizes:
        qualifier_combobox, size_combobox, _ = get_variant_creation_controls(driver)
        if should_select_qualifier:
            if qualifier_combobox is None:
                raise RuntimeError(
                    f"Variant Qualifier '{variant_qualifier}' was provided in Excel, "
                    "but no qualifier combobox ('Select Unit') was found on the page."
                )
            log_event("VARIANT", f"Creating size variant with qualifier '{variant_qualifier}' and size '{variant_size}'.")
            select_combobox_option(driver, qualifier_combobox, variant_qualifier, "Variant Qualifier")
            log_event("VARIANT", f"Selected Variant Qualifier: {variant_qualifier}")
            _, size_combobox, _ = get_variant_creation_controls(driver)
        else:
            log_event("VARIANT", f"Creating size variant without qualifier and size '{variant_size}'.")
        select_combobox_option(driver, size_combobox, variant_size, "Variant Size")
        log_event("VARIANT", f"Selected Variant Size: {variant_size}")
        click_variant_create_button(driver, variant_size)
        log_event("VARIANT", f"Clicked Create for variant size {variant_size}.")
        wait_for_variant_row_creation(driver, variant_qualifier, variant_size)

    source_size_text = build_variant_size_display_text(
        product_input_row.size,
        variant_qualifier,
    )
    log_event("VARIANT", f"Using source size row for copy/paste: {source_size_text}")
    selected_target_sizes = copy_source_variant_into_selected_rows(driver, source_size_text)
    update_pasted_variant_row_skus(
        config,
        listing_selection,
        driver,
        source_sku,
        selected_target_sizes,
    )
    log_event("VARIANT", "Variant row updates complete. Final save will be triggered at the end of the flow.")


def wait_for_changes_saved_toast(
    driver: webdriver.Firefox,
    pause_controller: PauseController,
    config: BotConfig,
    timeout_seconds: int = 20,
) -> None:
    deadline = datetime.now().timestamp() + timeout_seconds
    while datetime.now().timestamp() < deadline:
        checkpoint_pause(pause_controller, "Waiting for 'Changes saved!' toast", driver, config)
        if has_changes_saved_toast(driver):
            log_event("TOAST", "Detected success toast: Changes saved!")
            return
        sleep(0.25)

    raise TimeoutException("Timed out waiting for 'Changes saved!' toast after opening selling info tab.")


def has_changes_saved_toast(driver: webdriver.Firefox) -> bool:
    success_toast_locator = (
        By.XPATH,
        "//div[contains(@class,'toast-title-container')]//b[normalize-space()='Changes saved!']",
    )
    return bool(driver.find_elements(*success_toast_locator))


def get_tab_progress_text(driver: webdriver.Firefox, tab_xpath: str) -> str:
    for _attempt in range(3):
        try:
            tab_buttons = driver.find_elements(By.XPATH, tab_xpath)
            if not tab_buttons:
                return ""
            return " ".join((tab_buttons[0].text or "").split())
        except StaleElementReferenceException:
            sleep(0.3)
    return ""


def extract_nonzero_tab_progress(progress_text: str) -> tuple[int, int] | None:
    match = re.search(r"\((\d+)\s*/\s*(\d+)\)", progress_text)
    if not match:
        return None

    first_value = int(match.group(1))
    second_value = int(match.group(2))
    if first_value == 0:
        return None
    return first_value, second_value


def has_nonzero_page_switch_progress(
    driver: webdriver.Firefox,
    tab_sequence: list[tuple[str, str]] | tuple[tuple[str, str], ...],
) -> tuple[str, str, int, int] | None:
    for tab_label, tab_xpath in tab_sequence:
        progress_text = get_tab_progress_text(driver, tab_xpath)
        progress_values = extract_nonzero_tab_progress(progress_text)
        if progress_values is not None:
            return tab_label, progress_text, progress_values[0], progress_values[1]
    return None


def cycle_page_switch_verification_until_toast(
    driver: webdriver.Firefox,
    pause_controller: PauseController,
    config: BotConfig,
    cycle_label: str = "page",
    tab_sequence: list[tuple[str, str]] | tuple[tuple[str, str], ...] = (),
    timeout_seconds: int = 45,
) -> None:
    if not tab_sequence:
        raise ValueError("tab_sequence must contain at least one tab label/xpath pair.")

    deadline = datetime.now().timestamp() + timeout_seconds
    tab_index = 0

    log_event(
        "VERIFY",
        f"Starting IN THE PAGE SWITCH VERIFICATION CYCLE for {cycle_label} using "
        f"{'toast or counter progress' if USE_CHANGES_SAVED_TOAST_FOR_VERIFICATION else 'counter progress only'}.",
    )
    while datetime.now().timestamp() < deadline:
        checkpoint_pause(
            pause_controller,
            f"Cycling {cycle_label} page switches until save toast",
            driver,
            config,
        )
        # Small test-run mode: keep toast detection disabled and rely only on tab counters.
        if USE_CHANGES_SAVED_TOAST_FOR_VERIFICATION and has_changes_saved_toast(driver):
            log_event("TOAST", f"Detected success toast during {cycle_label} page-switch verification cycle.")
            return
        if tab_index >= 2:
            nonzero_progress = has_nonzero_page_switch_progress(driver, tab_sequence)
            if nonzero_progress is not None:
                tab_label, progress_text, first_value, second_value = nonzero_progress
                log_event(
                    "VERIFY",
                    f"Detected nonzero tab progress during {cycle_label} page-switch verification cycle: "
                    f"{tab_label} -> {progress_text} ({first_value}/{second_value}). Treating verification as complete.",
                )
                return

        current_tab_label, current_tab_xpath = tab_sequence[tab_index % len(tab_sequence)]
        open_tab_via_autogui(
            driver,
            current_tab_label,
            current_tab_xpath,
            settle_seconds=1,
        )
        tab_index += 1

    raise TimeoutException(
        f"Timed out while cycling {cycle_label} page switches waiting for verification progress."
    )


TAB_XPATHS = {
    "Additional Description": "//button[@role='tab' and .//span[contains(normalize-space(), 'Additional Description')]]",
    "Product Description": "//button[@role='tab' and .//span[contains(normalize-space(), 'Product Description')]]",
    "Price, Stock and Shipping Information": "//button[@role='tab' and .//span[contains(normalize-space(), 'Price, Stock and Shipping Information')]]",
    "Image addition": "//button[@role='tab' and .//span[contains(normalize-space(), 'Image addition')]]",
    "Variant addition": "//button[@role='tab' and .//span[contains(normalize-space(), 'Variant addition')]]",
}

LEGACY_PRODUCT_PAGE_FLOWS = {
    "jeans": (
        "additional_description",
        "product_description",
        "price_stock_shipping",
        "images",
        "variants",
    ),
    "shorts": (
        "additional_description",
        "product_description",
        "price_stock_shipping",
        "images",
    ),
}


@dataclass(slots=True)
class FlowState:
    price_fill_result: FillResult | None = None
    pending_image_folder_use: PendingImageFolderUse | None = None
    context: dict[str, object] = field(default_factory=dict)


@dataclass(slots=True)
class FlowStepDefinition:
    order: int
    step_id: str
    handler: str
    spec_file: str
    spec_payload: dict[str, object]


@dataclass(slots=True)
class FlowPageDefinition:
    order: int
    step_name: str
    page_file: str
    handler: str
    tab_label: str | None = None
    tab_xpath: str | None = None
    checkpoint_label: str | None = None
    worksheet_name: str | None = None
    field_json_config_attr: str | None = None
    excel_config_attr: str | None = None
    log_stage: str | None = None
    log_message: str | None = None
    filled_checkpoint_label: str | None = None
    verify: dict[str, object] | None = None
    snapshot_name: str | None = None
    brand_name: str | None = None
    locator_strategy: str | None = None
    source_snapshot: str | None = None
    upload_slot_ids: list[str] = field(default_factory=list)
    fields: list[FieldDefinition] = field(default_factory=list)


@dataclass(slots=True)
class FlowDefinition:
    surface: str
    product_type: str
    flow_name: str
    flow_directory: Path
    manifest_context: dict[str, object]
    steps: list[FlowStepDefinition]


def load_json_payload(json_path: Path) -> dict[str, object]:
    return json.loads(json_path.read_text(encoding="utf-8"))


def resolve_flow_directory(product_type: str, surface: str = DEFAULT_FLOW_SURFACE) -> Path:
    return FLOW_CONFIG_ROOT / f"{product_type}_{surface}"


def discover_flow_target_options() -> list[FlowTargetOption]:
    options: list[FlowTargetOption] = []
    if not FLOW_CONFIG_ROOT.exists():
        return options

    for flow_directory in sorted(FLOW_CONFIG_ROOT.iterdir(), key=lambda path: path.name.lower()):
        if not flow_directory.is_dir():
            continue
        if not (flow_directory / "flow.json").exists():
            continue
        if "_" not in flow_directory.name:
            continue
        product_type, surface = flow_directory.name.rsplit("_", 1)
        product_type = product_type.strip().lower()
        surface = surface.strip().lower()
        if not product_type or not surface:
            continue
        options.append(
            FlowTargetOption(
                product_type=product_type,
                surface=surface,
                flow_directory=flow_directory,
            )
        )
    return options


def load_listing_flow_definition(
    product_type: str,
    surface: str = DEFAULT_FLOW_SURFACE,
) -> FlowDefinition | None:
    flow_directory = resolve_flow_directory(product_type, surface)
    flow_manifest_path = flow_directory / "flow.json"
    if not flow_manifest_path.exists():
        return None

    payload = load_json_payload(flow_manifest_path)
    step_entries = payload.get("steps")
    if step_entries is None:
        legacy_page_entries = payload.get("pages", [])
        if not isinstance(legacy_page_entries, list) or not legacy_page_entries:
            raise ValueError(f"Flow manifest has no steps/pages: {flow_manifest_path}")
        step_entries = [
            {
                "order": page_entry["order"],
                "step_id": Path(str(page_entry["page_file"])).stem,
                "handler": "field_fill_page",
                "spec_file": page_entry["page_file"],
            }
            for page_entry in legacy_page_entries
        ]

    if not isinstance(step_entries, list) or not step_entries:
        raise ValueError(f"Flow manifest has no steps: {flow_manifest_path}")

    steps: list[FlowStepDefinition] = []
    for step_entry in step_entries:
        if not isinstance(step_entry, dict):
            raise ValueError(f"Invalid step entry in flow manifest: {flow_manifest_path}")
        spec_file = str(step_entry["spec_file"]).strip()
        step_path = flow_directory / spec_file
        if not step_path.exists():
            raise ValueError(f"Flow step JSON was not found: {step_path}")

        spec_payload = load_json_payload(step_path)
        steps.append(
            FlowStepDefinition(
                order=int(step_entry.get("order", spec_payload.get("order", 0))),
                step_id=str(step_entry.get("step_id", spec_payload.get("step_name", spec_file))),
                handler=str(step_entry.get("handler", spec_payload.get("handler", ""))),
                spec_file=spec_file,
                spec_payload=spec_payload,
            )
        )

    steps.sort(key=lambda step: step.order)
    return FlowDefinition(
        surface=str(payload.get("surface", surface)),
        product_type=str(payload.get("product_type", product_type)),
        flow_name=str(payload.get("flow_name", "listing_creation")),
        flow_directory=flow_directory,
        manifest_context=payload.get("context", {}) if isinstance(payload.get("context"), dict) else {},
        steps=steps,
    )


def build_page_definition_from_spec(
    step_definition: FlowStepDefinition,
) -> FlowPageDefinition:
    spec_payload = step_definition.spec_payload
    tab_payload = spec_payload.get("tab") if isinstance(spec_payload.get("tab"), dict) else {}
    data_source = (
        spec_payload.get("data_source")
        if isinstance(spec_payload.get("data_source"), dict)
        else {}
    )
    verification = (
        spec_payload.get("verification")
        if isinstance(spec_payload.get("verification"), dict)
        else None
    )
    upload_targets = (
        spec_payload.get("upload_targets")
        if isinstance(spec_payload.get("upload_targets"), dict)
        else {}
    )
    tab_locator_candidates = (
        tab_payload.get("locator_candidates")
        if isinstance(tab_payload, dict)
        else None
    )
    tab_xpath: str | None = None
    if isinstance(tab_locator_candidates, list):
        for candidate in tab_locator_candidates:
            if (
                isinstance(candidate, dict)
                and str(candidate.get("type", "")).lower() == "xpath"
                and candidate.get("value")
            ):
                tab_xpath = str(candidate["value"])
                break

    fields_payload = spec_payload.get("fields", [])
    return FlowPageDefinition(
        order=int(spec_payload.get("order", step_definition.order)),
        step_name=str(spec_payload.get("step_name", step_definition.step_id)),
        page_file=step_definition.spec_file,
        handler=step_definition.handler,
        tab_label=str(tab_payload.get("label")) if tab_payload.get("label") else None,
        tab_xpath=tab_xpath,
        checkpoint_label=(
            str(spec_payload["checkpoint_label"])
            if spec_payload.get("checkpoint_label")
            else None
        ),
        worksheet_name=(
            str(data_source["worksheet"])
            if isinstance(data_source, dict) and data_source.get("worksheet")
            else None
        ),
        field_json_config_attr=(
            str(spec_payload["field_json_config_attr"])
            if spec_payload.get("field_json_config_attr")
            else None
        ),
        excel_config_attr=(
            str(data_source["workbook_attr"])
            if isinstance(data_source, dict) and data_source.get("workbook_attr")
            else None
        ),
        log_stage=str(spec_payload["log_stage"]) if spec_payload.get("log_stage") else None,
        log_message=str(spec_payload["log_message"]) if spec_payload.get("log_message") else None,
        filled_checkpoint_label=(
            str(spec_payload["filled_checkpoint_label"])
            if spec_payload.get("filled_checkpoint_label")
            else None
        ),
        verify=verification,
        snapshot_name=(
            str(spec_payload["snapshot_name"])
            if spec_payload.get("snapshot_name")
            else None
        ),
        brand_name=str(spec_payload["brand_name"]) if spec_payload.get("brand_name") else None,
        locator_strategy=(
            str(spec_payload["locator_strategy"])
            if spec_payload.get("locator_strategy")
            else None
        ),
        source_snapshot=(
            str(spec_payload["source_snapshot"])
            if spec_payload.get("source_snapshot")
            else None
        ),
        upload_slot_ids=[
            str(slot_id)
            for slot_id in upload_targets.get("slot_ids", [])
            if str(slot_id).strip()
        ],
        fields=[
            build_field_definition(field_payload)
            for field_payload in fields_payload
            if isinstance(field_payload, dict)
        ],
    )


def resolve_runtime_reference(
    value: object,
    config: BotConfig,
    listing_selection: ListingSelection,
    flow_state: FlowState,
) -> object:
    if not isinstance(value, str) or not value.startswith("$"):
        return value

    if value.startswith("$context."):
        return flow_state.context.get(value.removeprefix("$context."))
    if value.startswith("$listing."):
        return getattr(listing_selection, value.removeprefix("$listing."))
    if value.startswith("$config."):
        return getattr(config, value.removeprefix("$config."))
    if value == "$state.price_fill_result":
        return flow_state.price_fill_result
    return value


def update_flow_context(
    flow_state: FlowState,
    save_to_context: object,
    saved_values: dict[str, object],
) -> None:
    if not isinstance(save_to_context, dict):
        return
    for context_key, source_key in save_to_context.items():
        if isinstance(source_key, str) and source_key in saved_values:
            flow_state.context[str(context_key)] = saved_values[source_key]


def flow_definition_saves_and_exits(flow_definition: FlowDefinition | None) -> bool:
    if flow_definition is None:
        return False

    for step in flow_definition.steps:
        if step.handler != "navigation_step":
            continue
        actions = step.spec_payload.get("actions", [])
        if not isinstance(actions, list):
            continue
        for action in actions:
            if (
                isinstance(action, dict)
                and str(action.get("type", "")).strip() == "click_save_and_go_back"
            ):
                return True
    return False


def run_navigation_step(
    driver: webdriver.Firefox,
    pause_controller: PauseController,
    config: BotConfig,
    listing_selection: ListingSelection,
    flow_state: FlowState,
    step_definition: FlowStepDefinition,
) -> None:
    actions = step_definition.spec_payload.get("actions", [])
    if not isinstance(actions, list):
        raise ValueError(f"Navigation step '{step_definition.step_id}' must define a list of actions.")

    for action in actions:
        if not isinstance(action, dict):
            raise ValueError(f"Navigation step '{step_definition.step_id}' has an invalid action entry.")
        action_type = str(action.get("type", "")).strip()
        checkpoint_label = (
            str(action["checkpoint_label"])
            if action.get("checkpoint_label")
            else None
        )

        if action_type == "open_listing_page":
            listing_url = resolve_runtime_reference(
                action.get("url", "$config.listing_url"),
                config,
                listing_selection,
                flow_state,
            )
            if not isinstance(listing_url, str):
                raise ValueError("open_listing_page action must resolve to a URL string.")
            log_event("NAV", f"Opening listing page: {listing_url}")
            open_listing_page(driver, listing_url)
            log_event("NAV", "Listing page opened in Firefox.")
        elif action_type == "dismiss_optional_ad_popup":
            dismiss_optional_ad_popup(
                driver,
                timeout_seconds=int(action.get("timeout_seconds", 5)),
            )
        elif action_type == "fill_brand_name":
            brand_name = resolve_runtime_reference(
                action.get("brand_name", "$context.brand_name"),
                config,
                listing_selection,
                flow_state,
            )
            if not isinstance(brand_name, str) or not brand_name.strip():
                raise ValueError("fill_brand_name action must resolve to a non-empty brand name.")
            fill_brand_name(driver, brand_name)
        elif action_type == "click_create_new_listing":
            click_create_new_listing(driver)
        elif action_type == "click_optional_continue":
            click_optional_continue(
                driver,
                timeout_seconds=int(action.get("timeout_seconds", 5)),
            )
        elif action_type == "click_save_and_go_back":
            click_save_and_go_back_button(driver)
            commit_pending_image_folder_exhaustion(flow_state)
        else:
            raise ValueError(
                f"Unsupported navigation action '{action_type}' in step '{step_definition.step_id}'."
            )

        if checkpoint_label:
            checkpoint_pause(pause_controller, checkpoint_label, driver, config)


def run_open_listing_bootstrap_step(
    driver: webdriver.Firefox,
    pause_controller: PauseController,
    config: BotConfig,
    listing_selection: ListingSelection,
    flow_definition: FlowDefinition | None,
) -> bool:
    if flow_definition is None:
        return False

    open_step = next(
        (
            step
            for step in flow_definition.steps
            if step.handler == "navigation_step" and step.step_id == "open_listing"
        ),
        None,
    )
    if open_step is None:
        return False

    flow_state = FlowState()
    flow_state.context.update(flow_definition.manifest_context)
    flow_state.context["brand_name"] = listing_selection.brand_name
    run_navigation_step(
        driver,
        pause_controller,
        config,
        listing_selection,
        flow_state,
        open_step,
    )
    return True


def resolve_preflight_listing_url(
    config: BotConfig,
    listing_selection: ListingSelection,
    flow_definition: FlowDefinition | None,
) -> str:
    if flow_definition is None:
        return config.listing_url

    open_step = next(
        (
            step
            for step in flow_definition.steps
            if step.handler == "navigation_step" and step.step_id == "open_listing"
        ),
        None,
    )
    if open_step is None:
        return config.listing_url

    actions = open_step.spec_payload.get("actions")
    if not isinstance(actions, list):
        return config.listing_url

    flow_state = FlowState()
    flow_state.context.update(flow_definition.manifest_context)
    flow_state.context["brand_name"] = listing_selection.brand_name

    for action in actions:
        if not isinstance(action, dict):
            continue
        if str(action.get("type", "")).strip() != "open_listing_page":
            continue
        resolved_url = resolve_runtime_reference(
            action.get("url", "$config.listing_url"),
            config,
            listing_selection,
            flow_state,
        )
        if isinstance(resolved_url, str) and resolved_url.strip():
            return resolved_url.strip()

    return config.listing_url


def run_login_precheck(
    config: BotConfig,
    listing_selection: ListingSelection,
    flow_definition: FlowDefinition | None,
) -> None:
    # Login precheck temporarily disabled.
    # target_url = resolve_preflight_listing_url(config, listing_selection, flow_definition)
    # check_script_path = PROJECT_ROOT / "check_logged_in.py"
    # command = [
    #     sys.executable,
    #     str(check_script_path),
    #     "--profile",
    #     config.profile_name,
    #     "--url",
    #     target_url,
    # ]
    # completed_process = subprocess.run(
    #     command,
    #     cwd=str(PROJECT_ROOT),
    #     check=False,
    # )
    # if completed_process.returncode != 0:
    #     raise RuntimeError(
    #         f"Login precheck failed before listing flow started (exit code {completed_process.returncode})."
    #     )
    return None


def open_flow_tab(
    driver: webdriver.Firefox,
    pause_controller: PauseController,
    config: BotConfig,
    tab_label: str,
    checkpoint_label: str,
    tab_xpath: str | None = None,
) -> None:
    resolved_tab_xpath = tab_xpath or TAB_XPATHS[tab_label]
    open_tab_via_autogui(driver, tab_label, resolved_tab_xpath)
    checkpoint_pause(pause_controller, checkpoint_label, driver, config)


def verify_flow_page_switch(
    driver: webdriver.Firefox,
    pause_controller: PauseController,
    config: BotConfig,
    cycle_label: str,
    tab_labels: tuple[str, str],
    checkpoint_label: str,
    tab_xpaths: tuple[str, str] | None = None,
) -> None:
    resolved_tab_sequence = []
    for index, tab_label in enumerate(tab_labels):
        tab_xpath = tab_xpaths[index] if tab_xpaths is not None else TAB_XPATHS[tab_label]
        resolved_tab_sequence.append((tab_label, tab_xpath))
    cycle_page_switch_verification_until_toast(
        driver,
        pause_controller,
        config,
        cycle_label=cycle_label,
        tab_sequence=resolved_tab_sequence,
    )
    checkpoint_pause(pause_controller, checkpoint_label, driver, config)


def run_page_verification_from_definition(
    driver: webdriver.Firefox,
    pause_controller: PauseController,
    config: BotConfig,
    page_definition: FlowPageDefinition,
) -> None:
    verify_payload = page_definition.verify
    if not verify_payload:
        return
    if not bool(verify_payload.get("enabled", True)):
        return

    tab_labels = verify_payload.get("tab_labels")
    if tab_labels is None and isinstance(verify_payload.get("tab_sequence"), list):
        tab_labels = verify_payload.get("tab_sequence")
    if not isinstance(tab_labels, list) or len(tab_labels) != 2:
        raise ValueError(
            f"Flow page '{page_definition.step_name}' must define exactly two verification tabs."
        )

    cycle_label = str(
        verify_payload.get("cycle_label")
        or verify_payload.get("page_name")
        or page_definition.step_name
    )
    checkpoint_label = str(
        verify_payload.get("checkpoint_label")
        or f"Verification completed for {page_definition.step_name}"
    )

    tab_xpaths_payload = verify_payload.get("tab_xpaths")
    tab_xpaths: tuple[str, str] | None = None
    if isinstance(tab_xpaths_payload, list) and len(tab_xpaths_payload) == 2:
        tab_xpaths = (str(tab_xpaths_payload[0]), str(tab_xpaths_payload[1]))

    verify_flow_page_switch(
        driver,
        pause_controller,
        config,
        cycle_label,
        (str(tab_labels[0]), str(tab_labels[1])),
        checkpoint_label,
        tab_xpaths=tab_xpaths,
    )


def load_page_input_row(
    config: BotConfig,
    listing_selection: ListingSelection,
    page_definition: FlowPageDefinition,
) -> ProductInputRow:
    if not page_definition.excel_config_attr:
        raise ValueError(f"Flow page '{page_definition.step_name}' is missing excel_config_attr.")
    workbook_path = getattr(config, page_definition.excel_config_attr)
    return load_product_input_row(
        workbook_path,
        listing_selection.kind,
        listing_selection.size,
        worksheet_name=page_definition.worksheet_name,
        surface=listing_selection.surface,
    )


def load_page_field_definitions(
    config: BotConfig,
    page_definition: FlowPageDefinition,
) -> list[FieldDefinition]:
    if page_definition.fields:
        return page_definition.fields
    if not page_definition.field_json_config_attr:
        raise ValueError(
            f"Flow page '{page_definition.step_name}' is missing field_json_config_attr."
        )
    return load_field_definitions(getattr(config, page_definition.field_json_config_attr))


def run_additional_description_flow_step(
    driver: webdriver.Firefox,
    pause_controller: PauseController,
    config: BotConfig,
    listing_selection: ListingSelection,
    page_definition: FlowPageDefinition | None = None,
    verify_after: bool = True,
) -> None:
    page_definition = page_definition or FlowPageDefinition(
        order=1,
        step_name="additional_description",
        page_file="",
        handler="additional_description",
        tab_label="Additional Description",
        checkpoint_label="Additional Description tab opened",
        worksheet_name=get_additional_description_sheet_name(listing_selection.product_type),
        field_json_config_attr="additional_description_json",
        excel_config_attr="additional_description_excel",
        log_stage="ADDL",
        log_message="Starting Additional Description field fill from Excel + JSON mapping...",
        filled_checkpoint_label="Additional Description fields filled",
        verify={
            "enabled": verify_after,
            "cycle_label": "Additional Description page",
            "tab_labels": ["Additional Description", "Product Description"],
            "checkpoint_label": "Changes saved detected after Additional Description click cycle",
        },
    )
    open_flow_tab(
        driver,
        pause_controller,
        config,
        page_definition.tab_label or "Additional Description",
        page_definition.checkpoint_label or "Additional Description tab opened",
        tab_xpath=page_definition.tab_xpath,
    )
    additional_description_row = load_page_input_row(config, listing_selection, page_definition)
    additional_description_field_definitions = load_page_field_definitions(config, page_definition)
    log_event(
        "DATA",
        f"Loaded Additional Description row: kind={additional_description_row.kind}, "
        f"size={additional_description_row.size}",
    )
    log_event(
        page_definition.log_stage or "ADDL",
        page_definition.log_message or "Starting Additional Description field fill from Excel + JSON mapping...",
    )
    fill_additional_description_fields(
        driver,
        additional_description_field_definitions,
        additional_description_row,
    )
    checkpoint_pause(
        pause_controller,
        page_definition.filled_checkpoint_label or "Additional Description fields filled",
        driver,
        config,
    )

    if verify_after:
        log_event(
            "VERIFY",
            "Additional Description filling completed. Starting VERIFYING CHANGES CYCLE.",
        )
        run_page_verification_from_definition(driver, pause_controller, config, page_definition)


def run_product_description_flow_step(
    driver: webdriver.Firefox,
    pause_controller: PauseController,
    config: BotConfig,
    listing_selection: ListingSelection,
    page_definition: FlowPageDefinition | None = None,
) -> None:
    page_definition = page_definition or FlowPageDefinition(
        order=2,
        step_name="product_description",
        page_file="",
        handler="product_description",
        tab_label="Product Description",
        checkpoint_label="Product Description tab opened",
        worksheet_name=get_product_description_sheet_name(listing_selection.product_type),
        field_json_config_attr="product_description_json",
        excel_config_attr="product_description_excel",
        log_stage="DESC",
        log_message="Starting Product Description field fill from Excel + JSON mapping...",
        filled_checkpoint_label="Product Description fields filled",
        verify={
            "enabled": True,
            "cycle_label": "Product Description page",
            "tab_labels": ["Product Description", "Price, Stock and Shipping Information"],
            "checkpoint_label": "Changes saved detected after Product Description click cycle",
        },
    )
    open_flow_tab(
        driver,
        pause_controller,
        config,
        page_definition.tab_label or "Product Description",
        page_definition.checkpoint_label or "Product Description tab opened",
        tab_xpath=page_definition.tab_xpath,
    )
    product_description_row = load_page_input_row(config, listing_selection, page_definition)
    product_description_field_definitions = load_page_field_definitions(config, page_definition)
    log_event(
        "DATA",
        f"Loaded Product Description row: kind={product_description_row.kind}, "
        f"size={product_description_row.size}",
    )
    log_event(
        page_definition.log_stage or "DESC",
        page_definition.log_message or "Starting Product Description field fill from Excel + JSON mapping...",
    )
    fill_product_description_fields(
        driver,
        product_description_field_definitions,
        product_description_row,
    )
    checkpoint_pause(
        pause_controller,
        page_definition.filled_checkpoint_label or "Product Description fields filled",
        driver,
        config,
    )
    log_event(
        "VERIFY",
        "Product Description filling completed. Starting VERIFYING CHANGES CYCLE.",
    )
    run_page_verification_from_definition(driver, pause_controller, config, page_definition)


def run_price_stock_shipping_flow_step(
    driver: webdriver.Firefox,
    pause_controller: PauseController,
    config: BotConfig,
    listing_selection: ListingSelection,
    page_definition: FlowPageDefinition | None = None,
) -> FillResult:
    page_definition = page_definition or FlowPageDefinition(
        order=3,
        step_name="price_stock_shipping",
        page_file="",
        handler="price_stock_shipping",
        tab_label="Price, Stock and Shipping Information",
        checkpoint_label="Selling info tab opened",
        field_json_config_attr="price_stock_shipping_json",
        excel_config_attr="price_stock_shipping_excel",
        filled_checkpoint_label="Price/Stock/Shipping fields filled",
        snapshot_name=PHASE_ONE_SNAPSHOT_NAME,
        verify={
            "enabled": True,
            "cycle_label": "SKU page",
            "tab_labels": ["Price, Stock and Shipping Information", "Image addition"],
            "checkpoint_label": "Changes saved detected after SKU-page click cycle",
        },
    )
    open_flow_tab(
        driver,
        pause_controller,
        config,
        page_definition.tab_label or "Price, Stock and Shipping Information",
        page_definition.checkpoint_label or "Selling info tab opened",
        tab_xpath=page_definition.tab_xpath,
    )
    phase_one_snapshot_path = save_named_html_snapshot(
        driver,
        config.snapshot_directory,
        page_definition.snapshot_name or PHASE_ONE_SNAPSHOT_NAME,
    )
    log_event("SNAPSHOT", f"Saved phase snapshot: {phase_one_snapshot_path}")
    product_input_row = load_page_input_row(config, listing_selection, page_definition)
    field_definitions = load_page_field_definitions(config, page_definition)
    log_event(
        "DATA",
        f"Loaded Price/Stock/Shipping row for filling: kind={product_input_row.kind}, "
        f"size={product_input_row.size}",
    )
    price_fill_result = fill_price_stock_shipping_fields(
        driver,
        field_definitions,
        product_input_row,
    )
    expected_field_values = build_expected_field_values(
        field_definitions,
        product_input_row,
        price_fill_result,
    )
    verify_and_refill_price_stock_shipping_fields(
        driver,
        field_definitions,
        expected_field_values,
        price_fill_result.skipped_fields,
    )
    checkpoint_pause(
        pause_controller,
        page_definition.filled_checkpoint_label or "Price/Stock/Shipping fields filled",
        driver,
        config,
    )
    log_event(
        "VERIFY",
        "Price/Stock/Shipping verification completed. Starting VERIFYING CHANGES CYCLE.",
    )
    run_page_verification_from_definition(driver, pause_controller, config, page_definition)
    return price_fill_result


def run_images_flow_step(
    driver: webdriver.Firefox,
    pause_controller: PauseController,
    config: BotConfig,
    listing_selection: ListingSelection,
    verify_before_variants: bool,
    flow_state: FlowState,
    page_definition: FlowPageDefinition | None = None,
) -> None:
    page_definition = page_definition or FlowPageDefinition(
        order=4,
        step_name="images",
        page_file="",
        handler="images",
        tab_label="Image addition",
        checkpoint_label="Images tab opened",
        filled_checkpoint_label="Images uploaded",
        brand_name=listing_selection.brand_name,
        verify={
            "enabled": verify_before_variants,
            "cycle_label": "Images page",
            "tab_labels": ["Image addition", "Variant addition"],
            "checkpoint_label": "Changes saved detected after Images page click cycle",
        },
    )
    open_flow_tab(
        driver,
        pause_controller,
        config,
        page_definition.tab_label or "Image addition",
        page_definition.checkpoint_label or "Images tab opened",
        tab_xpath=page_definition.tab_xpath,
    )
    brand_name = page_definition.brand_name or listing_selection.brand_name
    selected_image_folder = preview_selected_image_folder(
        config,
        brand_name,
        listing_selection.surface,
    )
    if selected_image_folder is None:
        raise ValueError(
            f"Images step could not find an available image folder for "
            f"{brand_name}/{listing_selection.surface} in {config.image_directory}."
        )

    checkpoint_pause(pause_controller, "Image folder selected", driver, config)
    upload_image_folder(
        driver,
        selected_image_folder,
        brand_name,
        listing_selection.surface,
        pause_controller,
        config,
        slot_ids=page_definition.upload_slot_ids or IMAGE_SLOT_IDS,
    )
    queue_image_folder_exhaustion(
        flow_state,
        selected_image_folder,
        brand_name,
        listing_selection.surface,
    )
    checkpoint_pause(
        pause_controller,
        page_definition.filled_checkpoint_label or "Images uploaded",
        driver,
        config,
    )

    if verify_before_variants:
        log_event(
            "VERIFY",
            "Images step completed. Starting VERIFYING CHANGES CYCLE before Variant page.",
        )
        run_page_verification_from_definition(driver, pause_controller, config, page_definition)


def run_variants_flow_step(
    driver: webdriver.Firefox,
    pause_controller: PauseController,
    config: BotConfig,
    listing_selection: ListingSelection,
    price_fill_result: FillResult | None,
    source_sku_override: str | None = None,
    page_definition: FlowPageDefinition | None = None,
) -> None:
    if price_fill_result is None and not source_sku_override:
        raise ValueError("Variant step needs the Price/Stock/Shipping fill result.")

    page_definition = page_definition or FlowPageDefinition(
        order=5,
        step_name="variants",
        page_file="",
        handler="variants",
        tab_label="Variant addition",
        checkpoint_label="Variant tab opened",
        worksheet_name=get_variants_sheet_name(listing_selection.product_type),
        excel_config_attr="variants_excel",
        filled_checkpoint_label="Variant rows created",
    )
    open_flow_tab(
        driver,
        pause_controller,
        config,
        page_definition.tab_label or "Variant addition",
        page_definition.checkpoint_label or "Variant tab opened",
        tab_xpath=page_definition.tab_xpath,
    )
    variant_row = load_page_input_row(config, listing_selection, page_definition)
    log_event(
        "DATA",
        f"Loaded Variant row: kind={variant_row.kind}, size={variant_row.size}",
    )
    log_event(
        page_definition.log_stage or "VARIANT",
        page_definition.log_message or "Starting Variant page creation loop from Excel mapping...",
    )
    source_sku = source_sku_override or ""
    if price_fill_result is not None:
        source_sku = price_fill_result.generated_values.get("Seller SKU ID", source_sku)
    fill_variant_page(
        config,
        listing_selection,
        driver,
        pause_controller,
        variant_row,
        source_sku,
    )
    checkpoint_pause(
        pause_controller,
        page_definition.filled_checkpoint_label or "Variant rows created",
        driver,
        config,
    )


def run_flow_step_from_definition(
    driver: webdriver.Firefox,
    pause_controller: PauseController,
    config: BotConfig,
    listing_selection: ListingSelection,
    step_definition: FlowStepDefinition,
    flow_state: FlowState,
    flow_step_ids: set[str],
) -> None:
    if step_definition.handler == "navigation_step":
        run_navigation_step(
            driver,
            pause_controller,
            config,
            listing_selection,
            flow_state,
            step_definition,
        )
        return

    page_definition = build_page_definition_from_spec(step_definition)
    verification_payload = (
        page_definition.verify if isinstance(page_definition.verify, dict) else {}
    )

    if step_definition.handler == "field_fill_page":
        field_fill_mode = str(step_definition.spec_payload.get("field_fill_mode", "")).strip()
        if field_fill_mode == "additional_description":
            verify_after = bool(verification_payload.get("enabled", True))
            run_additional_description_flow_step(
                driver,
                pause_controller,
                config,
                listing_selection,
                page_definition=page_definition,
                verify_after=verify_after,
            )
        elif field_fill_mode == "product_description":
            run_product_description_flow_step(
                driver,
                pause_controller,
                config,
                listing_selection,
                page_definition=page_definition,
            )
        elif field_fill_mode == "price_stock_shipping":
            flow_state.price_fill_result = run_price_stock_shipping_flow_step(
                driver,
                pause_controller,
                config,
                listing_selection,
                page_definition=page_definition,
            )
            if flow_state.price_fill_result.generated_values:
                saved_values: dict[str, object] = {}
                for generated_label, generated_value in flow_state.price_fill_result.generated_values.items():
                    default_context_key = f"{page_definition.step_name}.generated.{generated_label}"
                    flow_state.context[default_context_key] = generated_value
                    log_event(
                        "FLOW",
                        f"Saved generated value to context: {default_context_key} = {generated_value}",
                    )
                    saved_values[generated_label] = generated_value
                update_flow_context(
                    flow_state,
                    step_definition.spec_payload.get("save_to_context"),
                    saved_values,
                )
        else:
            raise ValueError(
                f"Unsupported field_fill_mode '{field_fill_mode}' in step '{step_definition.step_id}'."
            )
    elif step_definition.handler == "image_upload_page":
        image_source_payload = (
            step_definition.spec_payload.get("image_source")
            if isinstance(step_definition.spec_payload.get("image_source"), dict)
            else {}
        )
        if image_source_payload.get("brand_name_from"):
            resolved_brand_name = resolve_runtime_reference(
                image_source_payload.get("brand_name_from"),
                config,
                listing_selection,
                flow_state,
            )
            if isinstance(resolved_brand_name, str) and resolved_brand_name.strip():
                page_definition.brand_name = resolved_brand_name
        verify_before_variants = bool(
            verification_payload.get("enabled", "variants" in flow_step_ids)
        )
        run_images_flow_step(
            driver,
            pause_controller,
            config,
            listing_selection,
            verify_before_variants=verify_before_variants,
            flow_state=flow_state,
            page_definition=page_definition,
        )
    elif step_definition.handler == "variant_page":
        source_sku_override = resolve_runtime_reference(
            step_definition.spec_payload.get(
                "source_sku_from",
                "$context.price_stock_shipping.generated.Seller SKU ID",
            ),
            config,
            listing_selection,
            flow_state,
        )
        run_variants_flow_step(
            driver,
            pause_controller,
            config,
            listing_selection,
            flow_state.price_fill_result,
            source_sku_override=str(source_sku_override) if source_sku_override else None,
            page_definition=page_definition,
        )
    else:
        raise ValueError(f"Unknown flow handler: {step_definition.handler}")


def run_flow_page_from_definition(
    driver: webdriver.Firefox,
    pause_controller: PauseController,
    config: BotConfig,
    listing_selection: ListingSelection,
    page_definition: FlowPageDefinition,
    flow_state: FlowState,
    flow_step_names: set[str],
) -> None:
    # Legacy compatibility wrapper for the older hardcoded page-only flow path.
    if page_definition.handler == "additional_description":
        verify_after = bool((page_definition.verify or {}).get("enabled", True))
        run_additional_description_flow_step(
            driver,
            pause_controller,
            config,
            listing_selection,
            page_definition=page_definition,
            verify_after=verify_after,
        )
    elif page_definition.handler == "product_description":
        run_product_description_flow_step(
            driver,
            pause_controller,
            config,
            listing_selection,
            page_definition=page_definition,
        )
    elif page_definition.handler == "price_stock_shipping":
        flow_state.price_fill_result = run_price_stock_shipping_flow_step(
            driver,
            pause_controller,
            config,
            listing_selection,
            page_definition=page_definition,
        )
    elif page_definition.handler == "images":
        run_images_flow_step(
            driver,
            pause_controller,
            config,
            listing_selection,
            page_definition=page_definition,
            verify_before_variants="variants" in flow_step_names,
            flow_state=flow_state,
        )
    elif page_definition.handler == "variants":
        run_variants_flow_step(
            driver,
            pause_controller,
            config,
            listing_selection,
            flow_state.price_fill_result,
            page_definition=page_definition,
        )
    else:
        raise ValueError(f"Unknown flow handler: {page_definition.handler}")


def run_listing_page_flow(
    driver: webdriver.Firefox,
    pause_controller: PauseController,
    config: BotConfig,
    listing_selection: ListingSelection,
) -> FlowState:
    flow_definition = load_listing_flow_definition(
        listing_selection.product_type,
        listing_selection.surface,
    )
    flow_state = FlowState()
    if flow_definition is not None:
        flow_state.context.update(flow_definition.manifest_context)
        flow_state.context["brand_name"] = listing_selection.brand_name
        flow_step_ids = {step.step_id for step in flow_definition.steps}
        log_event(
            "FLOW",
            f"Running JSON flow {flow_definition.flow_name} for "
            f"{flow_definition.product_type}/{flow_definition.surface}: "
            f"{' -> '.join(step.step_id for step in flow_definition.steps)}",
        )
        for step_definition in flow_definition.steps:
            run_flow_step_from_definition(
                driver,
                pause_controller,
                config,
                listing_selection,
                step_definition,
                flow_state,
                flow_step_ids,
            )
        return flow_state

    flow_steps = LEGACY_PRODUCT_PAGE_FLOWS[listing_selection.product_type]
    log_event(
        "FLOW",
        f"Running legacy {listing_selection.product_type} flow: {' -> '.join(flow_steps)}",
    )
    for step_name in flow_steps:
        run_flow_page_from_definition(
            driver,
            pause_controller,
            config,
            listing_selection,
            FlowPageDefinition(order=0, step_name=step_name, page_file="", handler=step_name),
            flow_state,
            set(flow_steps),
        )
    return flow_state


def print_runtime_context(config: BotConfig) -> None:
    log_event("BOOT", "Flipkart lister bot starting...")
    log_event("BOOT", f"Target URL: {config.listing_url}")
    log_event("BOOT", f"Firefox profile: {config.profile_name}")
    log_event("BOOT", f"Firefox profile path: {config.firefox_profile_path}")
    log_event(
        "BOOT",
        f"Image directory: {config.image_directory if str(config.image_directory) else 'not configured yet'}",
    )
    log_event(
        "BOOT",
        f"Data directory: {config.data_directory if str(config.data_directory) else 'not configured yet'}",
    )
    log_event("BOOT", f"Price/Stock/Shipping Excel: {config.price_stock_shipping_excel}")
    log_event("BOOT", f"Price/Stock/Shipping JSON: {config.price_stock_shipping_json}")
    log_event("BOOT", f"Product Description Excel: {config.product_description_excel}")
    log_event("BOOT", f"Product Description JSON: {config.product_description_json}")
    log_event("BOOT", f"Additional Description Excel: {config.additional_description_excel}")
    log_event("BOOT", f"Additional Description JSON: {config.additional_description_json}")
    log_event("BOOT", f"Variants Excel: {config.variants_excel}")
    log_event("BOOT", f"Snapshot directory: {config.snapshot_directory}")
    log_event("BOOT", "Pause control: press Space in this terminal to pause at the next safe step.")


def run_single_listing_session(
    config: BotConfig,
    listing_selection: ListingSelection,
    json_flow_definition: FlowDefinition | None,
    run_index: int,
    total_runs: int,
) -> bool:
    set_current_run_label(f"run {run_index}/{total_runs}")
    try:
        log_event("BOOT", f"Launching Firefox WebDriver for run {run_index}/{total_runs}...")
        driver = build_firefox_driver(config)
        log_event("BOOT", f"Firefox WebDriver launched successfully for run {run_index}/{total_runs}.")
    except WebDriverException as exc:
        log_event(
            "ERROR",
            "Could not start Firefox WebDriver. Make sure Firefox is installed, the selected "
            "profile is valid, and that the same profile is not already open in another Firefox "
            "window. You can also set GECKODRIVER_PATH/FIREFOX_BINARY if needed.",
        )
        error_message = f"Run {run_index}/{total_runs} failed before browser launch: {exc}"
        write_latest_error(error_message)
        log_event("ERROR", error_message)
        return False

    pause_controller = PauseController()
    pause_controller.start()

    try:
        log_event("RUN", f"Starting run {run_index} of {total_runs}.")
        # Additional Description test-only flow kept commented out intentionally.
        # if additional_test_run_only:
        #     used_json_open_step = run_open_listing_bootstrap_step(
        #         driver,
        #         pause_controller,
        #         config,
        #         listing_selection,
        #         json_flow_definition,
        #     )
        #     if not used_json_open_step:
        #         log_event("NAV", f"Opening listing page: {config.listing_url}")
        #         open_listing_page(driver, config.listing_url)
        #         log_event("NAV", "Listing page opened in Firefox.")
        #         checkpoint_pause(pause_controller, "Listing page opened", driver, config)
        #         dismiss_optional_ad_popup(driver)
        #         checkpoint_pause(pause_controller, "Optional popup handling complete", driver, config)
        #         fill_brand_name(driver, listing_selection.brand_name)
        #         checkpoint_pause(pause_controller, "Brand entered", driver, config)
        #         click_create_new_listing(driver)
        #         checkpoint_pause(pause_controller, "Create new listing clicked", driver, config)
        #         click_optional_continue(driver)
        #         checkpoint_pause(pause_controller, "Optional continue handling complete", driver, config)
        #     log_event(
        #         "TEST",
        #         "Additional Description test mode is enabled. Skipping image upload, "
        #         "SKU page, and Product Description filling.",
        #     )
        #     run_additional_description_flow_step(
        #         driver,
        #         pause_controller,
        #         config,
        #         listing_selection,
        #         verify_after=False,
        #     )
        #     log_event(
        #         "DONE",
        #         f"Additional Description test flow completed for run {run_index}/{total_runs}.",
        #     )
        #     return

        if json_flow_definition is None:
            log_event("NAV", f"Opening listing page: {config.listing_url}")
            open_listing_page(driver, config.listing_url)
            log_event("NAV", "Listing page opened in Firefox.")
            checkpoint_pause(pause_controller, "Listing page opened", driver, config)
            dismiss_optional_ad_popup(driver)
            checkpoint_pause(pause_controller, "Optional popup handling complete", driver, config)
            fill_brand_name(driver, listing_selection.brand_name)
            checkpoint_pause(pause_controller, "Brand entered", driver, config)
            click_create_new_listing(driver)
            checkpoint_pause(pause_controller, "Create new listing clicked", driver, config)
            click_optional_continue(driver)
            checkpoint_pause(pause_controller, "Optional continue handling complete", driver, config)

        flow_state = run_listing_page_flow(driver, pause_controller, config, listing_selection)
        if flow_definition_saves_and_exits(json_flow_definition):
            log_event("DONE", "JSON flow already clicked Save & Go Back; skipping legacy final save click.")
        else:
            click_save_and_go_back_button(driver)
            commit_pending_image_folder_exhaustion(flow_state)
        log_event("DONE", f"{listing_selection.product_type.title()} flow completed for run {run_index}/{total_runs}.")
        log_event("BOOT", f"Waiting {SUCCESS_CLOSE_DELAY_SECONDS} seconds before closing browser.")
        sleep(SUCCESS_CLOSE_DELAY_SECONDS)
        return True
    except Exception as exc:
        try:
            snapshot_path = save_html_snapshot(
                driver,
                config.snapshot_directory,
                f"run {run_index} error before close",
            )
            log_event("ERROR", f"Saved failure snapshot before closing browser: {snapshot_path}")
        except Exception as snapshot_error:
            log_event("ERROR", f"Could not save failure snapshot: {snapshot_error}")
        error_message = f"Run {run_index}/{total_runs} failed: {exc}"
        write_latest_error(error_message)
        log_event("ERROR", error_message)
        log_event("RUN", f"Aborting current run {run_index}/{total_runs}; continuing with next run if available.")
        return False
    finally:
        pause_controller.stop()
        driver.quit()
        log_event("BOOT", f"Closed browser for run {run_index}/{total_runs}.")


def main() -> None:
    try:
        selected_profile = prompt_for_profile()
        run_count = prompt_for_run_count()
        listing_selection = prompt_for_listing_selection(selected_profile)
        json_flow_definition = load_listing_flow_definition(
            listing_selection.product_type,
            listing_selection.surface,
        )
        config = BotConfig(
            profile_name=selected_profile,
            image_directory=listing_selection.image_directory,
            product_description_excel=get_product_description_excel_path(listing_selection.product_type,listing_selection.surface),
            product_description_json=get_product_description_json_path(listing_selection.product_type,listing_selection.surface),
            additional_description_excel=get_additional_description_excel_path(listing_selection.product_type,listing_selection.surface),
            additional_description_json=get_additional_description_json_path(listing_selection.product_type,listing_selection.surface),
        )
        run_login_precheck(config, listing_selection, json_flow_definition)
        print_runtime_context(config)
        log_event("BOOT", f"Configured run count: {run_count}")
        log_event(
            "BOOT",
            f"Listing selection: type={listing_selection.product_type}, surface={listing_selection.surface}, "
            f"kind={listing_selection.kind}, size={listing_selection.size}, brand={listing_selection.brand_name}",
        )
        log_event(
            "BOOT",
            "JSON flow mode: "
            f"{'enabled' if json_flow_definition is not None else 'disabled'}",
        )
    except (ValueError, RuntimeError) as exc:
        raise SystemExit(str(exc)) from exc

    completed_runs = 0
    failed_runs = 0
    for run_index in range(1, run_count + 1):
        run_succeeded = run_single_listing_session(
            config,
            listing_selection,
            json_flow_definition,
            run_index,
            run_count,
        )
        if run_succeeded:
            completed_runs += 1
            try:
                record_path = record_successful_run(listing_selection, selected_profile)
                log_event(
                    "DONE",
                    f"Recorded successful run in Excel: {record_path.name}",
                )
            except Exception as exc:
                log_event("ERROR", f"Could not update successful run record Excel: {exc}")
        else:
            failed_runs += 1

    set_current_run_label("summary")
    log_event(
        "DONE",
        f"Batch finished. Successful run(s): {completed_runs}. Failed run(s): {failed_runs}.",
    )


if __name__ == "__main__":
    main()

