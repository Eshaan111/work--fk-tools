from __future__ import annotations

import csv
import json
import os
import re
import shutil
import subprocess
import time
import tkinter as tk
import urllib.request
import winsound
from dataclasses import dataclass
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import unquote, urlparse

import pyautogui
from openpyxl import load_workbook
from pynput import keyboard


# LAPTOP_NAME = "VAIO"
# LAPTOP_NAME = "ASUS"
LAPTOP_NAME = "HP"

PRABHU_FIREFOX_PROFILE_ASUS = Path(
    r"C:\Users\ESHAAN\Documents\Firefox-Profiles\0xe7h0bx.prabhu"
)
PRABHU_FIREFOX_PROFILE_VAIO = Path(
    r"C:\Users\SONY\AppData\Roaming\Mozilla\Firefox\Profiles\gm1pmawk.default-release"
)
PRABHU_FIREFOX_PROFILE_hp = Path(
    r"C:\Users\HP\AppData\Roaming\Mozilla\Firefox\Profiles\y6d16t5c.default-release"
)
IMAGE_PROMPTER_ROOT = Path(__file__).resolve().parent
ROOT_PATH = IMAGE_PROMPTER_ROOT


def image_prompter_path(*relative_parts: str) -> Path:
    return ROOT_PATH.joinpath(*relative_parts)


RUN_HELPERS_DIR = image_prompter_path("run-helpers")
FULL_GENERATED_IMAGES_DIR = image_prompter_path("FULL GENERATED IMAGES")
IMAGES_FINAL_DIR = image_prompter_path("IMAGES-FINAL")

PIXELS_FILE_ASUS = image_prompter_path("pixels-ASUS.json")
PIXELS_FILE_VAIO = image_prompter_path("pixels-VAIO.json")
PIXELS_FILE_HP = image_prompter_path("pixels-HP.json")
NO_BG_IMAGES_ROOT_ASUS = Path(r"C:\work-mom\NO-BG-IMAGES")
NO_BG_IMAGES_ROOT_VAIO = Path(r"C:\NO-BG-IMAGES")
NO_BG_IMAGES_ROOT_HP = Path(r"C:\NO-BG-IMAGES")

PRODUCT_IMAGE_FOLDER_OVERRIDES = {
    "WHITE-BAGGY-JEANS": "White-Baggy",
}
LAPTOP_CONFIGS = {
    "ASUS": {
        "firefox_profile": PRABHU_FIREFOX_PROFILE_ASUS,
        "no_bg_images_root": NO_BG_IMAGES_ROOT_ASUS,
        "pixel_json" : PIXELS_FILE_ASUS
    },
    "VAIO": {
        "firefox_profile": PRABHU_FIREFOX_PROFILE_VAIO,
        "no_bg_images_root": NO_BG_IMAGES_ROOT_VAIO,
        "pixel_json" : PIXELS_FILE_VAIO
    },
    "HP" : {
        "firefox_profile": PRABHU_FIREFOX_PROFILE_hp,
        "no_bg_images_root": NO_BG_IMAGES_ROOT_HP,
        "pixel_json" : PIXELS_FILE_HP
    }
}
ACTIVE_LAPTOP_CONFIG = LAPTOP_CONFIGS[LAPTOP_NAME.upper()]
PRABHU_FIREFOX_PROFILE = ACTIVE_LAPTOP_CONFIG["firefox_profile"]
NO_BG_IMAGES_ROOT = ACTIVE_LAPTOP_CONFIG["no_bg_images_root"]
PIXELS_FILE = ACTIVE_LAPTOP_CONFIG["pixel_json"]
USED_IMAGE_DESIGNS_WORKBOOK = image_prompter_path("USED-IMAGE-DESIGNS.xlsx")
PROMPT_TEMPLATE_PATH = image_prompter_path("image_edit_prompt_template.txt")
IMAGE_GENERATION_PROMPT_TEMPLATE_PATH = image_prompter_path("image_generation_prompt")
IMAGE_GENERATION_ABORT_PHRASES_PATH = image_prompter_path(
    "image_generation_abort_phrases.json"
)
DEFAULT_FIREFOX_BINARY = Path(r"C:\Program Files\Mozilla Firefox\firefox.exe")
FALLBACK_FIREFOX_BINARIES = [
    Path(r"C:\Program Files\Mozilla Firefox\firefox.exe"),
    Path(r"C:\Program Files (x86)\Mozilla Firefox\firefox.exe"),
]
PROMPT_PREVIEW_PATH = RUN_HELPERS_DIR / "generated_prompt_preview.txt"
LAST_FULL_CHAT_PATH = RUN_HELPERS_DIR / "last_full_chat.txt"
LATEST_RESPONSE_PATH = RUN_HELPERS_DIR / "latest_response.txt"
ALL_RESPONSES_PATH = RUN_HELPERS_DIR / "all_responses.txt"
PARSED_IDEAS_PATH = RUN_HELPERS_DIR / "parsed_latest_ideas.txt"
NEW_IDEAS_PATH = RUN_HELPERS_DIR / "new_ideas_not_in_excel.txt"
CURRENT_RUN_IDEA_PATH = RUN_HELPERS_DIR / "current_run_idea.json"
CURRENT_GENERATION_PROMPT_PATH = RUN_HELPERS_DIR / "current_generation_prompt.txt"
IMAGE_GENERATION_FINAL_CHAT_PATH = RUN_HELPERS_DIR / "image_generation_final_chat.txt"
IMAGE_GENERATION_TIMES_CSV_PATH = RUN_HELPERS_DIR / "image_generation_times.csv"
RUN_HELPER_PATHS = (
    PROMPT_PREVIEW_PATH,
    LAST_FULL_CHAT_PATH,
    LATEST_RESPONSE_PATH,
    ALL_RESPONSES_PATH,
    PARSED_IDEAS_PATH,
    NEW_IDEAS_PATH,
    CURRENT_RUN_IDEA_PATH,
    CURRENT_GENERATION_PROMPT_PATH,
    IMAGE_GENERATION_FINAL_CHAT_PATH,
)
LEGACY_RUN_HELPER_PATHS = (
    image_prompter_path("generated_prompt_preview.txt"),
    image_prompter_path("last_full_chat.txt"),
    image_prompter_path("latest_response.txt"),
    image_prompter_path("all_responses.txt"),
    image_prompter_path("parsed_latest_ideas.txt"),
    image_prompter_path("new_ideas_not_in_excel.txt"),
    image_prompter_path("current_run_idea.json"),
    image_prompter_path("current_generation_prompt.txt"),
    image_prompter_path("image_generation_final_chat.txt"),
)
CHATGPT_URL = "https://chatgpt.com"
IDEA_MARKER = "IDEA FOR BACKGROUND :"
SUPPORTED_IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp"}

pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.15
START_HOTKEY_KEY = keyboard.Key.right


def load_pixel_config() -> dict[str, object]:
    if not PIXELS_FILE.exists():
        raise FileNotFoundError(
            f"pixels.json not found at: {PIXELS_FILE}\n"
            "Run set_pixels.py first to capture your screen coordinates."
        )
    raw = json.loads(PIXELS_FILE.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError(f"{PIXELS_FILE.name} must contain a JSON object.")
    return raw


def load_pixels(raw: dict[str, object]) -> dict[str, tuple[int, int]]:
    return {
        key: (int(value["x"]), int(value["y"]))
        for key, value in raw.items()
        if isinstance(value, dict) and "x" in value and "y" in value
    }


def load_chatgpt_answering_pixel(
    raw: dict[str, object],
) -> tuple[tuple[int, int], tuple[int, int, int], int] | None:
    value = raw.get("chatgpt_answering", raw.get("generation_complete"))
    if value is None:
        return None
    if not isinstance(value, dict):
        raise ValueError("chatgpt_answering must be a JSON object.")

    position = value.get("position")
    rgb = value.get("rgb")
    tolerance = int(value.get("tolerance", 0))
    if not isinstance(position, list) or len(position) != 2:
        raise ValueError("chatgpt_answering.position must be [x, y].")
    if not isinstance(rgb, list) or len(rgb) != 3:
        raise ValueError("chatgpt_answering.rgb must be [red, green, blue].")
    if tolerance < 0:
        raise ValueError("chatgpt_answering.tolerance cannot be negative.")

    return (
        (int(position[0]), int(position[1])),
        (int(rgb[0]), int(rgb[1]), int(rgb[2])),
        tolerance,
    )


_PIXEL_CONFIG = load_pixel_config()
_PIXELS = load_pixels(_PIXEL_CONFIG)
CHATGPT_ANSWERING_PIXEL = load_chatgpt_answering_pixel(_PIXEL_CONFIG)
CHAT_CLICK_TARGET: tuple[int, int] = _PIXELS["chat_neutral_click"]
CHATGPT_PROMPT_BOX_PIXELS_VAIO = {
    "position": _PIXELS["prompt_box_initial"],
    "rgb": (230, 255, 255),
}
CHATGPT_PROMPT_BOX_PIXELS_VAIO_post_injection = {
    "position": _PIXELS["prompt_box_post_injection"],
    "rgb": (230, 255, 255),
}
CHATGPT_BOOT_FOCUS_CLICK_DURATION_SECONDS = 10
IDEA_RESPONSE_ABORT_TIMEOUT_SECONDS = 100
IDEA_RESPONSE_STUCK_PROMPT_RETRY_THRESHOLD = 8
INITIAL_PROMPT_SUBMISSION_WAIT_SECONDS = 10
INITIAL_PROMPT_COMPLETION_DETECTION_DELAY_SECONDS = 5
INITIAL_PROMPT_MIN_WORD_COUNT = 20
PROMPT_TEXT_PASTE_CHUNK_SIZE = 1200
PROMPT_TEXT_PASTE_CHUNK_DELAY_SECONDS = 0.25
IMAGE_GENERATION_POLL_INTERVAL_SECONDS = 2.0
IMAGE_GENERATION_ABORT_TIMEOUT_SECONDS = 240
IMAGE_GENERATION_MIN_WAIT_SECONDS = 20
IMAGE_GENERATION_STUCK_PROMPT_RETRY_THRESHOLD = 8
IMAGE_GENERATION_NOT_ANSWERING_FAILURE_LIMIT = 10
IMAGE_GENERATION_SUBMISSION_WAIT_SECONDS = 10
IMAGE_GENERATION_VERIFICATION_LIMIT = -1
POST_SAVE_EXTRACTION_WAIT_SECONDS = 5.0 
IMAGE_GENERATION_IN_PROGRESS_PHRASES = (
    "creating image",
    "generating image",
    "creating your image",
    "generating your image",
    "making image",
    "making your image",
    "editing image",
    "working on it",
)
IDEA_RESPONSE_IN_PROGRESS_PHRASES = (
    "analyzing image",
    "thinking",
    "working on it",
)


class GenerationStatusOverlay:
    """Small always-on-top window showing live batch generation progress."""

    def __init__(self) -> None:
        self.root = tk.Tk()
        self.root.title("Image Generation Status")
        self.root.attributes("-topmost", True)
        self.root.resizable(False, False)
        self.root.configure(bg="#17352d")

        width, height = 430, 130
        screen_width = self.root.winfo_screenwidth()
        self.root.geometry(f"{width}x{height}+{max(0, (screen_width - width) // 2)}+10")

        self.run_var = tk.StringVar(value="Run: preparing...")
        self.image_var = tk.StringVar(value="Waiting for: preparing images...")
        self.success_var = tk.StringVar(value="Successfully detected: 0")
        self.answering_failures_var = tk.StringVar(
            value=f"Consecutive NOT ANSWERING: 0/{IMAGE_GENERATION_NOT_ANSWERING_FAILURE_LIMIT}"
        )

        label_options = {
            "bg": "#17352d",
            "fg": "#ffffff",
            "anchor": "w",
            "padx": 12,
        }
        tk.Label(
            self.root,
            textvariable=self.run_var,
            font=("Segoe UI Semibold", 9),
            pady=6,
            **label_options,
        ).pack(fill="x")
        tk.Label(
            self.root,
            textvariable=self.image_var,
            font=("Segoe UI", 9),
            wraplength=402,
            justify="left",
            **label_options,
        ).pack(fill="x")
        tk.Label(
            self.root,
            textvariable=self.success_var,
            font=("Segoe UI Semibold", 9),
            fg="#78e6ad",
            **{key: value for key, value in label_options.items() if key != "fg"},
        ).pack(fill="x", pady=(4, 1))
        tk.Label(
            self.root,
            textvariable=self.answering_failures_var,
            font=("Segoe UI Semibold", 9),
            fg="#ffd166",
            **{key: value for key, value in label_options.items() if key != "fg"},
        ).pack(fill="x", pady=(1, 6))
        self.refresh()

    def refresh(self) -> None:
        self.root.update_idletasks()
        self.root.update()

    def set_run(self, current_run: int, total_runs: int) -> None:
        self.run_var.set(f"Run: {current_run} of {total_runs}")
        self.image_var.set("Waiting for: preparing images...")
        self.success_var.set("Successfully detected: 0")
        self.set_answering_failures(0, refresh=False)
        self.refresh()

    def set_waiting_image(self, image_index: int, total_images: int, image_path: Path) -> None:
        self.image_var.set(
            f"Waiting for image: {image_index} of {total_images} — {image_path.name}"
        )
        self.set_answering_failures(0, refresh=False)
        self.refresh()

    def set_success_count(self, successful_count: int, total_images: int) -> None:
        self.success_var.set(
            f"Successfully detected: {successful_count} of {total_images}"
        )
        self.refresh()

    def set_answering_failures(self, count: int, refresh: bool = True) -> None:
        self.answering_failures_var.set(
            "Consecutive NOT ANSWERING: "
            f"{count}/{IMAGE_GENERATION_NOT_ANSWERING_FAILURE_LIMIT}"
        )
        if refresh:
            self.refresh()

    def set_complete(self) -> None:
        self.image_var.set("Waiting for: run complete")
        self.set_answering_failures(0, refresh=False)
        self.refresh()

    def close(self) -> None:
        try:
            self.root.destroy()
        except tk.TclError:
            pass


STATUS_OVERLAY: GenerationStatusOverlay | None = None


class ImageGenerationBatchAbort(RuntimeError):
    """Raised when ChatGPT reports a terminal image-generation limit/error."""


class ImageGenerationReprompt(RuntimeError):
    """Raised when fallback pixel verification says to resubmit the same image."""


class ImageGenerationSkip(RuntimeError):
    """Raised when one image exceeds its overall generation wait limit."""


def append_image_generation_times(durations: list[float | None]) -> None:
    """Append one run's image-generation wait times to the persistent CSV."""
    IMAGE_GENERATION_TIMES_CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
    file_exists = IMAGE_GENERATION_TIMES_CSV_PATH.exists()
    normalized_durations = list(durations[:5])
    normalized_durations.extend([None] * (5 - len(normalized_durations)))

    with IMAGE_GENERATION_TIMES_CSV_PATH.open(
        "a",
        newline="",
        encoding="utf-8",
    ) as csv_file:
        writer = csv.writer(csv_file)
        if not file_exists or IMAGE_GENERATION_TIMES_CSV_PATH.stat().st_size == 0:
            writer.writerow(
                ["run_time", "image_1", "image_2", "image_3", "image_4", "image_5"]
            )
        writer.writerow(
            [datetime.now().isoformat(timespec="seconds")]
            + [f"{duration:.2f}" if duration is not None else "" for duration in normalized_durations]
        )

    print(f"Saved image generation timings to: {IMAGE_GENERATION_TIMES_CSV_PATH}")


def load_image_generation_abort_phrases() -> tuple[str, ...]:
    if not IMAGE_GENERATION_ABORT_PHRASES_PATH.exists():
        raise FileNotFoundError(
            "Image-generation abort phrase file was not found: "
            f"{IMAGE_GENERATION_ABORT_PHRASES_PATH}"
        )

    raw = json.loads(
        IMAGE_GENERATION_ABORT_PHRASES_PATH.read_text(encoding="utf-8")
    )
    if not isinstance(raw, list) or not all(isinstance(item, str) for item in raw):
        raise ValueError(
            f"{IMAGE_GENERATION_ABORT_PHRASES_PATH.name} must contain a JSON array of strings."
        )

    phrases = tuple(phrase.strip().casefold() for phrase in raw if phrase.strip())
    if not phrases:
        raise ValueError(
            f"{IMAGE_GENERATION_ABORT_PHRASES_PATH.name} must contain at least one non-empty phrase."
        )
    return phrases


IMAGE_GENERATION_ABORT_PHRASES = load_image_generation_abort_phrases()


@dataclass
class ProductPromptContext:
    product_kind: str
    product_color: str
    image_folder: Path
    image_paths: list[Path]
    used_phrases_csv: str
    prompt_text: str
    existing_phrases: list[str]


@dataclass
class BackgroundIdea:
    title: str
    visual_concept: str
    background_description: str


class GeneratedImageHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.generated_image_sources: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.casefold() != "img":
            return

        attrs_map = dict(attrs)
        alt_text = (attrs_map.get("alt") or "").strip()
        src = (attrs_map.get("src") or "").strip()
        if alt_text.startswith("Generated image:") and src:
            self.generated_image_sources.append(src)


def ensure_run_helpers_dir() -> None:
    RUN_HELPERS_DIR.mkdir(parents=True, exist_ok=True)


def migrate_legacy_run_helper_files() -> None:
    for legacy_path in LEGACY_RUN_HELPER_PATHS:
        if not legacy_path.exists() or legacy_path.parent == RUN_HELPERS_DIR:
            continue

        target_path = RUN_HELPERS_DIR / legacy_path.name
        if target_path.exists():
            target_path.unlink()
        shutil.move(str(legacy_path), str(target_path))


def reset_run_helper_files() -> None:
    for helper_path in RUN_HELPER_PATHS:
        if helper_path.exists():
            helper_path.unlink()


def initialize_run_helpers() -> None:
    ensure_run_helpers_dir()
    migrate_legacy_run_helper_files()
    reset_run_helper_files()


def load_kind_to_used_phrases() -> dict[str, list[str]]:
    if not USED_IMAGE_DESIGNS_WORKBOOK.exists():
        raise FileNotFoundError(
            f"Used image designs workbook was not found: {USED_IMAGE_DESIGNS_WORKBOOK}"
        )

    workbook = load_workbook(USED_IMAGE_DESIGNS_WORKBOOK, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}

    kind_to_phrases: dict[str, list[str]] = {}
    for row in rows[1:]:
        if not row:
            continue
        kind_value = row[0]
        if kind_value is None:
            continue

        kind_label = str(kind_value).strip()
        if not kind_label:
            continue

        phrases = [
            str(cell).strip()
            for cell in row[1:]
            if cell is not None and str(cell).strip()
        ]
        kind_to_phrases[kind_label.upper()] = phrases

    for configured_kind in PRODUCT_IMAGE_FOLDER_OVERRIDES:
        kind_to_phrases.setdefault(configured_kind, [])

    return kind_to_phrases


def prompt_for_kind(kind_to_phrases: dict[str, list[str]]) -> str:
    available_kinds = sorted(kind_to_phrases.keys())
    if not available_kinds:
        raise ValueError("No product kinds were found in USED-IMAGE-DESIGNS.xlsx.")

    print("Select the product kind to prepare:")
    for index, kind in enumerate(available_kinds, start=1):
        print(f"{index}. {kind}")

    while True:
        choice = input("Enter option number: ").strip()
        if not choice.isdigit():
            print("Please enter a valid number.")
            continue

        selected_index = int(choice)
        if 1 <= selected_index <= len(available_kinds):
            return available_kinds[selected_index - 1]

        print("Please choose one of the listed options.")


def prompt_for_product_color(product_kind: str) -> str:
    while True:
        product_color = input(
            f"What is the product color for {product_kind}? "
        ).strip()
        if product_color:
            return product_color
        print("Please enter the product color.")


def get_images_final_kind_dir(product_kind: str) -> Path:
    kind_folder_name = re.sub(r'[<>:"/\\|?*]+', "-", product_kind.strip()).strip(" .")
    if not kind_folder_name:
        raise ValueError("Product kind cannot be empty when creating IMAGES-FINAL folder.")
    return IMAGES_FINAL_DIR / kind_folder_name


def ensure_images_final_kind_folders(product_kinds: list[str]) -> None:
    IMAGES_FINAL_DIR.mkdir(parents=True, exist_ok=True)
    for product_kind in product_kinds:
        kind_dir = get_images_final_kind_dir(product_kind)
        (kind_dir / "0").mkdir(parents=True, exist_ok=True)


def prompt_for_loop_count() -> int:
    while True:
        choice = input("How many full cycles do you want to run? ").strip()
        if not choice.isdigit():
            print("Please enter a valid whole number.")
            continue

        loop_count = int(choice)
        if loop_count >= 1:
            return loop_count

        print("Please enter at least 1.")


def resolve_product_image_folder(product_kind: str) -> Path:
    if not NO_BG_IMAGES_ROOT.exists():
        raise FileNotFoundError(
            f"NO-BG-IMAGES root folder was not found: {NO_BG_IMAGES_ROOT}"
        )

    configured_folder_name = PRODUCT_IMAGE_FOLDER_OVERRIDES.get(
        product_kind.strip().upper(),
        product_kind,
    )
    normalized_target = configured_folder_name.strip().casefold()
    for folder in NO_BG_IMAGES_ROOT.iterdir():
        if folder.is_dir() and folder.name.strip().casefold() == normalized_target:
            return folder

    raise FileNotFoundError(
        f"Could not find an image folder for kind '{product_kind}' inside {NO_BG_IMAGES_ROOT}."
    )


def get_images_from_folder(folder_path: Path) -> list[Path]:
    image_files = sorted(
        [
            file_path
            for file_path in folder_path.iterdir()
            if file_path.is_file() and file_path.suffix.lower() in SUPPORTED_IMAGE_SUFFIXES
        ]
    )
    if not image_files:
        raise FileNotFoundError(f"No supported images were found in: {folder_path}")
    return image_files


def build_used_phrases_csv(phrases: list[str]) -> str:
    if not phrases:
        return "None used yet"
    return ", ".join(phrases)


def build_prompt_text(product_kind: str, used_phrases_csv: str) -> str:
    if not PROMPT_TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Prompt template was not found: {PROMPT_TEMPLATE_PATH}"
        )

    template = PROMPT_TEMPLATE_PATH.read_text(encoding="utf-8")
    return (
        template.replace("[PRODUCT_KIND]", product_kind)
        .replace("[USED_PHRASES_CSV]", used_phrases_csv)
        .strip()
    )


def build_image_generation_prompt(idea: BackgroundIdea, product_color: str) -> str:
    if not IMAGE_GENERATION_PROMPT_TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            "Image generation prompt template was not found: "
            f"{IMAGE_GENERATION_PROMPT_TEMPLATE_PATH}"
        )

    template = IMAGE_GENERATION_PROMPT_TEMPLATE_PATH.read_text(encoding="utf-8")
    background_description = (
        f"Title: {idea.title}\n"
        f"Visual Concept: {idea.visual_concept}\n"
        f"Background Description: {idea.background_description}"
    ).strip()
    return (
        template.replace("[INSERT BACKGROUND DESCRIPTION HERE]", background_description)
        .replace("[BACKGROUND DESCRIPTION]", background_description)
        .replace("[PRODUCT_COLOR]", product_color.strip())
        .strip()
    )


def prepare_product_prompt_context(
    product_kind: str | None = None,
    product_color: str | None = None,
) -> ProductPromptContext:
    initialize_run_helpers()
    kind_to_phrases = load_kind_to_used_phrases()
    selected_kind = product_kind or prompt_for_kind(kind_to_phrases)
    selected_color = product_color or prompt_for_product_color(selected_kind)
    image_folder = resolve_product_image_folder(selected_kind)
    image_paths = get_images_from_folder(image_folder)
    used_phrases_csv = build_used_phrases_csv(kind_to_phrases.get(selected_kind, []))
    prompt_text = build_prompt_text(selected_kind, used_phrases_csv)

    PROMPT_PREVIEW_PATH.write_text(prompt_text, encoding="utf-8")

    return ProductPromptContext(
        product_kind=selected_kind,
        product_color=selected_color,
        image_folder=image_folder,
        image_paths=image_paths,
        used_phrases_csv=used_phrases_csv,
        prompt_text=prompt_text,
        existing_phrases=kind_to_phrases.get(selected_kind, []),
    )


def get_firefox_binary() -> Path:
    firefox_binary = os.getenv("FIREFOX_BINARY")
    if firefox_binary:
        return Path(firefox_binary)

    for candidate in FALLBACK_FIREFOX_BINARIES:
        if candidate.exists():
            return candidate

    path_candidate = shutil.which("firefox.exe") or shutil.which("firefox")
    if path_candidate:
        return Path(path_candidate)

    return DEFAULT_FIREFOX_BINARY


def open_firefox_normal_window() -> None:
    firefox_binary = get_firefox_binary()
    if not firefox_binary.exists():
        raise FileNotFoundError(
            "Firefox binary was not found. Set FIREFOX_BINARY to your firefox.exe path "
            f"or install Firefox in a standard location. Last checked: {firefox_binary}"
        )

    if not PRABHU_FIREFOX_PROFILE.exists():
        raise FileNotFoundError(
            f"Firefox profile directory was not found: {PRABHU_FIREFOX_PROFILE}"
        )

    subprocess.Popen(
        [
            str(firefox_binary),
            "-profile",
            str(PRABHU_FIREFOX_PROFILE),
            CHATGPT_URL,
        ]
    )


def clear_clipboard() -> None:
    subprocess.run(
        ["powershell.exe", "-NoProfile", "-Command", "Set-Clipboard -Value $null"],
        check=True,
    )


def set_clipboard_text(text: str) -> None:
    if not text:
        clear_clipboard()
        return
    powershell_script = r"Set-Clipboard -Value ([Console]::In.ReadToEnd())"
    subprocess.run(
        ["powershell.exe", "-NoProfile", "-Command", powershell_script],
        input=text,
        text=True,
        check=True,
    )


def get_clipboard_text() -> str:
    result = subprocess.run(
        ["powershell.exe", "-NoProfile", "-Command", "Get-Clipboard"],
        capture_output=True,
        text=True,
        check=True,
    )
    return result.stdout.replace("\r\n", "\n").strip()


def set_clipboard_image(image_path: Path) -> None:
    if not image_path.exists():
        raise FileNotFoundError(f"Image file was not found: {image_path}")

    powershell_script = f"""
Add-Type -AssemblyName System.Windows.Forms
Add-Type -AssemblyName System.Drawing
$image = [System.Drawing.Image]::FromFile('{str(image_path).replace("'", "''")}')
[System.Windows.Forms.Clipboard]::SetImage($image)
$image.Dispose()
"""
    subprocess.run(
        ["powershell.exe", "-sta", "-Command", powershell_script],
        check=True,
    )


def split_text_for_composer_paste(text: str) -> list[str]:
    """Split large text so ChatGPT keeps it in the composer instead of an attachment."""
    if len(text) <= PROMPT_TEXT_PASTE_CHUNK_SIZE:
        return [text]
    return [
        text[start:start + PROMPT_TEXT_PASTE_CHUNK_SIZE]
        for start in range(0, len(text), PROMPT_TEXT_PASTE_CHUNK_SIZE)
    ]


def paste_text_via_clipboard(text: str, field_label: str) -> None:
    chunks = split_text_for_composer_paste(text)
    print(
        f"Pasting prompt text into {field_label} in {len(chunks)} "
        f"composer-safe chunk(s)..."
    )
    for chunk_index, chunk in enumerate(chunks, start=1):
        set_clipboard_text(chunk)
        time.sleep(0.35)
        pyautogui.hotkey("ctrl", "v")
        if chunk_index < len(chunks):
            time.sleep(PROMPT_TEXT_PASTE_CHUNK_DELAY_SECONDS)


def verify_initial_prompt_was_pasted() -> bool:
    print("Verifying that the initial prompt was pasted...")
    pyautogui.hotkey("ctrl", "a")
    time.sleep(0.3)
    pyautogui.hotkey("ctrl", "c")
    time.sleep(0.3)
    copied_text = get_clipboard_text()
    copied_word_count = len(copied_text.split())

    # Ctrl+A leaves the prompt selected. Move the caret to its end so that the
    # upcoming image paste is added to the prompt instead of replacing it.
    pyautogui.press("end")
    time.sleep(0.2)

    print(f"Initial prompt verification copied {copied_word_count} word(s).")
    if copied_word_count < INITIAL_PROMPT_MIN_WORD_COUNT:
        print(
            "Initial prompt paste was not detected: copied text contained fewer "
            f"than {INITIAL_PROMPT_MIN_WORD_COUNT} words."
        )
        return False

    print("Initial prompt paste verified successfully.")
    return True


def paste_image_via_clipboard(image_path: Path, field_label: str) -> None:
    print(f"Loading image into clipboard: {image_path}")
    set_clipboard_image(image_path)
    time.sleep(0.5)
    print(f"Pasting image into {field_label}...")
    pyautogui.hotkey("ctrl", "v")
    time.sleep(0.5)
    clear_clipboard()


def click_chat_copy_target() -> None:
    target_x, target_y = CHAT_CLICK_TARGET
    print(f"Clicking chat copy target at ({target_x}, {target_y}) before copy cycle...")
    pyautogui.moveTo(target_x, target_y, duration=0.2)
    pyautogui.click()


def hold_click_chatgpt_boot_focus_target() -> None:
    target_x, target_y = CHATGPT_PROMPT_BOX_PIXELS_VAIO["position"]
    print(
        "Clicking ChatGPT boot focus target "
        f"({target_x}, {target_y}) every 0.5 seconds for "
        f"{CHATGPT_BOOT_FOCUS_CLICK_DURATION_SECONDS} seconds..."
    )
    pyautogui.moveTo(target_x, target_y, duration=0.2)
    end_time = time.time() + CHATGPT_BOOT_FOCUS_CLICK_DURATION_SECONDS
    while time.time() < end_time:
        pyautogui.click()
        time.sleep(0.5)


def copy_full_chat_text_once() -> str:
    pyautogui.press("esc")
    time.sleep(0.25)
    click_chat_copy_target()
    time.sleep(0.25)
    pyautogui.hotkey("ctrl", "a")
    time.sleep(0.3)
    pyautogui.hotkey("ctrl", "c")
    time.sleep(0.3)
    return get_clipboard_text()


def find_latest_prompt_end(full_chat_text: str, prompt_text: str) -> int | None:
    """Return the end offset of the latest submitted prompt in copied chat text."""
    if not full_chat_text or not prompt_text:
        return None

    exact_prompt_index = full_chat_text.rfind(prompt_text)
    if exact_prompt_index != -1:
        return exact_prompt_index + len(prompt_text)

    # Copied ChatGPT transcripts can collapse whitespace or corrupt a character
    # inside a long prompt. Its final words are stable ASCII in both templates,
    # so use a whitespace-tolerant tail anchor without losing the raw-text
    # offset needed to isolate the current response.
    prompt_words = prompt_text.split()
    prompt_tail_words = prompt_words[-24:]
    prompt_tail = " ".join(prompt_tail_words)
    if len(prompt_tail) < 80:
        return None

    tail_pattern = r"\s+".join(re.escape(word) for word in prompt_tail_words)
    tail_matches = list(
        re.finditer(tail_pattern, full_chat_text, flags=re.IGNORECASE)
    )
    if not tail_matches:
        return None
    return tail_matches[-1].end()


def get_response_after_latest_prompt(
    full_chat_text: str,
    prompt_text: str,
) -> str | None:
    prompt_end = find_latest_prompt_end(full_chat_text, prompt_text)
    if prompt_end is None:
        return None
    return full_chat_text[prompt_end:].strip()


def wait_for_stable_full_chat_text(
    prompt_text: str = "",
    timeout_started_at: float | None = None,
) -> str | None:
    print(
        "Starting full-chat copy cycle every 0.5 seconds until two consecutive copies match and idea output is present..."
    )
    previous_copy: str | None = None
    attempt = 0
    stuck_counter = 0

    while True:
        attempt += 1
        current_copy = copy_full_chat_text_once()
        print(f"Captured full chat copy attempt {attempt}.")
        is_stable_copy = bool(current_copy and previous_copy == current_copy)
        current_response = (
            get_response_after_latest_prompt(current_copy, prompt_text)
            if prompt_text
            else current_copy
        )
        has_ideas = bool(
            current_response is not None and parse_ideas(current_response)
        )
        lower_copy = current_response.casefold() if current_response else ""
        still_in_progress = any(
            phrase in lower_copy for phrase in IDEA_RESPONSE_IN_PROGRESS_PHRASES
        )

        if is_stable_copy and has_ideas and not still_in_progress:
            print(
                "Detected stable copied chat text with parseable idea output. Treating output as complete."
            )
            return current_copy

        prompt_tail = prompt_text.split()[-10:]
        copy_tail = current_copy.split()[-10:]
        is_stuck = bool(prompt_tail and copy_tail == prompt_tail)
        print(
            "Stuck check: "
            f"{is_stuck} (counter: {stuck_counter + 1 if is_stuck else 0}/"
            f"{IDEA_RESPONSE_STUCK_PROMPT_RETRY_THRESHOLD})"
        )
        if is_stuck:
            stuck_counter += 1
            if stuck_counter >= IDEA_RESPONSE_STUCK_PROMPT_RETRY_THRESHOLD:
                print("Detected prompt still in input box (ENTER might have failed). Repressing ENTER...")
                pyautogui.click(CHATGPT_PROMPT_BOX_PIXELS_VAIO_post_injection["position"])
                time.sleep(0.5)
                pyautogui.press("enter")
                time.sleep(3)
                stuck_counter = 0
        else:
            stuck_counter = 0

        if is_stable_copy and not has_ideas:
            print(
                "Copied chat text is stable, but no parseable JSON or fallback idea blocks were found yet. Continuing to wait..."
            )

        if (
            timeout_started_at is not None
            and time.time() - timeout_started_at >= IDEA_RESPONSE_ABORT_TIMEOUT_SECONDS
        ):
            print(
                "Timed out after "
                f"{IDEA_RESPONSE_ABORT_TIMEOUT_SECONDS} seconds without detecting "
                "parseable idea JSON. Aborting this run."
            )
            return None

        previous_copy = current_copy
        time.sleep(0.5)


def load_existing_text(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8").replace("\r\n", "\n").strip()


def append_text_block(path: Path, text: str) -> None:
    existing_text = load_existing_text(path)
    combined_text = f"{existing_text}\n\n{text}".strip() if existing_text else text.strip()
    path.write_text(combined_text + "\n", encoding="utf-8")


def strip_duplicated_prompt_prefix(text: str, prompt_text: str) -> str:
    normalized_text = text.strip()
    normalized_prompt = prompt_text.strip()
    duplicated_prompt = normalized_prompt + normalized_prompt

    if normalized_prompt and normalized_text.startswith(duplicated_prompt):
        return normalized_text[len(normalized_prompt):].strip()

    return normalized_text


def normalize_phrase(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip().casefold()


def is_valid_ideas_payload(payload: dict[str, object]) -> bool:
    ideas = payload.get("ideas")
    if not isinstance(ideas, list) or not ideas:
        return False

    first_item = ideas[0]
    if not isinstance(first_item, dict):
        return False

    title = str(first_item.get("title", "")).strip()
    visual_concept = str(first_item.get("visual_concept", "")).strip()
    background_description = str(first_item.get("background_description", "")).strip()
    if not title or title == "...":
        return False
    if visual_concept == "..." or background_description == "...":
        return False
    return True


def extract_json_payload(text: str) -> dict[str, object] | None:
    decoder = json.JSONDecoder()
    candidate_start_indexes = [match.start() for match in re.finditer(r'\{\s*"ideas"\s*:', text)]
    valid_payloads: list[dict[str, object]] = []

    for start_index in candidate_start_indexes:
        try:
            payload, _ = decoder.raw_decode(text[start_index:].strip())
        except json.JSONDecodeError:
            continue

        if isinstance(payload, dict) and is_valid_ideas_payload(payload):
            valid_payloads.append(payload)

    if valid_payloads:
        return valid_payloads[-1]

    return None


def background_idea_from_json_item(item: dict[str, object]) -> BackgroundIdea | None:
    title = str(item.get("title", "")).strip()
    visual_concept = str(item.get("visual_concept", "")).strip()
    background_description = str(item.get("background_description", "")).strip()
    if not title:
        return None
    return BackgroundIdea(
        title=title,
        visual_concept=visual_concept,
        background_description=background_description,
    )


def background_idea_to_block(idea: BackgroundIdea) -> str:
    lines = [f"{IDEA_MARKER} {idea.title}"]
    if idea.visual_concept:
        lines.append(f"Visual Concept: {idea.visual_concept}")
    if idea.background_description:
        lines.append(f"Background Description: {idea.background_description}")
    return "\n".join(lines).strip()


def parse_ideas_from_json(text: str) -> list[BackgroundIdea]:
    payload = extract_json_payload(text)
    if not payload:
        return []

    ideas = payload.get("ideas")
    if not isinstance(ideas, list):
        return []

    parsed_ideas: list[BackgroundIdea] = []
    for item in ideas:
        if not isinstance(item, dict):
            continue
        parsed_idea = background_idea_from_json_item(item)
        if parsed_idea:
            parsed_ideas.append(parsed_idea)

    return parsed_ideas


def parse_ideas_from_marker_blocks(text: str) -> list[BackgroundIdea]:
    marker_matches = list(re.finditer(re.escape(IDEA_MARKER), text))
    if not marker_matches:
        return []

    ideas: list[BackgroundIdea] = []
    for index, match in enumerate(marker_matches):
        start = match.start()
        end = (
            marker_matches[index + 1].start()
            if index + 1 < len(marker_matches)
            else len(text)
        )
        block = text[start:end].strip()
        if not block:
            continue

        lines = [line.strip() for line in block.splitlines() if line.strip()]
        if not lines:
            continue

        first_line = lines[0]
        if not first_line.startswith(IDEA_MARKER):
            continue

        title = first_line[len(IDEA_MARKER):].strip()
        visual_concept = ""
        background_description = ""

        for line in lines[1:]:
            if line.startswith("Visual Concept:"):
                visual_concept = line.split(":", 1)[1].strip()
            elif line.startswith("Background Description:"):
                background_description = line.split(":", 1)[1].strip()

        if title and visual_concept and background_description:
            ideas.append(
                BackgroundIdea(
                    title=title,
                    visual_concept=visual_concept,
                    background_description=background_description,
                )
            )

    return ideas


def parse_ideas(text: str) -> list[BackgroundIdea]:
    json_ideas = parse_ideas_from_json(text)
    if json_ideas:
        return json_ideas
    return parse_ideas_from_marker_blocks(text)


def extract_idea_title(idea: BackgroundIdea) -> str:
    return idea.title.strip()


def format_idea_blocks(ideas: list[BackgroundIdea]) -> str:
    if not ideas:
        return ""
    return "\n\n".join(background_idea_to_block(idea) for idea in ideas).strip()


def get_new_ideas(ideas: list[BackgroundIdea], existing_phrases: list[str]) -> list[BackgroundIdea]:
    existing_phrase_keys = {normalize_phrase(phrase) for phrase in existing_phrases if phrase.strip()}
    new_ideas: list[BackgroundIdea] = []

    for idea in ideas:
        idea_title = extract_idea_title(idea)
        if normalize_phrase(idea_title) not in existing_phrase_keys:
            new_ideas.append(idea)

    return new_ideas


def choose_current_idea(ideas: list[BackgroundIdea], existing_phrases: list[str]) -> BackgroundIdea:
    new_ideas = get_new_ideas(ideas, existing_phrases)
    if new_ideas:
        return new_ideas[0]
    if ideas:
        return ideas[0]
    raise ValueError("No background ideas were parsed from the latest output.")


def append_phrase_to_workbook(product_kind: str, phrase_title: str) -> None:
    workbook = load_workbook(USED_IMAGE_DESIGNS_WORKBOOK)
    worksheet = workbook[workbook.sheetnames[0]]
    target_kind = normalize_phrase(product_kind)

    for row_index in range(2, worksheet.max_row + 1):
        kind_value = worksheet.cell(row=row_index, column=1).value
        if kind_value is None:
            continue
        if normalize_phrase(str(kind_value)) != target_kind:
            continue

        existing_row_phrases = []
        for column_index in range(2, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=row_index, column=column_index).value
            if cell_value is not None and str(cell_value).strip():
                existing_row_phrases.append(str(cell_value).strip())

        if normalize_phrase(phrase_title) in {
            normalize_phrase(phrase) for phrase in existing_row_phrases
        }:
            workbook.save(USED_IMAGE_DESIGNS_WORKBOOK)
            return

        target_column = 2
        while worksheet.cell(row=row_index, column=target_column).value not in (None, ""):
            target_column += 1

        worksheet.cell(row=row_index, column=target_column).value = phrase_title
        workbook.save(USED_IMAGE_DESIGNS_WORKBOOK)
        return

    new_row_index = worksheet.max_row + 1
    worksheet.cell(row=new_row_index, column=1).value = product_kind
    worksheet.cell(row=new_row_index, column=2).value = phrase_title
    workbook.save(USED_IMAGE_DESIGNS_WORKBOOK)


def save_current_run_idea(idea: BackgroundIdea) -> None:
    payload = {
        "title": idea.title,
        "visual_concept": idea.visual_concept,
        "background_description": idea.background_description,
    }
    CURRENT_RUN_IDEA_PATH.write_text(
        json.dumps(payload, indent=2, ensure_ascii=True) + "\n",
        encoding="utf-8",
    )


def extract_latest_output(full_chat_text: str, prompt_text: str) -> str:
    normalized_full_chat = full_chat_text.replace("\r\n", "\n").strip()
    current_response = get_response_after_latest_prompt(
        normalized_full_chat,
        prompt_text,
    )
    if current_response is not None:
        return strip_duplicated_prompt_prefix(current_response, prompt_text)

    previous_full_chat = load_existing_text(LAST_FULL_CHAT_PATH)
    previous_outputs = load_existing_text(ALL_RESPONSES_PATH)

    latest_output = normalized_full_chat

    if previous_full_chat and normalized_full_chat.startswith(previous_full_chat):
        latest_output = normalized_full_chat[len(previous_full_chat):].strip()
    else:
        if prompt_text and prompt_text in latest_output:
            latest_output = latest_output.split(prompt_text, 1)[-1].strip()
        if previous_outputs and previous_outputs in latest_output:
            latest_output = latest_output.replace(previous_outputs, "", 1).strip()

    return strip_duplicated_prompt_prefix(latest_output, prompt_text)


def capture_and_store_latest_output(
    prompt_text: str,
    timeout_started_at: float | None = None,
) -> str | None:
    click_chat_copy_target()
    # stable_full_chat_text = wait_for_stable_full_chat_text()
    stable_full_chat_text = wait_for_stable_full_chat_text(
        prompt_text,
        timeout_started_at,
    )
    if stable_full_chat_text is None:
        return None

    latest_output = extract_latest_output(stable_full_chat_text, prompt_text)

    LAST_FULL_CHAT_PATH.write_text(stable_full_chat_text + "\n", encoding="utf-8")
    LATEST_RESPONSE_PATH.write_text(latest_output + "\n", encoding="utf-8")
    if latest_output:
        append_text_block(ALL_RESPONSES_PATH, latest_output)

    print(f"Saved latest response to: {LATEST_RESPONSE_PATH}")
    print(f"Saved full chat snapshot to: {LAST_FULL_CHAT_PATH}")
    return latest_output


def save_parsed_idea_results(latest_output: str, existing_phrases: list[str]) -> list[BackgroundIdea]:
    ideas = parse_ideas(latest_output)
    new_ideas = get_new_ideas(ideas, existing_phrases)

    PARSED_IDEAS_PATH.write_text(
        format_idea_blocks(ideas) + ("\n" if ideas else ""),
        encoding="utf-8",
    )
    NEW_IDEAS_PATH.write_text(
        format_idea_blocks(new_ideas) + ("\n" if new_ideas else ""),
        encoding="utf-8",
    )

    print(f"Saved parsed idea blocks to: {PARSED_IDEAS_PATH}")
    print(f"Saved ideas not found in Excel to: {NEW_IDEAS_PATH}")
    return ideas


def prepare_current_generation_prompt(
    idea: BackgroundIdea,
    product_color: str,
) -> str:
    prompt_text = build_image_generation_prompt(idea, product_color)
    CURRENT_GENERATION_PROMPT_PATH.write_text(prompt_text + "\n", encoding="utf-8")
    print(f"Saved current generation prompt to: {CURRENT_GENERATION_PROMPT_PATH}")
    return prompt_text


def beep_ready_for_generation_prompt() -> None:
    winsound.MessageBeep(winsound.MB_ICONASTERISK)
    time.sleep(0.2)
    winsound.MessageBeep(winsound.MB_ICONASTERISK)


def beep_image_generation_complete() -> None:
    winsound.MessageBeep(winsound.MB_OK)
    time.sleep(0.25)
    winsound.MessageBeep(winsound.MB_OK)
    time.sleep(0.25)
    winsound.MessageBeep(winsound.MB_OK)


def beep_all_images_generation_complete() -> None:
    for frequency, duration_ms in (
        (880, 250),
        (988, 250),
        (1175, 450),
    ):
        winsound.Beep(frequency, duration_ms)
        time.sleep(0.12)


def save_generated_images_to_output_folder() -> None:
    output_folder = FULL_GENERATED_IMAGES_DIR.resolve()
    if not output_folder.exists():
        raise FileNotFoundError(
            f"Full generated images folder was not found: {output_folder}"
        )

    print(f"Saving generated images into: {output_folder}")
    pyautogui.hotkey("ctrl", "s")
    time.sleep(2.0)
    pyautogui.hotkey("alt", "d")
    time.sleep(0.35)
    set_clipboard_text(str(output_folder))
    time.sleep(0.35)
    pyautogui.hotkey("ctrl", "v")
    time.sleep(0.35)
    pyautogui.press("enter")
    time.sleep(0.5)

    for _ in range(6):
        pyautogui.press("tab")
        time.sleep(0.12)

    set_clipboard_text(time.strftime("%H-%M-%S"))
    time.sleep(0.2)
    pyautogui.hotkey("ctrl", "v")
    time.sleep(0.2)
    pyautogui.press("enter")


def get_latest_saved_html_path() -> Path:
    html_candidates = sorted(
        [
            path
            for path in FULL_GENERATED_IMAGES_DIR.iterdir()
            if path.is_file() and path.suffix.lower() in {".htm", ".html"}
        ],
        key=lambda path: path.stat().st_mtime,
    )
    if not html_candidates:
        raise FileNotFoundError(
            f"No saved HTML files were found in: {FULL_GENERATED_IMAGES_DIR}"
        )
    return html_candidates[-1]


def get_next_images_final_output_dir(product_kind: str) -> Path:
    images_final_kind_dir = get_images_final_kind_dir(product_kind)
    images_final_kind_dir.mkdir(parents=True, exist_ok=True)
    numeric_folders = [
        int(path.name)
        for path in images_final_kind_dir.iterdir()
        if path.is_dir() and path.name.isdigit() and path.name != "999"
    ]
    next_folder_number = max(numeric_folders, default=-1) + 1

    # Folder 999 is reserved and must neither affect numbering nor be replaced.
    while next_folder_number == 999 or (
        images_final_kind_dir / str(next_folder_number)
    ).exists():
        next_folder_number += 1

    output_dir = images_final_kind_dir / str(next_folder_number)
    output_dir.mkdir(parents=True, exist_ok=False)
    return output_dir


def extract_generated_image_sources_from_html(html_path: Path) -> list[str]:
    html_text = html_path.read_text(encoding="utf-8", errors="replace")
    parser = GeneratedImageHTMLParser()
    parser.feed(html_text)
    return parser.generated_image_sources


def copy_or_download_generated_image(
    image_src: str,
    html_path: Path,
    destination_path: Path,
) -> None:
    decoded_src = unquote(image_src)
    local_source_path = (html_path.parent / decoded_src).resolve()
    if local_source_path.exists():
        shutil.copy2(local_source_path, destination_path)
        return

    with urllib.request.urlopen(image_src) as response:
        destination_path.write_bytes(response.read())


def extract_generated_images_from_latest_saved_html(product_kind: str) -> Path | None:
    latest_html_path = get_latest_saved_html_path()
    generated_sources = extract_generated_image_sources_from_html(latest_html_path)
    if not generated_sources:
        print(
            "No generated-image <img> tags were found in the saved HTML, "
            f"so skipping final image extraction for this cycle: {latest_html_path}"
        )
        return None

    output_dir = get_next_images_final_output_dir(product_kind)
    print(f"Extracting generated images from: {latest_html_path}")
    print(f"Saving ordered generated images to: {output_dir}")

    for image_index, image_src in enumerate(generated_sources, start=1):
        parsed_src = urlparse(image_src)
        image_suffix = Path(parsed_src.path).suffix or ".png"
        destination_path = output_dir / f"{image_index}{image_suffix}"
        copy_or_download_generated_image(
            image_src=image_src,
            html_path=latest_html_path,
            destination_path=destination_path,
        )

    return output_dir


def is_image_generation_in_progress(full_chat_text: str) -> bool:
    normalized_text = full_chat_text.casefold()
    return any(
        phrase in normalized_text for phrase in IMAGE_GENERATION_IN_PROGRESS_PHRASES
    )


def find_image_generation_abort_phrase(full_chat_text: str) -> str | None:
    copied_text_tail = " ".join(full_chat_text.split()[-100:]).casefold()
    return next(
        (phrase for phrase in IMAGE_GENERATION_ABORT_PHRASES if phrase in copied_text_tail),
        None,
    )


def has_generated_image_confirmation(
    full_chat_text: str,
    generation_prompt_text: str,
) -> bool:
    current_response = get_response_after_latest_prompt(
        full_chat_text,
        generation_prompt_text,
    )
    return bool(
        current_response
        and "generated image:" in current_response.casefold()
    )


def chatgpt_answering_pixel_matches() -> tuple[bool, tuple[int, int, int]]:
    if CHATGPT_ANSWERING_PIXEL is None:
        raise RuntimeError("ChatGPT answering-state pixel is not configured.")

    position, expected_rgb, tolerance = CHATGPT_ANSWERING_PIXEL
    actual_rgb = tuple(int(channel) for channel in pyautogui.pixel(*position))
    matches = all(
        abs(actual - expected) <= tolerance
        for actual, expected in zip(actual_rgb, expected_rgb)
    )
    print(
        f"ChatGPT answering pixel check at {position}: actual={actual_rgb}, "
        f"expected={expected_rgb}, tolerance={tolerance}, matched={matches}."
    )
    return matches, actual_rgb


def wait_for_image_generation_completion(
    generation_prompt_text: str,
    image_wait_started_at: float,
) -> str | None:
    print(
        "Waiting for image-generation mode to end by polling copied chat text for generating-state phrases..."
    )
    time.sleep(IMAGE_GENERATION_MIN_WAIT_SECONDS)

    previous_copy: str | None = None
    attempt = 0
    stuck_counter = 0
    consecutive_not_answering_checks = 0

    if CHATGPT_ANSWERING_PIXEL is not None:
        print(
            "Answering-state grace period complete; beginning pixel checks. "
            f"A reprompt requires {IMAGE_GENERATION_NOT_ANSWERING_FAILURE_LIMIT} "
            "consecutive NOT ANSWERING readings."
        )

    while True:
        attempt += 1
        current_copy = copy_full_chat_text_once()
        print(f"Checked image-generation status attempt {attempt}.")
        abort_phrase = find_image_generation_abort_phrase(current_copy)
        if abort_phrase:
            IMAGE_GENERATION_FINAL_CHAT_PATH.write_text(
                current_copy + "\n",
                encoding="utf-8",
            )
            raise ImageGenerationBatchAbort(
                "Aborting the batch because the copied ChatGPT response matched "
                f"the configured terminal phrase: {abort_phrase!r}"
            )
        is_stable_copy = bool(current_copy and previous_copy == current_copy)
        current_response = get_response_after_latest_prompt(
            current_copy,
            generation_prompt_text,
        )
        has_generated_confirmation = has_generated_image_confirmation(
            current_copy,
            generation_prompt_text,
        )

        if (
            is_stable_copy
            and current_response is not None
            and not is_image_generation_in_progress(current_response)
            and has_generated_confirmation
        ):
            print(
                "Detected stable copied chat text with a generated-image confirmation after the injected prompt."
            )
            IMAGE_GENERATION_FINAL_CHAT_PATH.write_text(
                current_copy + "\n",
                encoding="utf-8",
            )
            beep_image_generation_complete()
            print(
                f"Saved final image-generation chat snapshot to: {IMAGE_GENERATION_FINAL_CHAT_PATH}"
            )
            return current_copy

        gen_prompt_tail = generation_prompt_text.split()[-10:]
        copy_tail = current_copy.split()[-10:]
        is_stuck = bool(gen_prompt_tail and copy_tail == gen_prompt_tail)
        print(
            "Stuck check: "
            f"{is_stuck} (counter: {stuck_counter + 1 if is_stuck else 0}/"
            f"{IMAGE_GENERATION_STUCK_PROMPT_RETRY_THRESHOLD})"
        )
        if is_stuck:
            stuck_counter += 1
            if stuck_counter >= IMAGE_GENERATION_STUCK_PROMPT_RETRY_THRESHOLD:
                print("Detected generation prompt still in input box. Repressing ENTER...")
                pyautogui.click(CHATGPT_PROMPT_BOX_PIXELS_VAIO_post_injection["position"])
                time.sleep(0.5)
                pyautogui.press("enter")
                time.sleep(3)
                stuck_counter = 0
        else:
            stuck_counter = 0

        if not has_generated_confirmation:
            print(
                "Copied chat text did not confirm 'Generated image:'. Continuing to wait..."
            )

        if not has_generated_confirmation and CHATGPT_ANSWERING_PIXEL is not None:
            is_answering, _actual_rgb = chatgpt_answering_pixel_matches()
            if is_answering:
                consecutive_not_answering_checks = 0
                if STATUS_OVERLAY is not None:
                    STATUS_OVERLAY.set_answering_failures(0)
                print("ChatGPT state: ANSWERING. Continuing to wait.")
            else:
                consecutive_not_answering_checks += 1
                if STATUS_OVERLAY is not None:
                    STATUS_OVERLAY.set_answering_failures(
                        consecutive_not_answering_checks
                    )
                print(
                    "ChatGPT state: NOT ANSWERING "
                    f"({consecutive_not_answering_checks}/"
                    f"{IMAGE_GENERATION_NOT_ANSWERING_FAILURE_LIMIT} consecutive checks)."
                )
                if (
                    consecutive_not_answering_checks
                    >= IMAGE_GENERATION_NOT_ANSWERING_FAILURE_LIMIT
                ):
                    raise ImageGenerationReprompt(
                        "ChatGPT remained NOT ANSWERING for the configured consecutive "
                        "check limit and no generated-image confirmation was found; "
                        "resubmitting the same prompt and image."
                    )

        if time.time() - image_wait_started_at >= IMAGE_GENERATION_ABORT_TIMEOUT_SECONDS:
            IMAGE_GENERATION_FINAL_CHAT_PATH.write_text(
                current_copy + "\n",
                encoding="utf-8",
            )
            raise ImageGenerationSkip(
                "Skipping this image after waiting "
                f"{IMAGE_GENERATION_ABORT_TIMEOUT_SECONDS} seconds without a "
                "verified generated-image response."
            )

        previous_copy = current_copy
        time.sleep(IMAGE_GENERATION_POLL_INTERVAL_SECONDS)


def run_generation_prompt_for_image(
    image_path: Path,
    generation_prompt_text: str,
) -> float:
    submission_attempt = 0
    image_wait_started_at: float | None = None
    while True:
        submission_attempt += 1
        print(
            "Starting follow-up image generation prompt and image paste flow "
            f"(submission attempt {submission_attempt})..."
        )
        pyautogui.press("w")
        time.sleep(0.8)
        paste_text_via_clipboard(generation_prompt_text, "focused ChatGPT prompt box")
        time.sleep(1.5)
        paste_image_via_clipboard(image_path, "focused ChatGPT prompt box")
        print(
            "Waiting "
            f"{IMAGE_GENERATION_SUBMISSION_WAIT_SECONDS} seconds before submitting "
            "the image generation prompt..."
        )
        time.sleep(IMAGE_GENERATION_SUBMISSION_WAIT_SECONDS)
        pyautogui.press("enter")
        if image_wait_started_at is None:
            image_wait_started_at = time.time()
        print("Pressed Enter to submit the image generation prompt.")
        try:
            wait_for_image_generation_completion(
                generation_prompt_text,
                image_wait_started_at,
            )
            return time.time() - image_wait_started_at
        except ImageGenerationReprompt as exc:
            print(exc)
            print(f"Reprompting the same image: {image_path}")


def run_generation_prompt_for_remaining_images(
    image_paths: list[Path],
    generation_prompt_text: str,
    product_kind: str,
) -> bool:
    if IMAGE_GENERATION_VERIFICATION_LIMIT == -1:
        target_verification_count = len(image_paths)
    elif IMAGE_GENERATION_VERIFICATION_LIMIT < -1:
        raise ValueError(
            "IMAGE_GENERATION_VERIFICATION_LIMIT must be -1 or a non-negative integer."
        )
    else:
        target_verification_count = min(
            IMAGE_GENERATION_VERIFICATION_LIMIT,
            len(image_paths),
        )

    if target_verification_count == 0:
        print(
            "IMAGE_GENERATION_VERIFICATION_LIMIT is 0, so skipping image verification and moving directly to the final save flow."
        )
        beep_all_images_generation_complete()
        save_generated_images_to_output_folder()
        time.sleep(POST_SAVE_EXTRACTION_WAIT_SECONDS)
        output_dir = extract_generated_images_from_latest_saved_html(product_kind)
        return output_dir is not None

    print(
        f"Will verify {target_verification_count} generated image(s) before the final beep/save flow."
    )

    successful_count = 0
    skipped_count = 0
    image_durations: list[float | None] = [None] * target_verification_count

    for image_index, image_path in enumerate(
        image_paths[:target_verification_count],
        start=1,
    ):
        print()
        print(
            f"Running image generation for image {image_index} of {target_verification_count}: {image_path}"
        )
        if STATUS_OVERLAY is not None:
            STATUS_OVERLAY.set_waiting_image(
                image_index,
                target_verification_count,
                image_path,
            )
        try:
            image_durations[image_index - 1] = run_generation_prompt_for_image(
                image_path,
                generation_prompt_text,
            )
        except ImageGenerationSkip as exc:
            skipped_count += 1
            print(exc)
            print(
                f"Skipped image {image_index} of {target_verification_count}: {image_path}"
            )
            continue

        successful_count += 1
        if STATUS_OVERLAY is not None:
            STATUS_OVERLAY.set_success_count(successful_count, target_verification_count)
        print(
            f"Confirmed generated image for image {image_index} of {target_verification_count}."
        )

    append_image_generation_times(image_durations)

    print(
        f"Image generation finished with {successful_count} confirmed and "
        f"{skipped_count} skipped image(s)."
    )
    if skipped_count == 0 and target_verification_count == len(image_paths):
        print("Confirmed generated images for every image in the folder.")
    else:
        print(
            "Reached the configured image-generation verification limit. Moving to the final beep/save flow."
        )
    beep_all_images_generation_complete()
    save_generated_images_to_output_folder()
    time.sleep(POST_SAVE_EXTRACTION_WAIT_SECONDS)
    output_dir = extract_generated_images_from_latest_saved_html(product_kind)
    return successful_count > 0 and output_dir is not None


def run_chatgpt_manual_browser_flow(context: ProductPromptContext) -> bool:
    print()
    print("Firefox will open as a normal browser window.")
    print("Then manually go to ChatGPT, open the page you want, and keep it visible.")
    print("Keep the ChatGPT prompt box focused before pressing the Right Arrow key.")
    print()

    open_firefox_normal_window()
    time.sleep(2)

    # wait_for_start_hotkey()
    hold_click_chatgpt_boot_focus_target()

    print()
    print("Starting focused-field prompt and image paste flow...")
    paste_text_via_clipboard(context.prompt_text, "focused ChatGPT prompt box")
    time.sleep(0.9)
    if not verify_initial_prompt_was_pasted():
        return False
    paste_image_via_clipboard(context.image_paths[0], "focused ChatGPT prompt box")
    print(
        "Waiting "
        f"{INITIAL_PROMPT_SUBMISSION_WAIT_SECONDS} seconds before submitting "
        "the ChatGPT prompt..."
    )
    time.sleep(INITIAL_PROMPT_SUBMISSION_WAIT_SECONDS)
    pyautogui.press("enter")
    prompt_submitted_at = time.time()
    print("Pressed Enter to submit the prompt")
    print(
        "Waiting "
        f"{INITIAL_PROMPT_COMPLETION_DETECTION_DELAY_SECONDS} seconds before "
        "starting output-completion detection..."
    )
    time.sleep(INITIAL_PROMPT_COMPLETION_DETECTION_DELAY_SECONDS)
    latest_output = capture_and_store_latest_output(
        context.prompt_text,
        prompt_submitted_at,
    )
    if not latest_output:
        print("No new latest output text could be isolated from the copied conversation.")
        return False

    ideas = save_parsed_idea_results(latest_output, context.existing_phrases)
    print("Captured latest output successfully.")
    if not ideas:
        print("No parsed ideas were found in the latest output.")
        return False

    new_titles = [idea.title for idea in get_new_ideas(ideas, context.existing_phrases)]
    if new_titles:
        print("Ideas not found in Excel:")
        for title in new_titles:
            print(f"- {title}")

    current_idea = choose_current_idea(ideas, context.existing_phrases)
    generation_prompt_text = prepare_current_generation_prompt(
        current_idea,
        context.product_color,
    )
    # beep_ready_for_generation_prompt()

    print(f"Selected CURRENT IDEA: {current_idea.title}")
    # print(
    #     "Image generation prompt is ready. Press the Right Arrow key to paste the generation prompt and image."
    # )
    # wait_for_start_hotkey()
    save_current_run_idea(current_idea)
    print(f"Saved current run idea JSON to: {CURRENT_RUN_IDEA_PATH}")
    generation_succeeded = run_generation_prompt_for_remaining_images(
        context.image_paths,
        generation_prompt_text,
        context.product_kind,
    )
    if not generation_succeeded:
        print(
            "Image generation did not produce a verified, extracted result. "
            "Leaving the idea unmarked so this cycle can retry."
        )
        return False

    append_phrase_to_workbook(context.product_kind, current_idea.title)
    return True


def wait_for_start_hotkey() -> None:
    print(
        "After ChatGPT is open and visible in Firefox, press the Right Arrow key from the browser to start."
    )

    start_detected = {"pressed": False}

    def on_press(key: keyboard.Key | keyboard.KeyCode) -> bool | None:
        try:
            if key == START_HOTKEY_KEY:
                start_detected["pressed"] = True
                print("Detected start hotkey: Right Arrow")
                return False
        except Exception:
            return None
        return None

    with keyboard.Listener(on_press=on_press) as listener:
        listener.join()

    if not start_detected["pressed"]:
        raise RuntimeError("Start hotkey was not detected.")


def main() -> None:
    global STATUS_OVERLAY

    loop_count = prompt_for_loop_count()
    kind_to_phrases = load_kind_to_used_phrases()
    ensure_images_final_kind_folders(sorted(kind_to_phrases.keys()))
    selected_kind = prompt_for_kind(kind_to_phrases)
    selected_color = prompt_for_product_color(selected_kind)
    STATUS_OVERLAY = GenerationStatusOverlay()

    for cycle_index in range(1, loop_count + 1):
        STATUS_OVERLAY.set_run(cycle_index, loop_count)
        while True:
            print()
            print(f"========== Starting cycle {cycle_index} of {loop_count} ==========")

            context = prepare_product_prompt_context(selected_kind, selected_color)

            print(f"Selected kind: {context.product_kind}")
            print(f"Product color: {context.product_color}")
            print(f"First image ready: {context.image_paths[0]}")
            print(f"Total images queued for generation: {len(context.image_paths)}")
            print(f"Prompt preview saved to: {PROMPT_PREVIEW_PATH}")

            if run_chatgpt_manual_browser_flow(context):
                break

            print("Restarting this cycle as a new run after failed prompt-paste verification.")

        print(f"========== Finished cycle {cycle_index} of {loop_count} ==========")
        STATUS_OVERLAY.set_complete()


if __name__ == "__main__":
    try:
        main()
    except ImageGenerationBatchAbort as exc:
        print()
        print("Image-generation batch stopped cleanly.")
        print(exc)



