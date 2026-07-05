from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import time
import urllib.request
import winsound
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import unquote, urlparse

import pyautogui
from openpyxl import load_workbook
from pynput import keyboard

# LAPTOP_NAME = "VAIO"
LAPTOP_NAME = "ASUS"

PRABHU_FIREFOX_PROFILE_ASUS = Path(
    r"C:\Users\ESHAAN\Documents\Firefox-Profiles\0xe7h0bx.prabhu"
)
PRABHU_FIREFOX_PROFILE_VAIO = Path(
    r"C:\Users\SONY\AppData\Roaming\Mozilla\Firefox\Profiles\gm1pmawk.default-release"
)
IMAGE_PROMPTER_ROOT = Path(__file__).resolve().parent
ROOT_PATH = IMAGE_PROMPTER_ROOT


def image_prompter_path(*relative_parts: str) -> Path:
    return ROOT_PATH.joinpath(*relative_parts)


RUN_HELPERS_DIR = image_prompter_path("run-helpers")
FULL_GENERATED_IMAGES_DIR = image_prompter_path("FULL GENERATED IMAGES")
IMAGES_FINAL_DIR = image_prompter_path("IMAGES-FINAL")
NO_BG_IMAGES_ROOT_ASUS = Path(r"C:\work-mom\NO-BG-IMAGES")
PIXELS_FILE_ASUS = image_prompter_path("pixels-ASUS.json")
PIXELS_FILE_VAIO = image_prompter_path("pixels-VAIO.json")
NO_BG_IMAGES_ROOT_VAIO = Path(r"C:\NO-BG-IMAGES")
LAPTOP_CONFIGS = {
    "ASUS": {
        "firefox_profile": PRABHU_FIREFOX_PROFILE_ASUS,
        "no_bg_images_root": NO_BG_IMAGES_ROOT_ASUS,
        "pixel_json" : PIXELS_FILE_ASUS
    },
    "VAIO": {
        "firefox_profile": PRABHU_FIREFOX_PROFILE_VAIO,
        "no_bg_images_root": NO_BG_IMAGES_ROOT_VAIO,
        "pixel_json" : PIXELS_FILE_VAIO
    },
}
ACTIVE_LAPTOP_CONFIG = LAPTOP_CONFIGS[LAPTOP_NAME.upper()]
PRABHU_FIREFOX_PROFILE = ACTIVE_LAPTOP_CONFIG["firefox_profile"]
NO_BG_IMAGES_ROOT = ACTIVE_LAPTOP_CONFIG["no_bg_images_root"]
PIXELS_FILE = ACTIVE_LAPTOP_CONFIG["pixel_json"]
USED_IMAGE_DESIGNS_WORKBOOK = image_prompter_path("USED-IMAGE-DESIGNS.xlsx")
PROMPT_TEMPLATE_PATH = image_prompter_path("image_edit_prompt_template.txt")
IMAGE_GENERATION_PROMPT_TEMPLATE_PATH = image_prompter_path("image_generation_prompt")
DEFAULT_FIREFOX_BINARY = Path(r"C:\Program Files\Mozilla Firefox\firefox.exe")
FALLBACK_FIREFOX_BINARIES = [
    Path(r"C:\Program Files\Mozilla Firefox\firefox.exe"),
    Path(r"C:\Program Files (x86)\Mozilla Firefox\firefox.exe"),
]
PROMPT_PREVIEW_PATH = RUN_HELPERS_DIR / "generated_prompt_preview.txt"
LAST_FULL_CHAT_PATH = RUN_HELPERS_DIR / "last_full_chat.txt"
LATEST_RESPONSE_PATH = RUN_HELPERS_DIR / "latest_response.txt"
ALL_RESPONSES_PATH = RUN_HELPERS_DIR / "all_responses.txt"
PARSED_IDEAS_PATH = RUN_HELPERS_DIR / "parsed_latest_ideas.txt"
NEW_IDEAS_PATH = RUN_HELPERS_DIR / "new_ideas_not_in_excel.txt"
CURRENT_RUN_IDEA_PATH = RUN_HELPERS_DIR / "current_run_idea.json"
CURRENT_GENERATION_PROMPT_PATH = RUN_HELPERS_DIR / "current_generation_prompt.txt"
IMAGE_GENERATION_FINAL_CHAT_PATH = RUN_HELPERS_DIR / "image_generation_final_chat.txt"
RUN_LOG_PATH = RUN_HELPERS_DIR / "run_log.txt"
IDEA_RESPONSE_POLL_HISTORY_PATH = RUN_HELPERS_DIR / "idea_response_poll_history.txt"
IMAGE_GENERATION_POLL_HISTORY_PATH = RUN_HELPERS_DIR / "image_generation_poll_history.txt"
RUN_HELPER_PATHS = (
    PROMPT_PREVIEW_PATH,
    LAST_FULL_CHAT_PATH,
    LATEST_RESPONSE_PATH,
    ALL_RESPONSES_PATH,
    PARSED_IDEAS_PATH,
    NEW_IDEAS_PATH,
    CURRENT_RUN_IDEA_PATH,
    CURRENT_GENERATION_PROMPT_PATH,
    IMAGE_GENERATION_FINAL_CHAT_PATH,
    RUN_LOG_PATH,
    IDEA_RESPONSE_POLL_HISTORY_PATH,
    IMAGE_GENERATION_POLL_HISTORY_PATH,
)
LEGACY_RUN_HELPER_PATHS = (
    image_prompter_path("generated_prompt_preview.txt"),
    image_prompter_path("last_full_chat.txt"),
    image_prompter_path("latest_response.txt"),
    image_prompter_path("all_responses.txt"),
    image_prompter_path("parsed_latest_ideas.txt"),
    image_prompter_path("new_ideas_not_in_excel.txt"),
    image_prompter_path("current_run_idea.json"),
    image_prompter_path("current_generation_prompt.txt"),
    image_prompter_path("image_generation_final_chat.txt"),
)
CHATGPT_URL = "https://chatgpt.com"
IDEA_MARKER = "IDEA FOR BACKGROUND :"
SUPPORTED_IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp"}
CHATGPT_WINDOW_HINTS = ("chatgpt", "openai")
BROWSER_WINDOW_HINTS = ("firefox",)
PLACEHOLDER_IDEA_VALUES = {"", "...", "n/a", "none", "null", "tbd"}

pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.15
START_HOTKEY_KEY = keyboard.Key.right


def load_pixels() -> dict[str, tuple[int, int]]:
    if not PIXELS_FILE.exists():
        raise FileNotFoundError(
            f"pixels.json not found at: {PIXELS_FILE}\n"
            "Run set_pixels.py first to capture your screen coordinates."
        )
    raw = json.loads(PIXELS_FILE.read_text(encoding="utf-8"))
    return {key: (int(val["x"]), int(val["y"])) for key, val in raw.items()}


_PIXELS = load_pixels()
CHAT_CLICK_TARGET: tuple[int, int] = _PIXELS["chat_neutral_click"]
CHATGPT_PROMPT_BOX_PIXELS_VAIO = {
    "position": _PIXELS["prompt_box_initial"],
    "rgb": (230, 255, 255),
}
CHATGPT_PROMPT_BOX_PIXELS_VAIO_post_injection = {
    "position": _PIXELS["prompt_box_post_injection"],
    "rgb": (230, 255, 255),
}
CHATGPT_BOOT_FOCUS_CLICK_DURATION_SECONDS = 10
IDEA_RESPONSE_ABORT_TIMEOUT_SECONDS = 80
IDEA_RESPONSE_STUCK_PROMPT_RETRY_THRESHOLD = 8
INITIAL_PROMPT_SUBMISSION_WAIT_SECONDS = 10
INITIAL_PROMPT_COMPLETION_DETECTION_DELAY_SECONDS = 5
IMAGE_GENERATION_POLL_INTERVAL_SECONDS = 2.0
IMAGE_GENERATION_ABORT_TIMEOUT_SECONDS = 240
IMAGE_GENERATION_MIN_WAIT_SECONDS = 12
IMAGE_GENERATION_STUCK_PROMPT_RETRY_THRESHOLD = 8
IMAGE_GENERATION_SUBMISSION_WAIT_SECONDS = 10
PROMPT_DISPATCH_TIMEOUT_SECONDS = 20
PROMPT_BOX_RESET_POLL_INTERVAL_SECONDS = 0.5
PROMPT_DISPATCH_RETRY_THRESHOLD = 4
PRE_PASTE_FOCUS_SETTLE_SECONDS = 0.35
POST_TEXT_PASTE_SETTLE_SECONDS = 0.9
POST_IMAGE_PASTE_SETTLE_SECONDS = 0.5
INITIAL_TEXT_INJECTION_RETRY_LIMIT = 3
IMAGE_GENERATION_VERIFICATION_LIMIT = -1
POST_SAVE_EXTRACTION_WAIT_SECONDS = 5.0 
IMAGE_GENERATION_IN_PROGRESS_PHRASES = (
    "creating image",
    "generating image",
    "creating your image",
    "generating your image",
    "making image",
    "making your image",
    "editing image",
    "working on it",
)
IDEA_RESPONSE_IN_PROGRESS_PHRASES = (
    "analyzing image",
    "thinking",
    "working on it",
)


@dataclass
class ProductPromptContext:
    product_kind: str
    image_folder: Path
    image_paths: list[Path]
    used_phrases_csv: str
    prompt_text: str
    existing_phrases: list[str]


@dataclass
class BackgroundIdea:
    title: str
    visual_concept: str
    background_description: str


@dataclass
class ImageGenerationRunResult:
    attempted_images: int
    verified_images: int
    failed_images: list[str]
    final_save_completed: bool
    extracted_output_dir: Path | None


class ImagePrompterError(Exception):
    pass


class FocusError(ImagePrompterError):
    pass


class ParseError(ImagePrompterError):
    pass


class VerificationError(ImagePrompterError):
    pass


class AutomationStepError(ImagePrompterError):
    def __init__(self, step_name: str, original_exception: Exception) -> None:
        super().__init__(f"{step_name} failed: {original_exception}")
        self.step_name = step_name
        self.original_exception = original_exception


class GeneratedImageHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.generated_image_sources: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.casefold() != "img":
            return

        attrs_map = dict(attrs)
        alt_text = (attrs_map.get("alt") or "").strip()
        src = (attrs_map.get("src") or "").strip()
        if alt_text.startswith("Generated image:") and src:
            self.generated_image_sources.append(src)


def ensure_run_helpers_dir() -> None:
    RUN_HELPERS_DIR.mkdir(parents=True, exist_ok=True)


def migrate_legacy_run_helper_files() -> None:
    for legacy_path in LEGACY_RUN_HELPER_PATHS:
        if not legacy_path.exists() or legacy_path.parent == RUN_HELPERS_DIR:
            continue

        target_path = RUN_HELPERS_DIR / legacy_path.name
        if target_path.exists():
            target_path.unlink()
        shutil.move(str(legacy_path), str(target_path))


def reset_run_helper_files() -> None:
    for helper_path in RUN_HELPER_PATHS:
        if helper_path.exists():
            helper_path.unlink()


def initialize_run_helpers() -> None:
    ensure_run_helpers_dir()
    migrate_legacy_run_helper_files()
    reset_run_helper_files()


def append_run_log(message: str) -> None:
    ensure_run_helpers_dir()
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    with RUN_LOG_PATH.open("a", encoding="utf-8") as log_file:
        log_file.write(f"[{timestamp}] {message}\n")


def record_poll_snapshot(path: Path, attempt: int, text: str) -> None:
    ensure_run_helpers_dir()
    with path.open("a", encoding="utf-8") as snapshot_file:
        snapshot_file.write(
            f"===== Attempt {attempt} @ {time.strftime('%Y-%m-%d %H:%M:%S')} =====\n"
        )
        snapshot_file.write(text.strip() + "\n\n")


def run_logged_step(step_name: str, callback, *args, **kwargs):
    append_run_log(f"START {step_name}")
    try:
        result = callback(*args, **kwargs)
    except Exception as exc:
        append_run_log(f"FAIL {step_name}: {exc}")
        raise AutomationStepError(step_name, exc) from exc
    append_run_log(f"OK {step_name}")
    return result


def load_kind_to_used_phrases() -> dict[str, list[str]]:
    if not USED_IMAGE_DESIGNS_WORKBOOK.exists():
        raise FileNotFoundError(
            f"Used image designs workbook was not found: {USED_IMAGE_DESIGNS_WORKBOOK}"
        )

    workbook = load_workbook(USED_IMAGE_DESIGNS_WORKBOOK, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}

    kind_to_phrases: dict[str, list[str]] = {}
    for row in rows[1:]:
        if not row:
            continue
        kind_value = row[0]
        if kind_value is None:
            continue

        kind_label = str(kind_value).strip()
        if not kind_label:
            continue

        phrases = [
            str(cell).strip()
            for cell in row[1:]
            if cell is not None and str(cell).strip()
        ]
        kind_to_phrases[kind_label.upper()] = phrases

    return kind_to_phrases


def prompt_for_kind(kind_to_phrases: dict[str, list[str]]) -> str:
    available_kinds = sorted(kind_to_phrases.keys())
    if not available_kinds:
        raise ValueError("No product kinds were found in USED-IMAGE-DESIGNS.xlsx.")

    print("Select the product kind to prepare:")
    for index, kind in enumerate(available_kinds, start=1):
        print(f"{index}. {kind}")

    while True:
        choice = input("Enter option number: ").strip()
        if not choice.isdigit():
            print("Please enter a valid number.")
            continue

        selected_index = int(choice)
        if 1 <= selected_index <= len(available_kinds):
            return available_kinds[selected_index - 1]

        print("Please choose one of the listed options.")


def get_images_final_kind_dir(product_kind: str) -> Path:
    kind_folder_name = re.sub(r'[<>:"/\\|?*]+', "-", product_kind.strip()).strip(" .")
    if not kind_folder_name:
        raise ValueError("Product kind cannot be empty when creating IMAGES-FINAL folder.")
    return IMAGES_FINAL_DIR / kind_folder_name


def ensure_images_final_kind_folders(product_kinds: list[str]) -> None:
    IMAGES_FINAL_DIR.mkdir(parents=True, exist_ok=True)
    for product_kind in product_kinds:
        kind_dir = get_images_final_kind_dir(product_kind)
        (kind_dir / "0").mkdir(parents=True, exist_ok=True)


def prompt_for_loop_count() -> int:
    while True:
        choice = input("How many full cycles do you want to run? ").strip()
        if not choice.isdigit():
            print("Please enter a valid whole number.")
            continue

        loop_count = int(choice)
        if loop_count >= 1:
            return loop_count

        print("Please enter at least 1.")


def resolve_product_image_folder(product_kind: str) -> Path:
    if not NO_BG_IMAGES_ROOT.exists():
        raise FileNotFoundError(
            f"NO-BG-IMAGES root folder was not found: {NO_BG_IMAGES_ROOT}"
        )

    normalized_target = product_kind.strip().casefold()
    for folder in NO_BG_IMAGES_ROOT.iterdir():
        if folder.is_dir() and folder.name.strip().casefold() == normalized_target:
            return folder

    raise FileNotFoundError(
        f"Could not find an image folder for kind '{product_kind}' inside {NO_BG_IMAGES_ROOT}."
    )


def get_images_from_folder(folder_path: Path) -> list[Path]:
    image_files = sorted(
        [
            file_path
            for file_path in folder_path.iterdir()
            if file_path.is_file() and file_path.suffix.lower() in SUPPORTED_IMAGE_SUFFIXES
        ]
    )
    if not image_files:
        raise FileNotFoundError(f"No supported images were found in: {folder_path}")
    return image_files


def build_used_phrases_csv(phrases: list[str]) -> str:
    if not phrases:
        return "None used yet"
    return ", ".join(phrases)


def build_prompt_text(product_kind: str, used_phrases_csv: str) -> str:
    if not PROMPT_TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Prompt template was not found: {PROMPT_TEMPLATE_PATH}"
        )

    template = PROMPT_TEMPLATE_PATH.read_text(encoding="utf-8")
    return (
        template.replace("[PRODUCT_KIND]", product_kind)
        .replace("[USED_PHRASES_CSV]", used_phrases_csv)
        .strip()
    )


def build_image_generation_prompt(idea: BackgroundIdea) -> str:
    if not IMAGE_GENERATION_PROMPT_TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            "Image generation prompt template was not found: "
            f"{IMAGE_GENERATION_PROMPT_TEMPLATE_PATH}"
        )

    template = IMAGE_GENERATION_PROMPT_TEMPLATE_PATH.read_text(encoding="utf-8")
    background_description = (
        f"Title: {idea.title}\n"
        f"Visual Concept: {idea.visual_concept}\n"
        f"Background Description: {idea.background_description}"
    ).strip()
    return (
        template.replace("[INSERT BACKGROUND DESCRIPTION HERE]", background_description)
        .replace("[BACKGROUND DESCRIPTION]", background_description)
        .strip()
    )


def prepare_product_prompt_context(product_kind: str | None = None) -> ProductPromptContext:
    initialize_run_helpers()
    kind_to_phrases = load_kind_to_used_phrases()
    selected_kind = product_kind or prompt_for_kind(kind_to_phrases)
    image_folder = resolve_product_image_folder(selected_kind)
    image_paths = get_images_from_folder(image_folder)
    used_phrases_csv = build_used_phrases_csv(kind_to_phrases.get(selected_kind, []))
    prompt_text = build_prompt_text(selected_kind, used_phrases_csv)

    PROMPT_PREVIEW_PATH.write_text(prompt_text, encoding="utf-8")

    return ProductPromptContext(
        product_kind=selected_kind,
        image_folder=image_folder,
        image_paths=image_paths,
        used_phrases_csv=used_phrases_csv,
        prompt_text=prompt_text,
        existing_phrases=kind_to_phrases.get(selected_kind, []),
    )


def get_firefox_binary() -> Path:
    firefox_binary = os.getenv("FIREFOX_BINARY")
    if firefox_binary:
        return Path(firefox_binary)

    for candidate in FALLBACK_FIREFOX_BINARIES:
        if candidate.exists():
            return candidate

    path_candidate = shutil.which("firefox.exe") or shutil.which("firefox")
    if path_candidate:
        return Path(path_candidate)

    return DEFAULT_FIREFOX_BINARY


def open_firefox_normal_window() -> None:
    firefox_binary = get_firefox_binary()
    if not firefox_binary.exists():
        raise FileNotFoundError(
            "Firefox binary was not found. Set FIREFOX_BINARY to your firefox.exe path "
            f"or install Firefox in a standard location. Last checked: {firefox_binary}"
        )

    if not PRABHU_FIREFOX_PROFILE.exists():
        raise FileNotFoundError(
            f"Firefox profile directory was not found: {PRABHU_FIREFOX_PROFILE}"
        )

    subprocess.Popen(
        [
            str(firefox_binary),
            "-profile",
            str(PRABHU_FIREFOX_PROFILE),
            CHATGPT_URL,
        ]
    )


def clear_clipboard() -> None:
    subprocess.run(
        ["powershell.exe", "-NoProfile", "-Command", "Set-Clipboard -Value $null"],
        check=True,
    )


def set_clipboard_text(text: str) -> None:
    if not text:
        clear_clipboard()
        return
    powershell_script = r"Set-Clipboard -Value ([Console]::In.ReadToEnd())"
    subprocess.run(
        ["powershell.exe", "-NoProfile", "-Command", powershell_script],
        input=text,
        text=True,
        check=True,
    )


def get_clipboard_text() -> str:
    result = subprocess.run(
        ["powershell.exe", "-NoProfile", "-Command", "Get-Clipboard"],
        capture_output=True,
        text=True,
        check=True,
    )
    return result.stdout.replace("\r\n", "\n").strip()


def set_clipboard_image(image_path: Path) -> None:
    if not image_path.exists():
        raise FileNotFoundError(f"Image file was not found: {image_path}")

    powershell_script = f"""
Add-Type -AssemblyName System.Windows.Forms
Add-Type -AssemblyName System.Drawing
$image = [System.Drawing.Image]::FromFile('{str(image_path).replace("'", "''")}')
[System.Windows.Forms.Clipboard]::SetImage($image)
$image.Dispose()
"""
    subprocess.run(
        ["powershell.exe", "-sta", "-Command", powershell_script],
        check=True,
    )


def get_active_window_title() -> str:
    try:
        active_window = pyautogui.getActiveWindow()
    except Exception:
        return ""
    if active_window is None:
        return ""
    return (active_window.title or "").strip()


def is_chatgpt_window_active(window_title: str) -> bool:
    normalized_title = window_title.casefold()
    return (
        any(hint in normalized_title for hint in CHATGPT_WINDOW_HINTS)
        and any(hint in normalized_title for hint in BROWSER_WINDOW_HINTS)
    )


def prompt_box_matches_expected_state(
    expected_prompt_state: str = "initial",
    tolerance: int = 20,
) -> bool:
    target = (
        CHATGPT_PROMPT_BOX_PIXELS_VAIO
        if expected_prompt_state == "initial"
        else CHATGPT_PROMPT_BOX_PIXELS_VAIO_post_injection
    )
    position = target["position"]
    expected_rgb = target["rgb"]
    try:
        current_rgb = pyautogui.pixel(position[0], position[1])
    except Exception:
        return False
    return all(abs(current_rgb[index] - expected_rgb[index]) <= tolerance for index in range(3))


def prompt_operator_to_recover(message: str) -> None:
    print(message)
    while True:
        choice = input("Type R to retry after fixing focus, or Q to abort this run: ").strip().upper()
        if choice == "R":
            return
        if choice == "Q":
            raise FocusError(message)
        print("Please type R or Q.")


def confirm_clean_prompt_box_visually(action_label: str) -> None:
    print(
        f"Pixel validation is still uncertain before {action_label}. "
        "Visually confirm that the ChatGPT prompt box is empty and focused."
    )
    while True:
        choice = input("Type C to continue, R to retry cleanup, or Q to abort this run: ").strip().upper()
        if choice == "C":
            append_run_log(
                f"Operator visually confirmed a clean prompt box before {action_label}."
            )
            return
        if choice == "R":
            clear_chatgpt_prompt_box()
            if prompt_box_matches_expected_state("initial"):
                append_run_log(
                    f"Prompt box cleanup succeeded after manual retry for {action_label}."
                )
                return
            continue
        if choice == "Q":
            raise FocusError(f"Operator aborted while confirming prompt box before {action_label}.")
        print("Please type C, R, or Q.")


def ensure_chatgpt_window_ready(
    action_label: str,
    *,
    require_prompt_box: bool = True,
    expected_prompt_state: str = "initial",
) -> None:
    while True:
        window_title = get_active_window_title()
        window_ok = is_chatgpt_window_active(window_title)
        prompt_ok = True
        if require_prompt_box:
            prompt_ok = prompt_box_matches_expected_state(expected_prompt_state)

        if window_ok and prompt_ok:
            return

        details = [f"Active window: {window_title or '<unknown>'}"]
        if require_prompt_box:
            details.append(f"Prompt box pixel check passed: {prompt_ok}")
        message = (
            f"Cannot safely {action_label}. "
            + " | ".join(details)
            + ". Bring ChatGPT in Firefox to the foreground and restore the prompt area."
        )
        append_run_log(message)
        prompt_operator_to_recover(message)


def focus_chatgpt_prompt_box(expected_prompt_state: str = "initial") -> None:
    ensure_chatgpt_window_ready(
        "focus the ChatGPT prompt box",
        require_prompt_box=False,
    )
    target = (
        CHATGPT_PROMPT_BOX_PIXELS_VAIO
        if expected_prompt_state == "initial"
        else CHATGPT_PROMPT_BOX_PIXELS_VAIO_post_injection
    )
    pyautogui.click(target["position"])


def clear_chatgpt_prompt_box() -> None:
    focus_chatgpt_prompt_box("initial")
    time.sleep(0.2)
    pyautogui.hotkey("ctrl", "a")
    time.sleep(0.15)
    pyautogui.press("backspace")
    time.sleep(0.35)
    pyautogui.press("esc")
    time.sleep(0.2)


def prepare_prompt_box_for_paste() -> None:
    ensure_chatgpt_window_ready(
        "prepare the ChatGPT prompt box for paste",
        require_prompt_box=False,
    )
    focus_chatgpt_prompt_box("initial")
    time.sleep(PRE_PASTE_FOCUS_SETTLE_SECONDS)


def copy_focused_prompt_box_text(expected_prompt_state: str = "post_injection") -> str:
    focus_chatgpt_prompt_box(expected_prompt_state)
    time.sleep(0.25)
    pyautogui.hotkey("ctrl", "a")
    time.sleep(0.2)
    pyautogui.hotkey("ctrl", "c")
    time.sleep(0.25)
    return get_clipboard_text()


def verify_initial_text_injection(prompt_text: str) -> None:
    normalized_prompt = normalize_phrase(prompt_text)

    for attempt in range(1, INITIAL_TEXT_INJECTION_RETRY_LIMIT + 1):
        copied_text = copy_focused_prompt_box_text("post_injection")
        normalized_copied_text = normalize_phrase(copied_text)
        injection_failed = "chatgpt" in normalized_copied_text
        injection_confirmed = bool(
            normalized_prompt
            and normalized_prompt[:120] in normalized_copied_text
        )

        if not injection_failed and injection_confirmed:
            append_run_log(
                f"Initial text injection confirmed on attempt {attempt}."
            )
            return

        append_run_log(
            "Initial text injection check failed on attempt "
            f"{attempt}. Copied text preview: {copied_text[:120]!r}"
        )
        if attempt < INITIAL_TEXT_INJECTION_RETRY_LIMIT:
            print("Initial text injection did not look correct. Re-pasting the prompt text...")
            paste_text_via_clipboard(prompt_text, "focused ChatGPT prompt box")
            continue

    raise VerificationError(
        "Initial text injection could not be confirmed. The prompt box copy still looked wrong."
    )


def wait_for_prompt_submission_dispatch(
    prompt_text: str,
    action_label: str,
    timeout_seconds: int = PROMPT_DISPATCH_TIMEOUT_SECONDS,
) -> None:
    print(f"Waiting for ChatGPT to accept the submitted {action_label}...")
    dispatch_started_at = time.time()
    attempt = 0
    stuck_counter = 0

    while True:
        attempt += 1
        prompt_box_reset = prompt_box_matches_expected_state("initial")
        current_copy = copy_full_chat_text_once()
        prompt_visible_in_chat = prompt_text.strip() and prompt_text.strip() in current_copy
        prompt_still_stuck_in_input = bool(prompt_text.split()[-10:] and current_copy.split()[-10:] == prompt_text.split()[-10:])

        if prompt_box_reset or (prompt_visible_in_chat and not prompt_still_stuck_in_input):
            append_run_log(
                f"Prompt dispatch confirmed for {action_label} on attempt {attempt}."
            )
            return

        if prompt_still_stuck_in_input:
            stuck_counter += 1
            if stuck_counter >= PROMPT_DISPATCH_RETRY_THRESHOLD:
                append_run_log(
                    f"Prompt still appears unsent for {action_label}; pressing Enter again."
                )
                focus_chatgpt_prompt_box("post_injection")
                time.sleep(0.3)
                pyautogui.press("enter")
                time.sleep(1.5)
                stuck_counter = 0
        else:
            stuck_counter = 0

        if time.time() - dispatch_started_at >= timeout_seconds:
            raise VerificationError(
                f"ChatGPT did not accept the submitted {action_label} within {timeout_seconds} seconds."
            )

        time.sleep(PROMPT_BOX_RESET_POLL_INTERVAL_SECONDS)


def paste_text_via_clipboard(text: str, field_label: str) -> None:
    prepare_prompt_box_for_paste()
    set_clipboard_text(text)
    time.sleep(PRE_PASTE_FOCUS_SETTLE_SECONDS)
    print(f"Pasting prompt text into {field_label}...")
    pyautogui.hotkey("ctrl", "v")
    time.sleep(POST_TEXT_PASTE_SETTLE_SECONDS)


def paste_image_via_clipboard(image_path: Path, field_label: str) -> None:
    ensure_chatgpt_window_ready(
        f"paste an image into {field_label}",
        require_prompt_box=False,
    )
    print(f"Loading image into clipboard: {image_path}")
    set_clipboard_image(image_path)
    time.sleep(PRE_PASTE_FOCUS_SETTLE_SECONDS)
    print(f"Pasting image into {field_label}...")
    pyautogui.hotkey("ctrl", "v")
    time.sleep(POST_IMAGE_PASTE_SETTLE_SECONDS)
    clear_clipboard()


def click_chat_copy_target() -> None:
    ensure_chatgpt_window_ready(
        "click the chat copy target",
        require_prompt_box=False,
    )
    target_x, target_y = CHAT_CLICK_TARGET
    print(f"Clicking chat copy target at ({target_x}, {target_y}) before copy cycle...")
    pyautogui.moveTo(target_x, target_y, duration=0.2)
    pyautogui.click()


def hold_click_chatgpt_boot_focus_target() -> None:
    target_x, target_y = CHATGPT_PROMPT_BOX_PIXELS_VAIO["position"]
    print(
        "Clicking ChatGPT boot focus target "
        f"({target_x}, {target_y}) every 0.5 seconds for "
        f"{CHATGPT_BOOT_FOCUS_CLICK_DURATION_SECONDS} seconds..."
    )
    pyautogui.moveTo(target_x, target_y, duration=0.2)
    end_time = time.time() + CHATGPT_BOOT_FOCUS_CLICK_DURATION_SECONDS
    while time.time() < end_time:
        pyautogui.click()
        time.sleep(0.5)


def copy_full_chat_text_once() -> str:
    ensure_chatgpt_window_ready(
        "copy the current ChatGPT conversation",
        require_prompt_box=False,
    )
    pyautogui.press("esc")
    time.sleep(0.25)
    click_chat_copy_target()
    time.sleep(0.25)
    pyautogui.hotkey("ctrl", "a")
    time.sleep(0.3)
    pyautogui.hotkey("ctrl", "c")
    time.sleep(0.3)
    return get_clipboard_text()


def extract_text_after_prompt(full_chat_text: str, prompt_text: str) -> str:
    normalized_chat = full_chat_text.replace("\r\n", "\n").strip()
    normalized_prompt = prompt_text.replace("\r\n", "\n").strip()
    if not normalized_chat or not normalized_prompt:
        return normalized_chat

    prompt_index = normalized_chat.rfind(normalized_prompt)
    if prompt_index == -1:
        return normalized_chat

    return normalized_chat[prompt_index + len(normalized_prompt):].strip()


def has_meaningful_trailing_content(
    full_chat_text: str,
    prompt_text: str,
    minimum_length: int = 20,
) -> bool:
    trailing_text = extract_text_after_prompt(full_chat_text, prompt_text)
    return len(trailing_text) >= minimum_length


def is_idea_response_in_progress(full_chat_text: str) -> bool:
    normalized_text = full_chat_text.casefold()
    return any(phrase in normalized_text for phrase in IDEA_RESPONSE_IN_PROGRESS_PHRASES)


def wait_for_stable_full_chat_text(
    prompt_text: str = "",
    timeout_started_at: float | None = None,
) -> str | None:
    print(
        "Starting full-chat copy cycle every 0.5 seconds until two consecutive copies match and idea output is present..."
    )
    previous_copy: str | None = None
    attempt = 0
    stuck_counter = 0

    while True:
        attempt += 1
        current_copy = copy_full_chat_text_once()
        record_poll_snapshot(IDEA_RESPONSE_POLL_HISTORY_PATH, attempt, current_copy)
        print(f"Captured full chat copy attempt {attempt}.")
        is_stable_copy = bool(current_copy and previous_copy == current_copy)
        has_ideas = bool(parse_ideas(current_copy))
        response_started = has_meaningful_trailing_content(current_copy, prompt_text)
        still_in_progress = is_idea_response_in_progress(current_copy)

        if is_stable_copy and response_started and has_ideas and not still_in_progress:
            print(
                "Detected stable copied chat text with parseable idea output and trailing response content."
            )
            return current_copy

        prompt_tail = prompt_text.split()[-10:]
        copy_tail = current_copy.split()[-10:]
        is_stuck = bool(prompt_tail and copy_tail == prompt_tail)
        print(
            "Stuck check: "
            f"{is_stuck} (counter: {stuck_counter + 1 if is_stuck else 0}/"
            f"{IDEA_RESPONSE_STUCK_PROMPT_RETRY_THRESHOLD})"
        )
        if is_stuck:
            stuck_counter += 1
            if stuck_counter >= IDEA_RESPONSE_STUCK_PROMPT_RETRY_THRESHOLD:
                print("Detected prompt still in input box (ENTER might have failed). Repressing ENTER...")
                focus_chatgpt_prompt_box("post_injection")
                time.sleep(0.5)
                pyautogui.press("enter")
                time.sleep(3)
                stuck_counter = 0
        else:
            stuck_counter = 0

        if is_stable_copy and not has_ideas:
            print(
                "Copied chat text is stable, but no fully parseable idea blocks were found yet. Continuing to wait..."
            )

        if (
            timeout_started_at is not None
            and time.time() - timeout_started_at >= IDEA_RESPONSE_ABORT_TIMEOUT_SECONDS
        ):
            print(
                "Timed out after "
                f"{IDEA_RESPONSE_ABORT_TIMEOUT_SECONDS} seconds without detecting "
                "a complete parseable idea response. Aborting this run."
            )
            return None

        previous_copy = current_copy
        time.sleep(0.5)


def load_existing_text(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8").replace("\r\n", "\n").strip()


def append_text_block(path: Path, text: str) -> None:
    existing_text = load_existing_text(path)
    combined_text = f"{existing_text}\n\n{text}".strip() if existing_text else text.strip()
    path.write_text(combined_text + "\n", encoding="utf-8")


def strip_duplicated_prompt_prefix(text: str, prompt_text: str) -> str:
    normalized_text = text.strip()
    normalized_prompt = prompt_text.strip()
    duplicated_prompt = normalized_prompt + normalized_prompt

    if normalized_prompt and normalized_text.startswith(duplicated_prompt):
        return normalized_text[len(normalized_prompt):].strip()

    return normalized_text


def normalize_phrase(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip().casefold()


def normalize_idea_field(value: object) -> str:
    return re.sub(r"\s+", " ", str(value)).strip()


def is_placeholder_idea_value(value: str) -> bool:
    return normalize_phrase(value) in PLACEHOLDER_IDEA_VALUES


def is_valid_ideas_payload(payload: dict[str, object]) -> bool:
    ideas = payload.get("ideas")
    if not isinstance(ideas, list) or not ideas:
        return False

    valid_items = 0
    for item in ideas:
        if not isinstance(item, dict):
            return False
        if background_idea_from_json_item(item) is not None:
            valid_items += 1

    return valid_items > 0


def extract_json_payload(text: str) -> dict[str, object] | None:
    decoder = json.JSONDecoder()
    candidate_start_indexes = [match.start() for match in re.finditer(r'\{\s*"ideas"\s*:', text)]
    valid_payloads: list[dict[str, object]] = []

    for start_index in candidate_start_indexes:
        try:
            payload, _ = decoder.raw_decode(text[start_index:].strip())
        except json.JSONDecodeError:
            continue

        if isinstance(payload, dict) and is_valid_ideas_payload(payload):
            valid_payloads.append(payload)

    if valid_payloads:
        return valid_payloads[-1]

    return None


def background_idea_from_json_item(item: dict[str, object]) -> BackgroundIdea | None:
    title = normalize_idea_field(item.get("title", ""))
    visual_concept = normalize_idea_field(item.get("visual_concept", ""))
    background_description = normalize_idea_field(item.get("background_description", ""))
    if (
        not title
        or not visual_concept
        or not background_description
        or is_placeholder_idea_value(title)
        or is_placeholder_idea_value(visual_concept)
        or is_placeholder_idea_value(background_description)
    ):
        return None
    return BackgroundIdea(
        title=title,
        visual_concept=visual_concept,
        background_description=background_description,
    )


def dedupe_ideas_by_title(ideas: list[BackgroundIdea]) -> list[BackgroundIdea]:
    seen_titles: set[str] = set()
    deduped_ideas: list[BackgroundIdea] = []
    for idea in ideas:
        normalized_title = normalize_phrase(idea.title)
        if normalized_title in seen_titles:
            continue
        seen_titles.add(normalized_title)
        deduped_ideas.append(idea)
    return deduped_ideas


def background_idea_to_block(idea: BackgroundIdea) -> str:
    lines = [f"{IDEA_MARKER} {idea.title}"]
    if idea.visual_concept:
        lines.append(f"Visual Concept: {idea.visual_concept}")
    if idea.background_description:
        lines.append(f"Background Description: {idea.background_description}")
    return "\n".join(lines).strip()


def parse_ideas_from_json(text: str) -> list[BackgroundIdea]:
    payload = extract_json_payload(text)
    if not payload:
        return []

    ideas = payload.get("ideas")
    if not isinstance(ideas, list):
        return []

    parsed_ideas: list[BackgroundIdea] = []
    rejected_items = 0
    for item in ideas:
        if not isinstance(item, dict):
            rejected_items += 1
            continue
        parsed_idea = background_idea_from_json_item(item)
        if parsed_idea is None:
            rejected_items += 1
            continue
        parsed_ideas.append(parsed_idea)

    if rejected_items:
        append_run_log(
            f"Rejected {rejected_items} malformed idea item(s) while parsing JSON response."
        )

    return dedupe_ideas_by_title(parsed_ideas)


def parse_ideas_from_marker_blocks(text: str) -> list[BackgroundIdea]:
    marker_matches = list(re.finditer(re.escape(IDEA_MARKER), text))
    if not marker_matches:
        return []

    ideas: list[BackgroundIdea] = []
    for index, match in enumerate(marker_matches):
        start = match.start()
        end = (
            marker_matches[index + 1].start()
            if index + 1 < len(marker_matches)
            else len(text)
        )
        block = text[start:end].strip()
        if not block:
            continue

        lines = [line.strip() for line in block.splitlines() if line.strip()]
        if not lines:
            continue

        first_line = lines[0]
        if not first_line.startswith(IDEA_MARKER):
            continue

        title = first_line[len(IDEA_MARKER):].strip()
        visual_concept = ""
        background_description = ""

        for line in lines[1:]:
            if line.startswith("Visual Concept:"):
                visual_concept = line.split(":", 1)[1].strip()
            elif line.startswith("Background Description:"):
                background_description = line.split(":", 1)[1].strip()

        if title and visual_concept and background_description:
            ideas.append(
                BackgroundIdea(
                    title=title,
                    visual_concept=visual_concept,
                    background_description=background_description,
                )
            )

    return ideas


def parse_ideas(text: str) -> list[BackgroundIdea]:
    json_ideas = parse_ideas_from_json(text)
    if json_ideas:
        return dedupe_ideas_by_title(json_ideas)
    return dedupe_ideas_by_title(parse_ideas_from_marker_blocks(text))


def extract_idea_title(idea: BackgroundIdea) -> str:
    return idea.title.strip()


def format_idea_blocks(ideas: list[BackgroundIdea]) -> str:
    if not ideas:
        return ""
    return "\n\n".join(background_idea_to_block(idea) for idea in ideas).strip()


def get_new_ideas(ideas: list[BackgroundIdea], existing_phrases: list[str]) -> list[BackgroundIdea]:
    existing_phrase_keys = {normalize_phrase(phrase) for phrase in existing_phrases if phrase.strip()}
    new_ideas: list[BackgroundIdea] = []

    for idea in ideas:
        idea_title = extract_idea_title(idea)
        if normalize_phrase(idea_title) not in existing_phrase_keys:
            new_ideas.append(idea)

    return new_ideas


def choose_current_idea(ideas: list[BackgroundIdea], existing_phrases: list[str]) -> BackgroundIdea:
    new_ideas = get_new_ideas(ideas, existing_phrases)
    if new_ideas:
        return new_ideas[0]
    if ideas:
        return ideas[0]
    raise ValueError("No background ideas were parsed from the latest output.")


def append_phrase_to_workbook(product_kind: str, phrase_title: str) -> None:
    workbook = load_workbook(USED_IMAGE_DESIGNS_WORKBOOK)
    worksheet = workbook[workbook.sheetnames[0]]
    target_kind = normalize_phrase(product_kind)

    for row_index in range(2, worksheet.max_row + 1):
        kind_value = worksheet.cell(row=row_index, column=1).value
        if kind_value is None:
            continue
        if normalize_phrase(str(kind_value)) != target_kind:
            continue

        existing_row_phrases = []
        for column_index in range(2, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=row_index, column=column_index).value
            if cell_value is not None and str(cell_value).strip():
                existing_row_phrases.append(str(cell_value).strip())

        if normalize_phrase(phrase_title) in {
            normalize_phrase(phrase) for phrase in existing_row_phrases
        }:
            workbook.save(USED_IMAGE_DESIGNS_WORKBOOK)
            return

        target_column = 2
        while worksheet.cell(row=row_index, column=target_column).value not in (None, ""):
            target_column += 1

        worksheet.cell(row=row_index, column=target_column).value = phrase_title
        workbook.save(USED_IMAGE_DESIGNS_WORKBOOK)
        return

    raise ValueError(f"Could not find workbook row for kind '{product_kind}'.")


def save_current_run_idea(idea: BackgroundIdea) -> None:
    payload = {
        "title": idea.title,
        "visual_concept": idea.visual_concept,
        "background_description": idea.background_description,
    }
    CURRENT_RUN_IDEA_PATH.write_text(
        json.dumps(payload, indent=2, ensure_ascii=True) + "\n",
        encoding="utf-8",
    )


def extract_latest_output(full_chat_text: str, prompt_text: str) -> str:
    normalized_full_chat = full_chat_text.replace("\r\n", "\n").strip()
    previous_full_chat = load_existing_text(LAST_FULL_CHAT_PATH)
    previous_outputs = load_existing_text(ALL_RESPONSES_PATH)

    latest_output = normalized_full_chat

    if previous_full_chat and normalized_full_chat.startswith(previous_full_chat):
        latest_output = normalized_full_chat[len(previous_full_chat):].strip()
    else:
        if prompt_text and prompt_text in latest_output:
            latest_output = latest_output.split(prompt_text, 1)[-1].strip()
        if previous_outputs and previous_outputs in latest_output:
            latest_output = latest_output.replace(previous_outputs, "", 1).strip()

    return strip_duplicated_prompt_prefix(latest_output, prompt_text)


def capture_and_store_latest_output(
    prompt_text: str,
    timeout_started_at: float | None = None,
) -> str | None:
    click_chat_copy_target()
    # stable_full_chat_text = wait_for_stable_full_chat_text()
    stable_full_chat_text = wait_for_stable_full_chat_text(
        prompt_text,
        timeout_started_at,
    )
    if stable_full_chat_text is None:
        return None

    latest_output = extract_latest_output(stable_full_chat_text, prompt_text)

    LAST_FULL_CHAT_PATH.write_text(stable_full_chat_text + "\n", encoding="utf-8")
    LATEST_RESPONSE_PATH.write_text(latest_output + "\n", encoding="utf-8")
    if latest_output:
        append_text_block(ALL_RESPONSES_PATH, latest_output)

    print(f"Saved latest response to: {LATEST_RESPONSE_PATH}")
    print(f"Saved full chat snapshot to: {LAST_FULL_CHAT_PATH}")
    return latest_output


def save_parsed_idea_results(latest_output: str, existing_phrases: list[str]) -> list[BackgroundIdea]:
    ideas = parse_ideas(latest_output)
    new_ideas = get_new_ideas(ideas, existing_phrases)

    PARSED_IDEAS_PATH.write_text(
        format_idea_blocks(ideas) + ("\n" if ideas else ""),
        encoding="utf-8",
    )
    NEW_IDEAS_PATH.write_text(
        format_idea_blocks(new_ideas) + ("\n" if new_ideas else ""),
        encoding="utf-8",
    )

    print(f"Saved parsed idea blocks to: {PARSED_IDEAS_PATH}")
    print(f"Saved ideas not found in Excel to: {NEW_IDEAS_PATH}")
    return ideas


def prepare_current_generation_prompt(idea: BackgroundIdea) -> str:
    prompt_text = build_image_generation_prompt(idea)
    CURRENT_GENERATION_PROMPT_PATH.write_text(prompt_text + "\n", encoding="utf-8")
    print(f"Saved current generation prompt to: {CURRENT_GENERATION_PROMPT_PATH}")
    return prompt_text


def beep_ready_for_generation_prompt() -> None:
    winsound.MessageBeep(winsound.MB_ICONASTERISK)
    time.sleep(0.2)
    winsound.MessageBeep(winsound.MB_ICONASTERISK)


def beep_image_generation_complete() -> None:
    winsound.MessageBeep(winsound.MB_OK)
    time.sleep(0.25)
    winsound.MessageBeep(winsound.MB_OK)
    time.sleep(0.25)
    winsound.MessageBeep(winsound.MB_OK)


def beep_all_images_generation_complete() -> None:
    for frequency, duration_ms in (
        (880, 250),
        (988, 250),
        (1175, 450),
    ):
        winsound.Beep(frequency, duration_ms)
        time.sleep(0.12)


def save_generated_images_to_output_folder() -> None:
    output_folder = FULL_GENERATED_IMAGES_DIR.resolve()
    if not output_folder.exists():
        raise FileNotFoundError(
            f"Full generated images folder was not found: {output_folder}"
        )

    ensure_chatgpt_window_ready(
        "open the browser save dialog",
        require_prompt_box=False,
    )
    print(f"Saving generated images into: {output_folder}")
    pyautogui.hotkey("ctrl", "s")
    time.sleep(2.0)
    pyautogui.hotkey("alt", "d")
    time.sleep(0.35)
    set_clipboard_text(str(output_folder))
    time.sleep(0.35)
    pyautogui.hotkey("ctrl", "v")
    time.sleep(0.35)
    pyautogui.press("enter")
    time.sleep(0.5)

    for _ in range(6):
        pyautogui.press("tab")
        time.sleep(0.12)

    set_clipboard_text(time.strftime("%H-%M-%S"))
    time.sleep(0.2)
    pyautogui.hotkey("ctrl", "v")
    time.sleep(0.2)
    pyautogui.press("enter")


def get_latest_saved_html_path() -> Path:
    html_candidates = sorted(
        [
            path
            for path in FULL_GENERATED_IMAGES_DIR.iterdir()
            if path.is_file() and path.suffix.lower() in {".htm", ".html"}
        ],
        key=lambda path: path.stat().st_mtime,
    )
    if not html_candidates:
        raise FileNotFoundError(
            f"No saved HTML files were found in: {FULL_GENERATED_IMAGES_DIR}"
        )
    return html_candidates[-1]


def get_next_images_final_output_dir(product_kind: str) -> Path:
    images_final_kind_dir = get_images_final_kind_dir(product_kind)
    images_final_kind_dir.mkdir(parents=True, exist_ok=True)
    numeric_folders = [
        int(path.name)
        for path in images_final_kind_dir.iterdir()
        if path.is_dir() and path.name.isdigit()
    ]
    next_folder_number = max(numeric_folders, default=-1) + 1
    output_dir = images_final_kind_dir / str(next_folder_number)
    output_dir.mkdir(parents=True, exist_ok=False)
    return output_dir


def extract_generated_image_sources_from_html(html_path: Path) -> list[str]:
    html_text = html_path.read_text(encoding="utf-8", errors="replace")
    parser = GeneratedImageHTMLParser()
    parser.feed(html_text)
    return parser.generated_image_sources


def copy_or_download_generated_image(
    image_src: str,
    html_path: Path,
    destination_path: Path,
) -> None:
    decoded_src = unquote(image_src)
    local_source_path = (html_path.parent / decoded_src).resolve()
    if local_source_path.exists():
        shutil.copy2(local_source_path, destination_path)
        return

    with urllib.request.urlopen(image_src) as response:
        destination_path.write_bytes(response.read())


def extract_generated_images_from_latest_saved_html(product_kind: str) -> Path | None:
    latest_html_path = get_latest_saved_html_path()
    generated_sources = extract_generated_image_sources_from_html(latest_html_path)
    if not generated_sources:
        print(
            "No generated-image <img> tags were found in the saved HTML, "
            f"so skipping final image extraction for this cycle: {latest_html_path}"
        )
        return None

    output_dir = get_next_images_final_output_dir(product_kind)
    print(f"Extracting generated images from: {latest_html_path}")
    print(f"Saving ordered generated images to: {output_dir}")

    for image_index, image_src in enumerate(generated_sources, start=1):
        parsed_src = urlparse(image_src)
        image_suffix = Path(parsed_src.path).suffix or ".png"
        destination_path = output_dir / f"{image_index}{image_suffix}"
        copy_or_download_generated_image(
            image_src=image_src,
            html_path=latest_html_path,
            destination_path=destination_path,
        )

    return output_dir


def is_image_generation_in_progress(full_chat_text: str) -> bool:
    normalized_text = full_chat_text.casefold()
    return any(
        phrase in normalized_text for phrase in IMAGE_GENERATION_IN_PROGRESS_PHRASES
    )


def has_generated_image_confirmation(
    full_chat_text: str,
    generation_prompt_text: str,
) -> bool:
    normalized_chat = full_chat_text.replace("\r\n", "\n").strip()
    normalized_prompt = generation_prompt_text.replace("\r\n", "\n").strip()
    if not normalized_chat or not normalized_prompt:
        return False

    prompt_index = normalized_chat.rfind(normalized_prompt)
    if prompt_index == -1:
        return False

    trailing_text = normalized_chat[prompt_index + len(normalized_prompt):]
    return "generated image:" in trailing_text.casefold()


def wait_for_image_generation_completion(generation_prompt_text: str) -> str | None:
    print(
        "Waiting for image-generation mode to end by polling copied chat text for generating-state phrases..."
    )
    start_time = time.time()
    time.sleep(IMAGE_GENERATION_MIN_WAIT_SECONDS)

    previous_copy: str | None = None
    attempt = 0
    stuck_counter = 0

    while True:
        attempt += 1
        current_copy = copy_full_chat_text_once()
        record_poll_snapshot(IMAGE_GENERATION_POLL_HISTORY_PATH, attempt, current_copy)
        print(f"Checked image-generation status attempt {attempt}.")
        is_stable_copy = bool(current_copy and previous_copy == current_copy)
        has_generated_confirmation = has_generated_image_confirmation(
            current_copy,
            generation_prompt_text,
        )
        response_started = has_meaningful_trailing_content(current_copy, generation_prompt_text)

        if (
            is_stable_copy
            and response_started
            and not is_image_generation_in_progress(current_copy)
            and has_generated_confirmation
        ):
            print(
                "Detected stable copied chat text with a generated-image confirmation after the injected prompt."
            )
            IMAGE_GENERATION_FINAL_CHAT_PATH.write_text(
                current_copy + "\n",
                encoding="utf-8",
            )
            beep_image_generation_complete()
            print(
                f"Saved final image-generation chat snapshot to: {IMAGE_GENERATION_FINAL_CHAT_PATH}"
            )
            return current_copy

        gen_prompt_tail = generation_prompt_text.split()[-10:]
        copy_tail = current_copy.split()[-10:]
        is_stuck = bool(gen_prompt_tail and copy_tail == gen_prompt_tail)
        print(
            "Stuck check: "
            f"{is_stuck} (counter: {stuck_counter + 1 if is_stuck else 0}/"
            f"{IMAGE_GENERATION_STUCK_PROMPT_RETRY_THRESHOLD})"
        )
        if is_stuck:
            stuck_counter += 1
            if stuck_counter >= IMAGE_GENERATION_STUCK_PROMPT_RETRY_THRESHOLD:
                print("Detected generation prompt still in input box. Repressing ENTER...")
                focus_chatgpt_prompt_box("post_injection")
                time.sleep(0.5)
                pyautogui.press("enter")
                time.sleep(3)
                stuck_counter = 0
        else:
            stuck_counter = 0

        if is_stable_copy and not has_generated_confirmation:
            print(
                "Copied chat text is stable, but 'Generated image:' was not found after the injected prompt yet. Continuing to wait..."
            )

        if time.time() - start_time >= IMAGE_GENERATION_ABORT_TIMEOUT_SECONDS:
            print(
                "Timed out after "
                f"{IMAGE_GENERATION_ABORT_TIMEOUT_SECONDS} seconds without verifying a generated image. "
                "Aborting this image run and continuing."
            )
            IMAGE_GENERATION_FINAL_CHAT_PATH.write_text(
                current_copy + "\n",
                encoding="utf-8",
            )
            return None

        previous_copy = current_copy
        time.sleep(IMAGE_GENERATION_POLL_INTERVAL_SECONDS)

def run_generation_prompt_for_image(
    image_path: Path,
    generation_prompt_text: str,
) -> str | None:
    print("Starting follow-up image generation prompt and image paste flow...")
    focus_chatgpt_prompt_box("initial")
    time.sleep(0.8)
    paste_text_via_clipboard(generation_prompt_text, "focused ChatGPT prompt box")
    time.sleep(1.5)
    paste_image_via_clipboard(image_path, "focused ChatGPT prompt box")
    print(
        "Waiting "
        f"{IMAGE_GENERATION_SUBMISSION_WAIT_SECONDS} seconds before submitting "
        "the image generation prompt..."
    )
    time.sleep(IMAGE_GENERATION_SUBMISSION_WAIT_SECONDS)
    ensure_chatgpt_window_ready(
        "submit the image generation prompt",
        require_prompt_box=False,
    )
    focus_chatgpt_prompt_box("post_injection")
    time.sleep(0.2)
    pyautogui.press("enter")
    print("Pressed Enter to submit the image generation prompt.")
    wait_for_prompt_submission_dispatch(
        generation_prompt_text,
        "image generation prompt",
    )
    return wait_for_image_generation_completion(generation_prompt_text)


def run_generation_prompt_for_remaining_images(
    image_paths: list[Path],
    generation_prompt_text: str,
    product_kind: str,
) -> ImageGenerationRunResult:
    if IMAGE_GENERATION_VERIFICATION_LIMIT == -1:
        target_verification_count = len(image_paths)
    elif IMAGE_GENERATION_VERIFICATION_LIMIT < -1:
        raise ValueError(
            "IMAGE_GENERATION_VERIFICATION_LIMIT must be -1 or a non-negative integer."
        )
    else:
        target_verification_count = min(
            IMAGE_GENERATION_VERIFICATION_LIMIT,
            len(image_paths),
        )

    if target_verification_count == 0:
        print(
            "IMAGE_GENERATION_VERIFICATION_LIMIT is 0, so skipping image verification and moving directly to the final save flow."
        )
        save_generated_images_to_output_folder()
        time.sleep(POST_SAVE_EXTRACTION_WAIT_SECONDS)
        extracted_output_dir = extract_generated_images_from_latest_saved_html(product_kind)
        beep_all_images_generation_complete()
        return ImageGenerationRunResult(
            attempted_images=0,
            verified_images=0,
            failed_images=[],
            final_save_completed=True,
            extracted_output_dir=extracted_output_dir,
        )

    print(
        f"Will verify {target_verification_count} generated image(s) before the final save flow."
    )

    verified_images = 0
    failed_images: list[str] = []

    for image_index, image_path in enumerate(
        image_paths[:target_verification_count],
        start=1,
    ):
        print()
        print(
            f"Running image generation for image {image_index} of {target_verification_count}: {image_path}"
        )
        final_chat_text = run_generation_prompt_for_image(image_path, generation_prompt_text)
        if final_chat_text is None:
            failure_message = (
                f"Image {image_index} could not be verified. Skipping final save flow for this cycle."
            )
            print(failure_message)
            append_run_log(failure_message)
            failed_images.append(str(image_path))
            continue

        verified_images += 1
        print(
            f"Confirmed generated image for image {image_index} of {target_verification_count}."
        )

    if verified_images != target_verification_count:
        print(
            "Not every required image was verified, so the script will not claim completion or trigger the final save flow."
        )
        return ImageGenerationRunResult(
            attempted_images=target_verification_count,
            verified_images=verified_images,
            failed_images=failed_images,
            final_save_completed=False,
            extracted_output_dir=None,
        )

    if target_verification_count == len(image_paths):
        print("Confirmed generated images for every image in the folder.")
    else:
        print(
            "Reached the configured image-generation verification limit. Moving to the final save flow."
        )

    save_generated_images_to_output_folder()
    time.sleep(POST_SAVE_EXTRACTION_WAIT_SECONDS)
    extracted_output_dir = extract_generated_images_from_latest_saved_html(product_kind)
    beep_all_images_generation_complete()
    return ImageGenerationRunResult(
        attempted_images=target_verification_count,
        verified_images=verified_images,
        failed_images=failed_images,
        final_save_completed=True,
        extracted_output_dir=extracted_output_dir,
    )

def run_chatgpt_manual_browser_flow(context: ProductPromptContext) -> None:
    print()
    print("Firefox will open as a normal browser window.")
    print("Then manually go to ChatGPT, open the page you want, and keep it visible.")
    print("Keep the ChatGPT prompt box focused before pressing the Right Arrow key.")
    print()

    run_logged_step("open Firefox", open_firefox_normal_window)
    time.sleep(2)
    run_logged_step("focus ChatGPT boot target", hold_click_chatgpt_boot_focus_target)

    print()
    print("Starting focused-field prompt and image paste flow...")
    run_logged_step(
        "paste initial text prompt",
        paste_text_via_clipboard,
        context.prompt_text,
        "focused ChatGPT prompt box",
    )
    run_logged_step(
        "verify initial text injection",
        verify_initial_text_injection,
        context.prompt_text,
    )
    time.sleep(0.9)
    run_logged_step(
        "paste initial reference image",
        paste_image_via_clipboard,
        context.image_paths[0],
        "focused ChatGPT prompt box",
    )
    print(
        "Waiting "
        f"{INITIAL_PROMPT_SUBMISSION_WAIT_SECONDS} seconds before submitting "
        "the ChatGPT prompt..."
    )
    time.sleep(INITIAL_PROMPT_SUBMISSION_WAIT_SECONDS)
    ensure_chatgpt_window_ready(
        "submit the initial ChatGPT prompt",
        require_prompt_box=False,
    )
    focus_chatgpt_prompt_box("post_injection")
    time.sleep(0.2)
    pyautogui.press("enter")
    prompt_submitted_at = time.time()
    print("Pressed Enter to submit the prompt")
    wait_for_prompt_submission_dispatch(
        context.prompt_text,
        "initial idea prompt",
    )
    print(
        "Waiting "
        f"{INITIAL_PROMPT_COMPLETION_DETECTION_DELAY_SECONDS} seconds before "
        "starting output-completion detection..."
    )
    time.sleep(INITIAL_PROMPT_COMPLETION_DETECTION_DELAY_SECONDS)
    latest_output = run_logged_step(
        "capture latest idea response",
        capture_and_store_latest_output,
        context.prompt_text,
        prompt_submitted_at,
    )
    if not latest_output:
        raise VerificationError(
            "No new latest output text could be isolated from the copied conversation."
        )

    ideas = run_logged_step(
        "parse latest idea response",
        save_parsed_idea_results,
        latest_output,
        context.existing_phrases,
    )
    print("Captured latest output successfully.")
    if not ideas:
        raise ParseError("No parsed ideas were found in the latest output.")

    new_titles = [idea.title for idea in get_new_ideas(ideas, context.existing_phrases)]
    if new_titles:
        print("Ideas not found in Excel:")
        for title in new_titles:
            print(f"- {title}")

    current_idea = choose_current_idea(ideas, context.existing_phrases)
    generation_prompt_text = run_logged_step(
        "prepare current generation prompt",
        prepare_current_generation_prompt,
        current_idea,
    )

    print(f"Selected CURRENT IDEA: {current_idea.title}")
    save_current_run_idea(current_idea)
    print(f"Saved current run idea JSON to: {CURRENT_RUN_IDEA_PATH}")
    generation_result = run_logged_step(
        "generate images for remaining product files",
        run_generation_prompt_for_remaining_images,
        context.image_paths,
        generation_prompt_text,
        context.product_kind,
    )

    if not generation_result.final_save_completed:
        raise VerificationError(
            f"Generation finished with only {generation_result.verified_images}/"
            f"{generation_result.attempted_images} verified image(s)."
        )

    run_logged_step(
        "record successful idea in workbook",
        append_phrase_to_workbook,
        context.product_kind,
        current_idea.title,
    )

def wait_for_start_hotkey() -> None:
    print(
        "After ChatGPT is open and visible in Firefox, press the Right Arrow key from the browser to start."
    )

    start_detected = {"pressed": False}

    def on_press(key: keyboard.Key | keyboard.KeyCode) -> bool | None:
        try:
            if key == START_HOTKEY_KEY:
                start_detected["pressed"] = True
                print("Detected start hotkey: Right Arrow")
                return False
        except Exception:
            return None
        return None

    with keyboard.Listener(on_press=on_press) as listener:
        listener.join()

    if not start_detected["pressed"]:
        raise RuntimeError("Start hotkey was not detected.")


def main() -> None:
    loop_count = prompt_for_loop_count()
    kind_to_phrases = load_kind_to_used_phrases()
    ensure_images_final_kind_folders(sorted(kind_to_phrases.keys()))
    selected_kind = prompt_for_kind(kind_to_phrases)

    append_run_log("Starting image_prompter run.")

    for cycle_index in range(1, loop_count + 1):
        print()
        print(f"========== Starting cycle {cycle_index} of {loop_count} ==========")
        append_run_log(f"Starting cycle {cycle_index} of {loop_count}.")

        try:
            context = run_logged_step(
                "prepare product prompt context",
                prepare_product_prompt_context,
                selected_kind,
            )

            print(f"Selected kind: {context.product_kind}")
            print(f"First image ready: {context.image_paths[0]}")
            print(f"Total images queued for generation: {len(context.image_paths)}")
            print(f"Prompt preview saved to: {PROMPT_PREVIEW_PATH}")

            run_chatgpt_manual_browser_flow(context)
        except AutomationStepError as exc:
            print(f"Cycle {cycle_index} failed during {exc.step_name}: {exc.original_exception}")
            append_run_log(
                f"Cycle {cycle_index} failed during {exc.step_name}: {exc.original_exception}"
            )
            continue
        except ImagePrompterError as exc:
            print(f"Cycle {cycle_index} aborted: {exc}")
            append_run_log(f"Cycle {cycle_index} aborted: {exc}")
            continue
        except Exception as exc:
            print(f"Cycle {cycle_index} crashed unexpectedly: {exc}")
            append_run_log(f"Cycle {cycle_index} crashed unexpectedly: {exc}")
            continue

        print(f"========== Finished cycle {cycle_index} of {loop_count} ==========")
        append_run_log(f"Finished cycle {cycle_index} of {loop_count}.")


if __name__ == "__main__":
    main()



